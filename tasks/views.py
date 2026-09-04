from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Q, Count, Avg
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from permits.models import UserProfile, Department, DepartmentUnit, ModuleRoleAssignment
from permits.module_roles import module_role
from system_admin.models import SystemSetting
from .models import Task, TaskAssignment, TaskUpdate
from .forms import (
    TaskCreateForm,
    TaskUpdateForm,
    TaskManagementNoteForm,
    TaskReturnForm,
    TaskReassignForm,
)
from .access import (
    EXECUTIVE_TASK_ROLES,
    executive_assignee_queryset,
    executive_department_codes,
)


# --------------------------------------------------
# Helpers
# --------------------------------------------------
def _get_system_setting():
    setting, created = SystemSetting.objects.get_or_create(id=1)
    return setting

def _get_profile(user):
    profile, created = UserProfile.objects.get_or_create(user=user)
    return profile

def _get_user_role(user):
    """
    Return one standardized role code for Task permissions.
    """
    if user.is_superuser:
        return "ADMIN"

    assigned_role = module_role(
        user,
        ModuleRoleAssignment.Module.TASK,
        priority=("DIRECTOR", "ASSISTANT_DIRECTOR", "HEAD_OF_UNIT"),
    )
    if assigned_role:
        return assigned_role

    profile = _get_profile(user)

    approval_role = getattr(
        profile,
        "approval_role",
        None,
    )

    approval_role_code = getattr(
        approval_role,
        "code",
        None,
    )

    if approval_role_code:
        role_code = (
            approval_role_code
            .strip()
            .upper()
        )

        if role_code == "SYSTEM_ADMIN":
            return "ADMIN"

        return role_code

    profile_role = getattr(
        profile,
        "role",
        None,
    )

    if profile_role:
        role_code = (
            profile_role
            .strip()
            .upper()
        )

        if role_code == "SYSTEM_ADMIN":
            return "ADMIN"

        return role_code

    groups = {
        group_name.strip().upper()
        for group_name in user.groups.values_list(
            "name",
            flat=True,
        )
        if group_name
    }

    group_priority = [
        "SYSTEM_ADMIN",
        "ADMIN",
        "DIRECTOR",
        "ASSISTANT_DIRECTOR",
        "HEAD_OF_UNIT",
        "DIVISION_BUDGET_OFFICER",
        "ACCOUNTANT",
        "REQUESTER",
    ]

    for group_name in group_priority:
        if group_name in groups:
            if group_name == "SYSTEM_ADMIN":
                return "ADMIN"

            return group_name

    return "STAFF"


def _get_role_reason_label(user):
    role_label_map = {
        "ADMIN": "Admin",
        "DIRECTOR": "Director",
        "ADRD": "ADRD",
        "ADSTI": "ADSTI",
        "HEAD_OF_UNIT": "Head of Unit",
        "PERMANENT_SECRETARY": "Permanent Secretary",
        "DPS_HES": "Deputy Permanent Secretary - Higher Education and Science",
        "DPS_BE": "Deputy Permanent Secretary - Basic Education",
        "COMMISSIONER_EDUCATION": "Commissioner of Education",
    }

    return role_label_map.get(_get_user_role(user), "Manager")


def _is_admin(user):
    return user.is_superuser or _get_user_role(user) == "ADMIN"


def _is_director_level(user):
    return (
        user.is_superuser
        or _get_user_role(user) in ["DIRECTOR", "ADRD", "ADSTI"]
        or _get_user_role(user) in EXECUTIVE_TASK_ROLES
    )


def _is_head_of_unit(user):
    return user.is_superuser or _get_user_role(user) == "HEAD_OF_UNIT"


def _can_create_task(user):
    return _get_user_role(user) in [
        "ADMIN",
        "DIRECTOR",
        "ASSISTANT_DIRECTOR",
        "HEAD_OF_UNIT",
    ] or _get_user_role(user) in EXECUTIVE_TASK_ROLES


def _is_dsti_department_wide_assistant_director(user):
    """
    ADRD and ADSTI are DSTI-wide Assistant Directors and may assign
    tasks across every unit in DSTI.
    """
    profile = _get_profile(user)

    return (
        _get_user_role(user) == "ASSISTANT_DIRECTOR"
        and profile.department
        and profile.department.code == "DSTI"
        and user.username.strip().lower() in [
            "adsti",
            "adrd",
        ]
    )


def _get_assistant_director_task_staff_queryset(user):
    """
    Return staff an Assistant Director may receive task assignments.

    - adsti / adrd: every active user in DSTI.
    - Other Assistant Directors: active users in their exact unit.
    """
    profile = _get_profile(user)

    users = User.objects.filter(
        is_active=True,
    ).select_related(
        "profile",
        "profile__department",
        "profile__department_unit",
    )

    if (
        _get_user_role(user) != "ASSISTANT_DIRECTOR"
        or not profile.department_id
    ):
        return users.none()

    users = users.filter(
        profile__department_id=profile.department_id
    )

    if _is_dsti_department_wide_assistant_director(user):
        return users.order_by(
            "first_name",
            "last_name",
            "username",
        )

    if not profile.department_unit_id:
        return users.none()

    return users.filter(
        profile__department_unit_id=(
            profile.department_unit_id
        )
    ).order_by(
        "first_name",
        "last_name",
        "username",
    )


def _configure_assistant_director_task_form(form, user):
    """
    Restrict TaskCreateForm choices for Assistant Directors.

    Server-side validation in create_task remains authoritative; these
    queryset changes also give the user the correct dropdown choices.
    """
    if _get_user_role(user) != "ASSISTANT_DIRECTOR":
        return form

    profile = _get_profile(user)

    if "department" in form.fields:
        form.fields["department"].queryset = (
            Department.objects.filter(
                pk=profile.department_id
            )
            if profile.department_id
            else Department.objects.none()
        )
        form.fields["department"].initial = (
            profile.department_id
        )
        form.fields["department"].disabled = True

    if "department_unit" in form.fields:
        if _is_dsti_department_wide_assistant_director(user):
            form.fields["department_unit"].queryset = (
                DepartmentUnit.objects.filter(
                    department_id=profile.department_id,
                    is_active=True,
                )
            )
        elif profile.department_unit_id:
            form.fields["department_unit"].queryset = (
                DepartmentUnit.objects.filter(
                    pk=profile.department_unit_id,
                    department_id=profile.department_id,
                    is_active=True,
                )
            )
            form.fields["department_unit"].initial = (
                profile.department_unit_id
            )
            form.fields["department_unit"].disabled = True
        else:
            form.fields["department_unit"].queryset = (
                DepartmentUnit.objects.none()
            )

    if "assigned_users" in form.fields:
        form.fields["assigned_users"].queryset = (
            _get_assistant_director_task_staff_queryset(user)
        )

    return form


def _can_view_task(user, task):
    if user.is_superuser:
        return True

    profile = _get_profile(user)
    role = _get_user_role(user)

    # System Administrator can view every task.
    if role == "ADMIN":
        return True

    # A user can always view a task they created.
    if task.created_by == user:
        return True

    # A user can always view a task directly assigned to them.
    if task.assignments.filter(assigned_to=user).exists():
        return True

    user_department_id = profile.department_id
    user_department_unit_id = profile.department_unit_id

    task_department_id = task.department_id
    task_department_unit_id = task.department_unit_id

    if role in EXECUTIVE_TASK_ROLES:
        department_codes = executive_department_codes(role)
        if department_codes is None:
            return True
        return bool(
            task.department
            and task.department.code in department_codes
        ) or task.assignments.filter(
            assigned_to__profile__department__code__in=department_codes
        ).exists()

    # Director: access tasks belonging to the Director's department.
    if role == "DIRECTOR":
        if user_department_id and task_department_id:
            return task_department_id == user_department_id

        # Backward compatibility for older tasks.
        return bool(
            profile.unit_name
            and task.unit_name
            and task.unit_name == profile.unit_name
        )

    # Assistant Director scope.
    if role == "ASSISTANT_DIRECTOR":
        if not user_department_id:
            return False

        if _is_dsti_department_wide_assistant_director(user):
            return (
                task_department_id == user_department_id
                or task.assignments.filter(
                    assigned_to__profile__department_id=(
                        user_department_id
                    )
                ).exists()
            )

        if not user_department_unit_id:
            return False

        return (
            (
                task_department_id == user_department_id
                and task_department_unit_id
                == user_department_unit_id
            )
            or task.assignments.filter(
                assigned_to__profile__department_id=(
                    user_department_id
                ),
                assigned_to__profile__department_unit_id=(
                    user_department_unit_id
                ),
            ).exists()
        )

    # Head of Unit:
    # access tasks belonging to their own department unit.
    if role == "HEAD_OF_UNIT":
        if user_department_unit_id and task_department_unit_id:
            return task_department_unit_id == user_department_unit_id

        # If the department has no units, compare departments.
        if user_department_id and task_department_id:
            return task_department_id == user_department_id

        # Backward compatibility for older tasks.
        return bool(
            profile.unit_name
            and task.unit_name
            and task.unit_name == profile.unit_name
        )

    # Ordinary staff reach this point only when the task was neither
    # created by them nor directly assigned to them.
    return False


def _can_manage_task(user, task):
    if task.status == "COMPLETED" or (task.progress_percent or 0) >= 100:
        return False

    if user.is_superuser:
        return True

    profile = _get_profile(user)

    role = _get_user_role(user)

    if role == "ADMIN":
        return True

    if role in ["DIRECTOR", "ADRD", "ADSTI"]:
        return True

    if role == "HEAD_OF_UNIT":
        return task.unit_name == profile.unit_name

    if task.created_by == user:
        return True

    return False


def _can_update_task(user, task):
    if task.status in ["ON_HOLD", "CANCELLED", "RETURNED", "COMPLETED"]:
        return False

    if (task.progress_percent or 0) >= 100:
        return False

    if _can_manage_task(user, task):
        return True

    return task.assignments.filter(
        assigned_to=user,
        status__in=["ASSIGNED", "ACCEPTED", "IN_PROGRESS"]
    ).exists()


def _get_filtered_assignments(request):
    status = request.GET.get("status", "").strip()
    employee = request.GET.get("employee", "").strip()
    assignment_status = request.GET.get("assignment_status", "").strip()
    task_status = request.GET.get("task_status", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    created_by = request.GET.get("created_by", "").strip()
    scope = request.GET.get("scope", "").strip()

    now = timezone.now()
    profile = _get_profile(request.user)
    role = _get_user_role(request.user)

    base_assignments = TaskAssignment.objects.select_related(
        "task",
        "task__created_by",
        "task__department",
        "task__department_unit",
        "assigned_by",
        "assigned_to",
        "assigned_to__profile",
        "assigned_to__profile__department",
        "assigned_to__profile__department_unit",
    )

    

    # --------------------------------------------------
    # SYSTEM ADMIN / ADMIN
    # Can view all task assignments.
    # --------------------------------------------------
    # Personal task view used by the "My Tasks" navigation button.
    if scope == "mine":
        assignments = base_assignments.filter(
            assigned_to=request.user
        )

        employee_choices = User.objects.filter(
            pk=request.user.pk
        )

    # System Administrator / Admin can view all assignments.
    elif role == "ADMIN":
        assignments = base_assignments.all()

        employee_choices = User.objects.filter(
            is_active=True
        ).select_related(
            "profile",
            "profile__department",
            "profile__department_unit",
        ).order_by(
            "first_name",
            "last_name",
            "username"
        )

    elif role in EXECUTIVE_TASK_ROLES:
        department_codes = executive_department_codes(role)
        if department_codes is None:
            assignments = base_assignments.all()
            employee_choices = executive_assignee_queryset(role)
        else:
            assignments = base_assignments.filter(
                Q(task__department__code__in=department_codes)
                | Q(assigned_to__profile__department__code__in=department_codes)
                | Q(task__created_by=request.user)
                | Q(assigned_to=request.user)
            ).distinct()
            employee_choices = executive_assignee_queryset(role)

    # --------------------------------------------------
    # DIRECTOR
    # Can view assignments from the Director's own department.
    # --------------------------------------------------
    elif role == "DIRECTOR":
        if profile.department_id:
            department_unit_codes = list(
                DepartmentUnit.objects.filter(
                    department_id=profile.department_id,
                    is_active=True,
                ).values_list("code", flat=True)
            )

            legacy_scope_codes = department_unit_codes[:]

            if profile.department:
                legacy_scope_codes.append(profile.department.code)

            assignments = base_assignments.filter(
                Q(task__department_id=profile.department_id)
                | Q(
                    task__department__isnull=True,
                    task__unit_name__in=legacy_scope_codes,
                )
            ).distinct()

            employee_choices = User.objects.filter(
                is_active=True,
                profile__department_id=profile.department_id
            ).select_related(
                "profile",
                "profile__department",
                "profile__department_unit",
            ).order_by(
                "first_name",
                "last_name",
                "username"
            )

        elif profile.unit_name:
            # Backward compatibility for older tasks
            # that do not yet have a Department record.
            assignments = base_assignments.filter(
                task__department__isnull=True,
                task__unit_name=profile.unit_name
            )

            employee_choices = User.objects.filter(
                is_active=True,
                profile__unit_name=profile.unit_name
            ).select_related(
                "profile"
            ).order_by(
                "first_name",
                "last_name",
                "username"
            )

        else:
            assignments = base_assignments.none()
            employee_choices = User.objects.none()

    # --------------------------------------------------
    # ASSISTANT DIRECTOR
    # Unit-specific AD: exact unit tasks and every assignment made to
    # a member of that unit, including assignments made by a Director.
    # DSTI adsti/adrd: department-wide DSTI scope.
    # --------------------------------------------------
    elif role == "ASSISTANT_DIRECTOR":
        if (
            profile.department_id
            and _is_dsti_department_wide_assistant_director(
                request.user
            )
        ):
            assignments = base_assignments.filter(
                Q(
                    task__department_id=profile.department_id
                )
                | Q(
                    assigned_to__profile__department_id=(
                        profile.department_id
                    )
                )
                | Q(task__created_by=request.user)
                | Q(assigned_to=request.user)
            ).distinct()

            employee_choices = User.objects.filter(
                is_active=True,
                profile__department_id=profile.department_id,
            ).select_related(
                "profile",
                "profile__department",
                "profile__department_unit",
            ).order_by(
                "first_name",
                "last_name",
                "username",
            )

        elif (
            profile.department_id
            and profile.department_unit_id
        ):
            assignments = base_assignments.filter(
                Q(
                    task__department_id=profile.department_id,
                    task__department_unit_id=(
                        profile.department_unit_id
                    ),
                )
                | Q(
                    assigned_to__profile__department_id=(
                        profile.department_id
                    ),
                    assigned_to__profile__department_unit_id=(
                        profile.department_unit_id
                    ),
                )
                | Q(task__created_by=request.user)
                | Q(assigned_to=request.user)
            ).distinct()

            employee_choices = User.objects.filter(
                is_active=True,
                profile__department_id=profile.department_id,
                profile__department_unit_id=(
                    profile.department_unit_id
                ),
            ).select_related(
                "profile",
                "profile__department",
                "profile__department_unit",
            ).order_by(
                "first_name",
                "last_name",
                "username",
            )

        else:
            assignments = base_assignments.filter(
                Q(task__created_by=request.user)
                | Q(assigned_to=request.user)
            ).distinct()
            employee_choices = User.objects.filter(
                pk=request.user.pk
            )

       # --------------------------------------------------
    # HEAD OF UNIT
    # Can view assignments involving their department unit.
    # For departments without units, use the department.
    # --------------------------------------------------
    elif role == "HEAD_OF_UNIT":
        if profile.department_unit_id:
            unit_id = profile.department_unit_id

            legacy_unit_codes = []

            if profile.department_unit:
                legacy_unit_codes.append(
                    profile.department_unit.code
                )

            if profile.unit_name:
                legacy_unit_codes.append(
                    profile.unit_name
                )

            assignments = base_assignments.filter(
                Q(task__department_unit_id=unit_id)
                |
                Q(
                    assigned_to__profile__department_unit_id=unit_id
                )
                |
                Q(task__created_by=request.user)
                |
                Q(assigned_to=request.user)
                |
                Q(
                    task__department__isnull=True,
                    task__unit_name__in=legacy_unit_codes,
                )
            ).distinct()

            employee_choices = User.objects.filter(
                is_active=True,
                profile__department_unit_id=unit_id
            ).select_related(
                "profile",
                "profile__department",
                "profile__department_unit",
            ).order_by(
                "first_name",
                "last_name",
                "username"
            )

        elif profile.department_id:
            assignments = base_assignments.filter(
                Q(task__department_id=profile.department_id)
                |
                Q(
                    assigned_to__profile__department_id=profile.department_id
                )
                |
                Q(task__created_by=request.user)
                |
                Q(assigned_to=request.user)
            ).distinct()

            employee_choices = User.objects.filter(
                is_active=True,
                profile__department_id=profile.department_id
            ).select_related(
                "profile",
                "profile__department",
                "profile__department_unit",
            ).order_by(
                "first_name",
                "last_name",
                "username"
            )

        elif profile.unit_name:
            assignments = base_assignments.filter(
                Q(
                    task__department__isnull=True,
                    task__unit_name=profile.unit_name,
                )
                |
                Q(
                    assigned_to__profile__unit_name=profile.unit_name
                )
                |
                Q(task__created_by=request.user)
                |
                Q(assigned_to=request.user)
            ).distinct()

            employee_choices = User.objects.filter(
                is_active=True,
                profile__unit_name=profile.unit_name
            ).select_related(
                "profile"
            ).order_by(
                "first_name",
                "last_name",
                "username"
            )

        else:
            assignments = base_assignments.filter(
                Q(task__created_by=request.user)
                |
                Q(assigned_to=request.user)
            ).distinct()

            employee_choices = User.objects.filter(
                pk=request.user.pk
            )

    # --------------------------------------------------
    # ORDINARY STAFF
    # Can see only assignments made directly to them.
    # --------------------------------------------------
    else:
        assignments = base_assignments.filter(
            assigned_to=request.user
        )

        employee_choices = User.objects.filter(
            pk=request.user.pk
        )

    assignments = assignments.order_by("-assigned_at")

    if status == "OVERDUE":
        assignments = assignments.filter(
            task__due_date__lt=now
        ).exclude(
            task__status__in=["COMPLETED", "CANCELLED"]
        ).exclude(
            status="COMPLETED"
        )

    elif status == "ACCEPTED":
        # Acceptance is historical: once an assignee accepts an assignment,
        # keep it in this filter even after work starts or is completed.
        assignments = assignments.filter(
            accepted_at__isnull=False
        )

    elif status == "NOT_ACCEPTED":
        # Assignments for which the assignee has not recorded acceptance.
        # Completed/cancelled records are excluded because they are no
        # longer actionable outstanding assignments.
        assignments = assignments.filter(
            accepted_at__isnull=True
        ).exclude(
            status__in=["COMPLETED", "CANCELLED"]
        ).exclude(
            task__status__in=["COMPLETED", "CANCELLED"]
        )

    elif status == "IN_PROGRESS":
        # The dashboard button represents the overall Task Status.
        # Include every assignment row belonging to an in-progress
        # task, even when an individual assignment is still ASSIGNED
        # or ACCEPTED. Assignment-specific filtering remains available
        # through the separate Assignment Status dropdown below.
        assignments = assignments.filter(
            task__status="IN_PROGRESS"
        )

    elif status:
        assignments = assignments.filter(status=status)

    if employee:
        assignments = assignments.filter(
            assigned_to_id=employee
        )

    if assignment_status:
        assignments = assignments.filter(
            status=assignment_status
        )

    if task_status:
        assignments = assignments.filter(
            task__status=task_status
        )

    if date_from:
        assignments = assignments.filter(
            assigned_at__date__gte=date_from
        )

    if date_to:
        assignments = assignments.filter(
            assigned_at__date__lte=date_to
        )

    if created_by == "me":
        assignments = assignments.filter(
            task__created_by=request.user
        )

    return {
        "assignments": assignments,
        "employee_choices": employee_choices,
        "selected_status": status,
        "selected_employee": employee,
        "selected_assignment_status": assignment_status,
        "selected_task_status": task_status,
        "selected_date_from": date_from,
        "selected_date_to": date_to,
        "selected_created_by": created_by,
        "selected_scope": scope,
        "profile": profile,
    }


def _safe_full_name(user):
    if not user:
        return "-"
    return (user.get_full_name() or user.username or "-").strip()


def _date_diff_days_late(due_dt, end_dt):
    if not due_dt or not end_dt:
        return 0
    due_local = timezone.localtime(due_dt) if timezone.is_aware(due_dt) else due_dt
    end_local = timezone.localtime(end_dt) if timezone.is_aware(end_dt) else end_dt
    return max((end_local.date() - due_local.date()).days, 0)


def _date_diff_days_remaining(now_dt, due_dt):
    if not due_dt:
        return None
    now_local = timezone.localtime(now_dt) if timezone.is_aware(now_dt) else now_dt
    due_local = timezone.localtime(due_dt) if timezone.is_aware(due_dt) else due_dt
    return (due_local.date() - now_local.date()).days


def _is_open_task(task):
    return task.status not in ["COMPLETED", "CANCELLED", "RETURNED"] and (task.progress_percent or 0) < 100


def _latest_assignment_for_task(task):
    try:
        return task.assignments.all().order_by("-assigned_at").first()
    except Exception:
        return None

@login_required
def load_department_units(request):
    department_id = request.GET.get("department")
    profile = _get_profile(request.user)
    role = _get_user_role(request.user)

    units = DepartmentUnit.objects.none()

    if role in EXECUTIVE_TASK_ROLES:
        allowed_codes = executive_department_codes(role)
        department = Department.objects.filter(pk=department_id, is_active=True)
        if allowed_codes is not None:
            department = department.filter(code__in=allowed_codes)
        if not department.exists():
            return JsonResponse({"units": []})

    if role == "ASSISTANT_DIRECTOR":
        if (
            not profile.department_id
            or str(profile.department_id) != str(department_id)
        ):
            return JsonResponse({"units": []})

        if (
            not _is_dsti_department_wide_assistant_director(
                request.user
            )
            and profile.department_unit_id
        ):
            units = DepartmentUnit.objects.filter(
                pk=profile.department_unit_id,
                department_id=profile.department_id,
                is_active=True,
            ).order_by("code")

            return JsonResponse({
                "units": [
                    {
                        "id": unit.id,
                        "name": f"{unit.code} - {unit.name}",
                    }
                    for unit in units
                ]
            })

    if department_id:
        units = DepartmentUnit.objects.filter(
            department_id=department_id,
            is_active=True
        ).order_by("code")

    data = [
        {
            "id": unit.id,
            "name": f"{unit.code} - {unit.name}",
        }
        for unit in units
    ]

    return JsonResponse({
        "units": data
    })


@login_required
def load_department_staff(request):
    department_id = request.GET.get("department")
    department_unit_id = request.GET.get("department_unit")
    assignee_scope = request.GET.get("assignee_scope", "UNIT")

    profile = _get_profile(request.user)
    role = _get_user_role(request.user)

    if role in EXECUTIVE_TASK_ROLES:
        users = executive_assignee_queryset(role).exclude(pk=request.user.pk)
        allowed_codes = executive_department_codes(role)
        if department_id:
            selected_department = Department.objects.filter(pk=department_id).first()
            if (
                not selected_department
                or (
                    allowed_codes is not None
                    and selected_department.code not in allowed_codes
                )
            ):
                return JsonResponse({"staff": []})
            users = users.filter(
                Q(profile__department_id=department_id)
                | Q(profile__department__isnull=True)
            )
        if assignee_scope == "UNIT" and department_unit_id:
            users = users.filter(profile__department_unit_id=department_unit_id)
        return JsonResponse({
            "staff": [
                {
                    "id": user.id,
                    "name": user.get_full_name().strip() or user.username,
                }
                for user in users
            ]
        })

    if not department_id:
        return JsonResponse({
            "staff": []
        })

    if role == "ASSISTANT_DIRECTOR":
        if (
            not profile.department_id
            or str(profile.department_id) != str(department_id)
        ):
            return JsonResponse({"staff": []})

        users = _get_assistant_director_task_staff_queryset(
            request.user
        )

        if not _is_dsti_department_wide_assistant_director(
            request.user
        ):
            department_unit_id = str(
                profile.department_unit_id or ""
            )
            assignee_scope = "UNIT"

    else:
        users = User.objects.filter(
            is_active=True,
            profile__department_id=department_id
        ).select_related(
            "profile",
            "profile__department",
            "profile__department_unit",
        )

    # UNIT means only staff from the selected Department Unit.
    if assignee_scope == "UNIT" and department_unit_id:
        users = users.filter(
            profile__department_unit_id=department_unit_id
        )

    # OTHER means all eligible staff in the same Department.
    # The Department Unit filter is intentionally ignored.
    users = users.order_by(
        "first_name",
        "last_name",
        "username"
    )

    staff_data = []

    for user in users:
        profile = getattr(user, "profile", None)
        full_name = user.get_full_name().strip() or user.username

        department_code = (
            profile.department.code
            if profile and profile.department
            else "No Department"
        )

        unit_code = (
            profile.department_unit.code
            if profile and profile.department_unit
            else ""
        )

        if assignee_scope == "OTHER" and unit_code:
            display_name = f"{full_name} — {unit_code}"
        else:
            display_name = full_name

        staff_data.append({
            "id": user.id,
            "name": display_name,
        })

    return JsonResponse({
        "staff": staff_data
    })

def _get_visible_task_scope(
    user,
    strict_department=False,
):
    """
    Return department-aware Task and TaskAssignment querysets.

    All current Tasks already have:
    - department
    - department_unit

    Therefore, visibility is controlled using those relationships rather
    than shared legacy unit codes such as CCU.

    Rules:
    - ADMIN: all tasks and assignments.
    - DIRECTOR: own department.
    - ASSISTANT_DIRECTOR:
        * DSTI adsti/adrd -> all DSTI tasks and assignments;
        * unit-specific AD -> exact unit tasks and assignments to members
          of that unit, including assignments made by the Director.
    - HEAD_OF_UNIT: own unit.
    - Ordinary staff: only tasks they created or received directly.
    - strict_department=True prevents Directors and Assistant Directors
      from including cross-department tasks through personal assignment.
    """
    profile = _get_profile(user)
    role = _get_user_role(user)

    tasks = (
        Task.objects
        .select_related(
            "created_by",
            "department",
            "department_unit",
        )
        .prefetch_related(
            "assignments",
            "assignments__assigned_to",
            "assignments__assigned_to__profile",
        )
    )

    assignments = (
        TaskAssignment.objects
        .select_related(
            "task",
            "task__created_by",
            "task__department",
            "task__department_unit",
            "assigned_by",
            "assigned_to",
            "assigned_to__profile",
            "assigned_to__profile__department",
            "assigned_to__profile__department_unit",
        )
    )

    # --------------------------------------------------
    # SYSTEM ADMIN
    # --------------------------------------------------
    if role == "ADMIN":
        return (
            tasks.all(),
            assignments.all(),
        )

    if role in EXECUTIVE_TASK_ROLES:
        department_codes = executive_department_codes(role)
        if department_codes is None:
            return tasks.all(), assignments.all()
        return (
            tasks.filter(
                Q(department__code__in=department_codes)
                | Q(assignments__assigned_to__profile__department__code__in=department_codes)
                | Q(created_by=user)
                | Q(assignments__assigned_to=user)
            ).distinct(),
            assignments.filter(
                Q(task__department__code__in=department_codes)
                | Q(assigned_to__profile__department__code__in=department_codes)
                | Q(task__created_by=user)
                | Q(assigned_to=user)
            ).distinct(),
        )

    direct_task_access = (
        Q(created_by=user)
        | Q(assignments__assigned_to=user)
    )

    direct_assignment_access = (
        Q(task__created_by=user)
        | Q(assigned_to=user)
    )

    # --------------------------------------------------
    # DIRECTOR
    # Sees only Tasks belonging to their Department.
    # --------------------------------------------------
    if role == "DIRECTOR":
        if not profile.department_id:
            if strict_department:
                return (
                    tasks.none(),
                    assignments.none(),
                )

            return (
                tasks.filter(
                    direct_task_access
                ).distinct(),
                assignments.filter(
                    direct_assignment_access
                ).distinct(),
            )

        task_scope = Q(
            department_id=profile.department_id
        )

        assignment_scope = Q(
            task__department_id=profile.department_id
        )

        if not strict_department:
            task_scope |= direct_task_access
            assignment_scope |= direct_assignment_access

        return (
            tasks.filter(
                task_scope
            ).distinct(),
            assignments.filter(
                assignment_scope
            ).distinct(),
        )

    # --------------------------------------------------
    # ASSISTANT DIRECTOR
    #
    # DSTI-wide Assistant Directors adsti/adrd see DSTI.
    # Other Assistant Directors see their exact unit, including tasks
    # assigned to members of that unit by the Department Director.
    # --------------------------------------------------
    if role == "ASSISTANT_DIRECTOR":
        if (
            profile.department_id
            and _is_dsti_department_wide_assistant_director(user)
        ):
            task_scope = (
                Q(department_id=profile.department_id)
                | Q(
                    assignments__assigned_to__profile__department_id=(
                        profile.department_id
                    )
                )
            )

            assignment_scope = (
                Q(task__department_id=profile.department_id)
                | Q(
                    assigned_to__profile__department_id=(
                        profile.department_id
                    )
                )
            )

        elif (
            profile.department_id
            and profile.department_unit_id
        ):
            task_scope = Q(
                department_id=profile.department_id,
                department_unit_id=profile.department_unit_id,
            ) | Q(
                assignments__assigned_to__profile__department_id=(
                    profile.department_id
                ),
                assignments__assigned_to__profile__department_unit_id=(
                    profile.department_unit_id
                ),
            )

            assignment_scope = Q(
                task__department_id=profile.department_id,
                task__department_unit_id=(
                    profile.department_unit_id
                ),
            ) | Q(
                assigned_to__profile__department_id=(
                    profile.department_id
                ),
                assigned_to__profile__department_unit_id=(
                    profile.department_unit_id
                ),
            )

        else:
            return (
                tasks.filter(
                    direct_task_access
                ).distinct(),
                assignments.filter(
                    direct_assignment_access
                ).distinct(),
            )

        if not strict_department:
            task_scope |= direct_task_access
            assignment_scope |= direct_assignment_access

        return (
            tasks.filter(
                task_scope
            ).distinct(),
            assignments.filter(
                assignment_scope
            ).distinct(),
        )

        # --------------------------------------------------
    # HEAD OF UNIT
    # Sees only tasks belonging to the exact Department Unit.
    # --------------------------------------------------
    if role == "HEAD_OF_UNIT":
        if (
            profile.department_id
            and profile.department_unit_id
        ):
            visible_tasks = tasks.filter(
                Q(
                    department_id=profile.department_id,
                    department_unit_id=profile.department_unit_id,
                )
                | Q(
                    assignments__assigned_to__profile__department_id=(
                        profile.department_id
                    ),
                    assignments__assigned_to__profile__department_unit_id=(
                        profile.department_unit_id
                    ),
                )
                | Q(created_by=user)
                | Q(assignments__assigned_to=user)
            ).distinct()

            visible_assignments = assignments.filter(
                Q(
                    task__department_id=profile.department_id,
                    task__department_unit_id=(
                        profile.department_unit_id
                    ),
                )
                | Q(
                    assigned_to__profile__department_id=(
                        profile.department_id
                    ),
                    assigned_to__profile__department_unit_id=(
                        profile.department_unit_id
                    ),
                )
                | Q(task__created_by=user)
                | Q(assigned_to=user)
            ).distinct()

            return (
                visible_tasks,
                visible_assignments,
            )

        # Head of Unit in a department without department units.
        if profile.department_id:
            visible_tasks = tasks.filter(
                Q(
                    department_id=profile.department_id
                )
                | direct_task_access
            ).distinct()

            visible_assignments = assignments.filter(
                Q(
                    task__department_id=profile.department_id
                )
                | direct_assignment_access
            ).distinct()

            return (
                visible_tasks,
                visible_assignments,
            )

        return (
            tasks.filter(
                direct_task_access
            ).distinct(),
            assignments.filter(
                direct_assignment_access
            ).distinct(),
        )

    # --------------------------------------------------
    # ORDINARY STAFF
    # --------------------------------------------------
    # Personal access only. Department membership alone never grants
    # visibility into another employee's tasks.
    return (
        tasks.filter(
            direct_task_access
        ).distinct(),
        assignments.filter(
            direct_assignment_access
        ).distinct(),
    )

# --------------------------------------------------
# Dashboard
# --------------------------------------------------
@login_required
def task_dashboard(request):
    setting = _get_system_setting()

    if not setting.open_task_enabled:
        messages.error(
            request,
            "Open Task module is currently disabled by the System Administrator."
        )
        return redirect("system_home")

    profile = _get_profile(request.user)
    role = _get_user_role(request.user)
    now = timezone.now()

    visible_tasks, visible_assignments = _get_visible_task_scope(request.user)

    show_director_analytics = role in [
        "ADMIN",
        "DIRECTOR",
        "ASSISTANT_DIRECTOR",
        "HEAD_OF_UNIT",
    ] or role in EXECUTIVE_TASK_ROLES

    overdue_tasks = visible_tasks.filter(
        due_date__lt=now
    ).exclude(
        status__in=["COMPLETED", "CANCELLED"]
    )

    created_tasks = visible_tasks.filter(created_by=request.user)
    pending_tasks = visible_tasks.filter(status="PENDING")
    on_hold_tasks = visible_tasks.filter(status="ON_HOLD")
    cancelled_tasks = visible_tasks.filter(status="CANCELLED")
    returned_tasks = visible_tasks.filter(status="RETURNED")
    completed_tasks = visible_tasks.filter(status="COMPLETED")

    total_tasks = visible_tasks.count()
    completion_rate = (
        round((completed_tasks.count() / total_tasks) * 100, 1)
        if total_tasks > 0
        else 0
    )

    context = {
        "profile": profile,
        "role": role,
        "my_task_count": total_tasks,
        "my_assignment_count": visible_assignments.count(),
        "created_task_count": created_tasks.count(),
        "overdue_task_count": overdue_tasks.count(),
        "pending_task_count": pending_tasks.count(),
        "on_hold_task_count": on_hold_tasks.count(),
        "cancelled_task_count": cancelled_tasks.count(),
        "returned_task_count": returned_tasks.count(),
        "completed_task_count": completed_tasks.count(),
        "completion_rate": completion_rate,
        "can_create_task": (
            setting.allow_task_creation
            and _can_create_task(request.user)
        ),
        "recent_assignments": visible_assignments.order_by("-assigned_at")[:10],
        "recent_created_tasks": created_tasks.order_by("-created_at")[:10],
        "show_director_analytics": show_director_analytics,
    }

    if show_director_analytics:
        analytics_tasks, analytics_assignments = (
            _get_visible_task_scope(
                request.user,
                strict_department=True,
            )
        )

        units = (
            analytics_tasks
            .exclude(unit_name__isnull=True)
            .exclude(unit_name="")
            .values_list("unit_name", flat=True)
            .distinct()
            .order_by("unit_name")
        )

        task_counts = (
            analytics_tasks
            .exclude(unit_name__isnull=True)
            .exclude(unit_name="")
            .values("unit_name")
            .annotate(
                total=Count("id"),
                completed=Count(
                    "id",
                    filter=Q(status="COMPLETED")
                ),
            )
        )

        task_map = {
            row["unit_name"]: row
            for row in task_counts
        }

        tasks_per_unit = []
        completion_rate_by_unit = []

        for unit in units:
            row = task_map.get(unit)
            total = row["total"] if row else 0
            completed = row["completed"] if row else 0
            rate = (
                round((completed / total) * 100, 1)
                if total > 0
                else 0
            )

            tasks_per_unit.append({
                "unit_name": unit,
                "total": total,
            })

            completion_rate_by_unit.append({
                "unit_name": unit,
                "total": total,
                "completed": completed,
                "rate": rate,
            })

        top_performers_rows = (
            analytics_assignments
            .values(
                "assigned_to__id",
                "assigned_to__username",
                "assigned_to__first_name",
                "assigned_to__last_name",
                "assigned_to__profile__unit_name",
            )
            .annotate(
                total_assignments=Count("id"),
                completed_assignments=Count(
                    "id",
                    filter=Q(status="COMPLETED")
                ),
                avg_progress=Avg("progress_percent"),
            )
            .order_by(
                "-completed_assignments",
                "-avg_progress",
                "-total_assignments"
            )[:10]
        )

        top_performers = [
            {
                "user_id": row["assigned_to__id"],
                "name": (
                    f"{row['assigned_to__first_name']} "
                    f"{row['assigned_to__last_name']}"
                ).strip() or row["assigned_to__username"],
                "unit_name": (
                    row["assigned_to__profile__unit_name"] or "-"
                ),
                "total_assignments": row["total_assignments"],
                "completed_assignments": row["completed_assignments"],
                "avg_progress": round(
                    row["avg_progress"] or 0,
                    1
                ),
                "completion_rate": round(
                    (
                        row["completed_assignments"]
                        / row["total_assignments"]
                    ) * 100,
                    1
                ) if row["total_assignments"] > 0 else 0,
            }
            for row in top_performers_rows
        ]

        context.update({
            "tasks_per_unit": tasks_per_unit,
            "completion_rate_by_unit": completion_rate_by_unit,
            "top_performers": top_performers,
        })

    return render(
        request,
        "tasks/task_dashboard.html",
        context
    )


# --------------------------------------------------
# My Tasks
# --------------------------------------------------
@login_required
def my_tasks(request):
    data = _get_filtered_assignments(request)

    context = {
        "assignments": data["assignments"],
        "selected_status": data["selected_status"],
        "employee_choices": data["employee_choices"],
        "selected_employee": data["selected_employee"],
        "selected_assignment_status": data["selected_assignment_status"],
        "selected_task_status": data["selected_task_status"],
        "selected_date_from": data["selected_date_from"],
        "selected_date_to": data["selected_date_to"],
        "selected_created_by": data["selected_created_by"],
        "selected_scope": data["selected_scope"],
        "is_filter_manager": (
            request.user.is_superuser
            or data["profile"].role in [
                "ADMIN",
                "DIRECTOR",
                "ADRD",
                "ADSTI",
                "HEAD_OF_UNIT",
            ]
            or _get_user_role(request.user) in EXECUTIVE_TASK_ROLES
        ),
        "profile": data["profile"],
        "allow_task_export": _get_system_setting().allow_task_export,
    }

    return render(
        request,
        "tasks/my_tasks.html",
        context
    )


# --------------------------------------------------
# Export Filtered Tasks to Excel
# --------------------------------------------------
@login_required
def export_tasks_excel(request):
    setting = _get_system_setting()

    if not setting.allow_task_export:
        messages.error(request, "Task export is currently disabled by the System Administrator.")
        return redirect("my_tasks")
    data = _get_filtered_assignments(request)
    assignments = data["assignments"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Tasks"

    headers = [
        "Task Title",
        "Employee Assigned",
        "Employee Unit",
        "Assigned By",
        "Created By",
        "Task Status",
        "Assignment Status",
        "Priority",
        "Assignment Progress %",
        "Overall Task Progress %",
        "Responsible Unit",
        "Start Date",
        "Due Date",
        "Assigned At",
        "Accepted At",
        "Started At",
        "Last Updated At",
        "Completed At",
    ]
    ws.append(headers)

    for item in assignments:
        assigned_user = item.assigned_to
        assigned_profile = getattr(assigned_user, "profile", None)

        ws.append([
            item.task.title,
            assigned_user.get_full_name() or assigned_user.username,
            getattr(assigned_profile, "unit_name", "") or "",
            item.assigned_by.get_full_name() or item.assigned_by.username,
            item.task.created_by.get_full_name() or item.task.created_by.username,
            item.task.get_status_display(),
            item.get_status_display(),
            item.task.get_priority_display(),
            item.progress_percent,
            item.task.progress_percent,
            item.task.unit_name or "",
            item.task.start_date.strftime("%Y-%m-%d %H:%M") if item.task.start_date else "",
            item.task.due_date.strftime("%Y-%m-%d %H:%M") if item.task.due_date else "",
            item.assigned_at.strftime("%Y-%m-%d %H:%M") if item.assigned_at else "",
            item.accepted_at.strftime("%Y-%m-%d %H:%M") if item.accepted_at else "",
            item.started_at.strftime("%Y-%m-%d %H:%M") if item.started_at else "",
            item.last_updated_at.strftime("%Y-%m-%d %H:%M") if item.last_updated_at else "",
            item.completed_at.strftime("%Y-%m-%d %H:%M") if item.completed_at else "",
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                value_length = len(str(cell.value)) if cell.value else 0
                if value_length > max_length:
                    max_length = value_length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="filtered_tasks.xlsx"'

    wb.save(response)
    return response


# --------------------------------------------------
# Create Task
# --------------------------------------------------
@login_required
def create_task(request):
    setting = _get_system_setting()

    if not setting.allow_task_creation:
        messages.error(
            request,
            "Task creation is currently disabled by the System Administrator."
        )
        return redirect("task_dashboard")

    if not _can_create_task(request.user):
        return HttpResponseForbidden(
            "You are not allowed to create tasks."
        )

    profile = _get_profile(request.user)

    # Keep the existing parameters for compatibility
    selected_unit = request.GET.get("unit_name", "").strip()
    assignee_scope = request.GET.get("assignee_scope", "").strip() or "UNIT"

    if request.method == "POST":
        form = TaskCreateForm(
            request.POST,
            request.FILES,
            user=request.user
        )
        form = _configure_assistant_director_task_form(
            form,
            request.user,
        )

        if form.is_valid():

            # Block assignment before saving anything
            if not setting.allow_task_assignment:
                messages.error(
                    request,
                    "Task assignment is currently disabled by the System Administrator."
                )
                return redirect("task_dashboard")

            task = form.save(commit=False)
            task.created_by = request.user
            task.status = "PENDING"
            task.progress_percent = 0

            role = _get_user_role(request.user)
            selected_department = form.cleaned_data.get(
                "department"
            )
            selected_department_unit = form.cleaned_data.get(
                "department_unit"
            )
            assigned_users = form.cleaned_data[
                "assigned_users"
            ]

            if role == "ASSISTANT_DIRECTOR":
                allowed_staff = (
                    _get_assistant_director_task_staff_queryset(
                        request.user
                    )
                )
                allowed_staff_ids = set(
                    allowed_staff.values_list(
                        "id",
                        flat=True,
                    )
                )
                submitted_staff_ids = {
                    user.id
                    for user in assigned_users
                }

                if not profile.department_id:
                    return HttpResponseForbidden(
                        "Your account has no Department assignment."
                    )

                if (
                    not selected_department
                    or selected_department.id
                    != profile.department_id
                ):
                    return HttpResponseForbidden(
                        "Assistant Directors may create tasks only "
                        "within their assigned Department."
                    )

                if _is_dsti_department_wide_assistant_director(
                    request.user
                ):
                    if (
                        selected_department_unit
                        and selected_department_unit.department_id
                        != profile.department_id
                    ):
                        return HttpResponseForbidden(
                            "The selected unit is outside DSTI."
                        )
                else:
                    if (
                        not profile.department_unit_id
                        or not selected_department_unit
                        or selected_department_unit.id
                        != profile.department_unit_id
                    ):
                        return HttpResponseForbidden(
                            "Assistant Directors may create tasks only "
                            "for their assigned Department Unit."
                        )

                if not submitted_staff_ids.issubset(
                    allowed_staff_ids
                ):
                    return HttpResponseForbidden(
                        "One or more selected staff members are outside "
                        "your permitted task-assignment scope."
                    )

            # Save the new department structure
            task.department = selected_department
            task.department_unit = selected_department_unit

            # Preserve the old unit_name field for compatibility
            cleaned_unit_name = (
                form.cleaned_data.get("unit_name") or ""
            ).strip()

            if cleaned_unit_name:
                task.unit_name = cleaned_unit_name
            elif profile.department_unit:
                task.unit_name = profile.department_unit.code[:20]
            elif profile.department:
                task.unit_name = profile.department.code[:20]
            else:
                task.unit_name = profile.unit_name

            attachment = request.FILES.get("attachment")
            if attachment:
                task.attachment = attachment

            with transaction.atomic():
                task.save()

                for user in assigned_users:
                    TaskAssignment.objects.create(
                        task=task,
                        assigned_to=user,
                        assigned_by=request.user,
                        status="ASSIGNED",
                        progress_percent=0,
                        carried_forward_progress=0,
                    )

            task.refresh_from_assignments()

            messages.success(
                request,
                "Task created and assigned successfully."
            )
            return redirect("task_detail", pk=task.pk)

    else:
        form = TaskCreateForm(
            user=request.user,
            selected_unit=selected_unit,
            assignee_scope=assignee_scope,
        )
        form = _configure_assistant_director_task_form(
            form,
            request.user,
        )

    return render(
        request,
        "tasks/create_task.html",
        {
            "form": form,
        }
    )


# --------------------------------------------------
# Task Detail + Updates
# --------------------------------------------------
@login_required
def task_detail(request, pk):
    setting = _get_system_setting()

    task = get_object_or_404(
        Task.objects.select_related("created_by", "management_note_by").prefetch_related(
            "assignments__assigned_to",
            "assignments__assigned_to__profile",
            "assignments__assigned_by",
            "updates__updated_by",
            "updates__assignment",
        ),
        pk=pk
    )

    if not _can_view_task(request.user, task):
        return HttpResponseForbidden("You are not allowed to view this task.")

    my_assignment = task.assignments.filter(assigned_to=request.user).order_by("-assigned_at").first()
    is_completed_locked = task.status == "COMPLETED" or (task.progress_percent or 0) >= 100
    can_manage = _can_manage_task(request.user, task)
    can_delete = (task.created_by == request.user) and not is_completed_locked
    can_update = _can_update_task(request.user, task) and not is_completed_locked

    role_reason_label = _get_role_reason_label(request.user)
    saved_role_reason_label = task.management_note_role or role_reason_label

    if request.method == "POST":
        if not setting.allow_task_progress_update:
            messages.error(request, "Task progress update is currently disabled by the System Administrator.")
            return redirect("task_detail", pk=task.pk)

        if not can_update:
            return HttpResponseForbidden("You are not allowed to update this task.")

        update_form = TaskUpdateForm(request.POST, request.FILES)

        if update_form.is_valid():
            new_progress = update_form.cleaned_data.get("progress_percent") or 0
            new_progress = max(0, min(int(new_progress), 100))

            if new_progress >= 100 and not setting.allow_task_completion:
                messages.error(request, "Task completion is currently disabled by the System Administrator.")
                return redirect("task_detail", pk=task.pk)

            update = update_form.save(commit=False)
            update.task = task
            update.updated_by = request.user

            target_assignment = None

            if my_assignment:
                target_assignment = my_assignment
            elif _can_manage_task(request.user, task):
                target_assignment = task.assignments.filter(
                    status__in=["ASSIGNED", "ACCEPTED", "IN_PROGRESS"]
                ).order_by("-assigned_at").first()

                if not target_assignment:
                    target_assignment = task.assignments.order_by("-assigned_at").first()

            if target_assignment:
                update.assignment = target_assignment

            update.save()

            if target_assignment and "progress_percent" in update_form.cleaned_data:
                now = timezone.now()
                target_assignment.progress_percent = new_progress
                target_assignment.last_updated_at = now

                if new_progress > 0 and not target_assignment.started_at:
                    target_assignment.started_at = now

                baseline = getattr(target_assignment, "carried_forward_progress", 0) or 0
                total_progress = baseline + new_progress

                if total_progress >= 100:
                    target_assignment.status = "COMPLETED"
                    target_assignment.completed_at = now
                    target_assignment.progress_percent = max(0, 100 - baseline)
                elif new_progress > 0:
                    target_assignment.status = "IN_PROGRESS"
                    target_assignment.completed_at = None
                else:
                    if target_assignment.status == "COMPLETED":
                        target_assignment.status = "ASSIGNED"
                    target_assignment.completed_at = None

                target_assignment.save(update_fields=[
                    "progress_percent",
                    "status",
                    "started_at",
                    "last_updated_at",
                    "completed_at",
                    "updated_at",
                ])

                task.refresh_from_assignments()

            messages.success(request, "Task update added successfully.")
            return redirect("task_detail", pk=task.pk)
    else:
        initial_progress = 0
        if my_assignment:
            initial_progress = my_assignment.progress_percent or 0

        update_form = TaskUpdateForm(initial={
            "progress_percent": initial_progress
        })

    context = {
        "task": task,
        "assignments": task.assignments.all().order_by("-assigned_at"),
        "updates": task.updates.all().order_by("-created_at"),
        "my_assignment": my_assignment,
        "can_manage": can_manage,
        "can_delete": can_delete,
        "can_update": can_update,
        "update_form": update_form,
        "role_reason_label": role_reason_label,
        "saved_role_reason_label": saved_role_reason_label,
        "is_completed_locked": is_completed_locked,
        "allow_task_progress_update": setting.allow_task_progress_update,
        "allow_task_completion": setting.allow_task_completion,
    }

    return render(request, "tasks/task_detail.html", context)


# --------------------------------------------------
# Hold Task
# --------------------------------------------------
@login_required
def hold_task(request, pk):
    task = get_object_or_404(Task, pk=pk)

    if task.status == "COMPLETED" or (task.progress_percent or 0) >= 100:
        messages.error(request, "A completed task is view-only and cannot be put on hold.")
        return redirect("task_detail", pk=task.pk)

    if not _can_manage_task(request.user, task):
        return HttpResponseForbidden("You are not allowed to hold this task.")

    if task.status == "CANCELLED":
        messages.error(request, "A cancelled task cannot be put on hold.")
        return redirect("task_detail", pk=task.pk)

    if request.method != "POST":
        return HttpResponseForbidden("Invalid request method.")

    form = TaskManagementNoteForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Reason is required when putting a task on hold.")
        return redirect("task_detail", pk=task.pk)

    task.status = "ON_HOLD"
    task.management_note = form.cleaned_data["management_note"]
    task.management_note_updated_at = timezone.now()
    task.management_note_by = request.user
    task.management_note_role = _get_role_reason_label(request.user)
    task.save(update_fields=[
        "status",
        "management_note",
        "management_note_updated_at",
        "management_note_by",
        "management_note_role",
        "completed_at",
        "updated_at",
    ])

    messages.success(request, "Task has been placed on hold.")
    return redirect("task_detail", pk=task.pk)


# --------------------------------------------------
# Resume Task
# --------------------------------------------------
@login_required
def resume_task(request, pk):
    task = get_object_or_404(Task, pk=pk)

    if not _can_manage_task(request.user, task):
        return HttpResponseForbidden("You are not allowed to resume this task.")

    if task.status == "CANCELLED":
        messages.error(request, "A cancelled task cannot be resumed.")
        return redirect("task_detail", pk=task.pk)

    if task.status != "ON_HOLD":
        messages.info(request, "This task is not currently on hold.")
        return redirect("task_detail", pk=task.pk)

    task.status = "PENDING"
    task.save(update_fields=["status", "updated_at"])
    task.refresh_from_assignments()

    messages.success(request, "Task has been resumed.")
    return redirect("task_detail", pk=task.pk)


# --------------------------------------------------
# Cancel Task
# --------------------------------------------------
@login_required
def cancel_task(request, pk):
    task = get_object_or_404(Task, pk=pk)

    if task.status == "COMPLETED" or (task.progress_percent or 0) >= 100:
        messages.error(request, "A completed task is view-only and cannot be cancelled.")
        return redirect("task_detail", pk=task.pk)

    if not _can_manage_task(request.user, task):
        return HttpResponseForbidden("You are not allowed to cancel this task.")

    if request.method != "POST":
        return HttpResponseForbidden("Invalid request method.")

    form = TaskManagementNoteForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Reason is required when cancelling a task.")
        return redirect("task_detail", pk=task.pk)

    task.status = "CANCELLED"
    task.completed_at = None
    task.management_note = form.cleaned_data["management_note"]
    task.management_note_updated_at = timezone.now()
    task.management_note_by = request.user
    task.management_note_role = _get_role_reason_label(request.user)
    task.save(update_fields=[
        "status",
        "completed_at",
        "management_note",
        "management_note_updated_at",
        "management_note_by",
        "management_note_role",
        "updated_at",
    ])

    messages.success(request, "Task has been cancelled.")
    return redirect("task_detail", pk=task.pk)


# --------------------------------------------------
# Delete Task
# --------------------------------------------------
@login_required
def delete_task(request, pk):
    task = get_object_or_404(Task, pk=pk)

    if task.status == "COMPLETED" or (task.progress_percent or 0) >= 100:
        messages.error(request, "A completed task is view-only and cannot be deleted.")
        return redirect("task_detail", pk=task.pk)

    if task.created_by != request.user:
        return HttpResponseForbidden("Only the user who created this task can remove it.")

    if request.method == "POST":
        task.delete()
        messages.success(request, "Task removed successfully.")
        return redirect("task_dashboard")

    return render(request, "tasks/delete_task.html", {"task": task})


# --------------------------------------------------
# Accept Task
# --------------------------------------------------
@login_required
def accept_task(request, pk):
    assignment = get_object_or_404(
        TaskAssignment.objects.select_related("task", "assigned_to"),
        pk=pk
    )

    if assignment.assigned_to != request.user:
        return HttpResponseForbidden("You are not allowed to accept this task.")

    if assignment.task.status == "CANCELLED":
        messages.error(request, "This task has been cancelled and cannot be accepted.")
        return redirect("task_detail", pk=assignment.task.pk)

    if assignment.task.status == "ON_HOLD":
        messages.error(request, "This task is currently on hold and cannot be accepted.")
        return redirect("task_detail", pk=assignment.task.pk)

    if assignment.task.status == "RETURNED":
        messages.error(request, "This task has been returned and cannot be accepted until it is re-assigned.")
        return redirect("task_detail", pk=assignment.task.pk)

    if assignment.status == "ASSIGNED":
        assignment.status = "ACCEPTED"
        assignment.accepted_at = timezone.now()
        assignment.save(update_fields=["status", "accepted_at", "updated_at"])
        messages.success(request, "Task accepted successfully.")

    return redirect("task_detail", pk=assignment.task.pk)


# --------------------------------------------------
# Complete Task
# --------------------------------------------------
@login_required
def complete_task(request, pk):
    setting = _get_system_setting()

    if not setting.allow_task_completion:
        messages.error(request, "Task completion is currently disabled by the System Administrator.")
        return redirect("task_dashboard")
    assignment = get_object_or_404(
        TaskAssignment.objects.select_related("task", "assigned_to"),
        pk=pk
    )

    if assignment.task.status == "COMPLETED" or (assignment.task.progress_percent or 0) >= 100:
        messages.info(request, "This task is already completed and is now view-only.")
        return redirect("task_detail", pk=assignment.task.pk)

    if assignment.assigned_to != request.user and not _can_manage_task(request.user, assignment.task):
        return HttpResponseForbidden("You are not allowed to complete this task.")

    if assignment.task.status == "CANCELLED":
        messages.error(request, "This task has been cancelled and cannot be completed.")
        return redirect("task_detail", pk=assignment.task.pk)

    if assignment.task.status == "ON_HOLD":
        messages.error(request, "This task is currently on hold and cannot be completed.")
        return redirect("task_detail", pk=assignment.task.pk)

    if assignment.task.status == "RETURNED":
        messages.error(request, "This task has been returned and cannot be completed until it is re-assigned.")
        return redirect("task_detail", pk=assignment.task.pk)

    now = timezone.now()
    baseline = getattr(assignment, "carried_forward_progress", 0) or 0
    assignment.progress_percent = max(0, 100 - baseline)
    assignment.status = "COMPLETED"
    assignment.completed_at = now
    assignment.last_updated_at = now

    if not assignment.started_at:
        assignment.started_at = now

    assignment.save(update_fields=[
        "progress_percent",
        "status",
        "started_at",
        "last_updated_at",
        "completed_at",
        "updated_at",
    ])

    assignment.task.refresh_from_assignments()

    messages.success(request, "Task marked as completed.")
    return redirect("task_detail", pk=assignment.task.pk)


# --------------------------------------------------
# Returned Task View
# --------------------------------------------------
@login_required
def return_task(request, pk):
    task = get_object_or_404(Task, pk=pk)

    if task.status == "COMPLETED" or (task.progress_percent or 0) >= 100:
        messages.error(request, "A completed task is view-only and cannot be returned.")
        return redirect("task_detail", pk=task.pk)

    profile = _get_profile(request.user)

    if _get_user_role(request.user) in ["DIRECTOR"]:
        return HttpResponseForbidden("Directors cannot return tasks.")

    assignment = task.assignments.filter(
        assigned_to=request.user,
        status__in=["ASSIGNED", "ACCEPTED", "IN_PROGRESS"]
    ).order_by("-assigned_at").first()

    if not assignment:
        return HttpResponseForbidden("You are not assigned to this task.")

    if task.status in ["COMPLETED", "CANCELLED", "ON_HOLD", "RETURNED"]:
        messages.error(request, "This task cannot be returned.")
        return redirect("task_detail", pk=task.pk)

    if request.method == "POST":
        form = TaskReturnForm(request.POST)

        if form.is_valid():
            reason = form.cleaned_data["return_reason"]
            now = timezone.now()

            assignment.status = "RETURNED"
            assignment.returned_reason = reason
            assignment.returned_at = now
            assignment.last_updated_at = now
            assignment.completed_at = None
            assignment.save(update_fields=[
                "status",
                "returned_reason",
                "returned_at",
                "last_updated_at",
                "completed_at",
                "updated_at",
            ])

            effective_progress = min(
                100,
                (assignment.carried_forward_progress or 0) + (assignment.progress_percent or 0)
            )

            task.status = "RETURNED"
            task.progress_percent = effective_progress
            task.returned_reason = reason
            task.returned_at = now
            task.returned_by = request.user
            task.completed_at = None
            task.save(update_fields=[
                "status",
                "progress_percent",
                "returned_reason",
                "returned_at",
                "returned_by",
                "completed_at",
                "updated_at",
            ])

            messages.success(request, "Task returned successfully to the creator.")
            return redirect("task_detail", pk=task.pk)
    else:
        form = TaskReturnForm()

    return render(request, "tasks/return_task.html", {
        "task": task,
        "form": form,
    })


# --------------------------------------------------
# Re-Assign Returned Task View
# --------------------------------------------------
@login_required
def reassign_returned_task(request, pk):
    task = get_object_or_404(Task, pk=pk)

    if task.status == "COMPLETED" or (task.progress_percent or 0) >= 100:
        messages.error(request, "A completed task is view-only and cannot be re-assigned.")
        return redirect("task_detail", pk=task.pk)

    if not _can_manage_task(request.user, task):
        return HttpResponseForbidden("You are not allowed to re-assign this task.")

    if task.status != "RETURNED":
        messages.error(request, "Only returned tasks can be re-assigned.")
        return redirect("task_detail", pk=task.pk)

    if request.method == "POST":
        form = TaskReassignForm(request.POST, user=request.user, task=task)

        if form.is_valid():
            new_user = form.cleaned_data["assigned_to"]
            now = timezone.now()

            latest_returned = task.assignments.filter(
                status="RETURNED"
            ).order_by("-assigned_at").first()

            if latest_returned:
                inherited_progress = min(
                    100,
                    (latest_returned.carried_forward_progress or 0) +
                    (latest_returned.progress_percent or 0)
                )
            else:
                inherited_progress = task.progress_percent or 0

            with transaction.atomic():
                TaskAssignment.objects.create(
                    task=task,
                    assigned_to=new_user,
                    assigned_by=request.user,
                    status="ASSIGNED",
                    progress_percent=0,
                    carried_forward_progress=inherited_progress,
                    accepted_at=None,
                    started_at=None,
                    last_updated_at=None,
                    completed_at=None,
                )

                task.status = "PENDING"
                task.progress_percent = inherited_progress
                task.returned_reason = ""
                task.returned_at = None
                task.returned_by = None
                task.completed_at = None
                task.save(update_fields=[
                    "status",
                    "progress_percent",
                    "returned_reason",
                    "returned_at",
                    "returned_by",
                    "completed_at",
                    "updated_at",
                ])

            messages.success(
                request,
                f"Task re-assigned successfully to {new_user.get_full_name() or new_user.username}."
            )
            return redirect("task_detail", pk=task.pk)
    else:
        form = TaskReassignForm(user=request.user, task=task)

    return render(request, "tasks/reassign_returned_task.html", {
        "task": task,
        "form": form,
    })


# --------------------------------------------------
# Analytics in the Director's Dashboard
# --------------------------------------------------
@login_required
def task_analytics(request):
    profile = _get_profile(request.user)

    if not (
        request.user.is_superuser
        or _get_user_role(request.user) in ["ADMIN", "DIRECTOR", "ADRD", "ADSTI", "HEAD_OF_UNIT"]
        or _get_user_role(request.user) in EXECUTIVE_TASK_ROLES
    ):
        return HttpResponseForbidden("You are not allowed to view analytics.")

    now = timezone.now()

    all_tasks, all_assignments = (
        _get_visible_task_scope(
            request.user,
            strict_department=True,
        )
    )

    total_tasks = all_tasks.count()
    completed_tasks = all_tasks.filter(status="COMPLETED").count()
    overdue_tasks = all_tasks.filter(
        due_date__lt=now
    ).exclude(
        status__in=["COMPLETED", "CANCELLED", "RETURNED"]
    ).count()

    completion_rate = round((completed_tasks / total_tasks) * 100, 1) if total_tasks > 0 else 0

    completed_task_list = [task for task in all_tasks if task.status == "COMPLETED"]

    on_time_completed = 0
    late_completed = 0
    late_delay_days_values = []

    for task in completed_task_list:
        if not task.due_date or not task.completed_at:
            on_time_completed += 1
            continue

        days_late = _date_diff_days_late(task.due_date, task.completed_at)
        if days_late > 0:
            late_completed += 1
            late_delay_days_values.append(days_late)
        else:
            on_time_completed += 1

    open_overdue_tasks = [
        task for task in all_tasks
        if task.due_date and task.due_date < now and _is_open_task(task)
    ]

    overdue_delay_days_values = [
        _date_diff_days_late(task.due_date, now)
        for task in open_overdue_tasks
    ]

    combined_delay_values = late_delay_days_values + overdue_delay_days_values
    avg_delay_days = round(sum(combined_delay_values) / len(combined_delay_values), 1) if combined_delay_values else 0

    due_soon_list = []
    for task in all_tasks:
        if not task.due_date:
            continue
        if not _is_open_task(task):
            continue

        days_remaining = _date_diff_days_remaining(now, task.due_date)
        if days_remaining is None:
            continue

        if 0 <= days_remaining <= 3:
            latest_assignment = _latest_assignment_for_task(task)
            due_soon_list.append({
                "task_id": task.id,
                "title": task.title,
                "unit_name": task.unit_name or "-",
                "assignee_name": _safe_full_name(latest_assignment.assigned_to) if latest_assignment else "-",
                "progress": task.progress_percent or 0,
                "due_date": timezone.localtime(task.due_date).strftime("%Y-%m-%d %H:%M") if timezone.is_aware(task.due_date) else task.due_date.strftime("%Y-%m-%d %H:%M"),
                "days_remaining": days_remaining,
            })

    due_soon_list = sorted(due_soon_list, key=lambda x: (x["days_remaining"], x["progress"]))[:10]
    due_soon_tasks = len(due_soon_list)

    overdue_1_3 = 0
    overdue_4_7 = 0
    overdue_8_14 = 0
    overdue_15_plus = 0

    for task in open_overdue_tasks:
        days_late = _date_diff_days_late(task.due_date, now)

        if 1 <= days_late <= 3:
            overdue_1_3 += 1
        elif 4 <= days_late <= 7:
            overdue_4_7 += 1
        elif 8 <= days_late <= 14:
            overdue_8_14 += 1
        elif days_late >= 15:
            overdue_15_plus += 1

    units = (
        all_tasks
        .exclude(unit_name__isnull=True)
        .exclude(unit_name="")
        .values_list("unit_name", flat=True)
        .distinct()
        .order_by("unit_name")
    )

    task_counts = (
        all_tasks
        .exclude(unit_name__isnull=True)
        .exclude(unit_name="")
        .values("unit_name")
        .annotate(
            total=Count("id"),
            completed=Count("id", filter=Q(status="COMPLETED")),
        )
    )

    task_map = {
        row["unit_name"]: row
        for row in task_counts
    }

    tasks_per_unit = []
    completion_rate_by_unit = []

    for unit in units:
        row = task_map.get(unit)

        total = row["total"] if row else 0
        completed = row["completed"] if row else 0
        rate = round((completed / total) * 100, 1) if total > 0 else 0

        tasks_per_unit.append({
            "unit_name": unit,
            "total": total,
        })

        completion_rate_by_unit.append({
            "unit_name": unit,
            "total": total,
            "completed": completed,
            "rate": rate,
        })

    delay_by_unit = []

    for unit in units:
        unit_tasks = [task for task in all_tasks if (task.unit_name or "") == unit]
        unit_total_completed = 0
        unit_on_time_completed = 0
        unit_late_completed = 0
        unit_overdue = 0
        unit_delay_values = []

        for task in unit_tasks:
            if task.status == "COMPLETED":
                unit_total_completed += 1
                if task.due_date and task.completed_at:
                    days_late = _date_diff_days_late(task.due_date, task.completed_at)
                    if days_late > 0:
                        unit_late_completed += 1
                        unit_delay_values.append(days_late)
                    else:
                        unit_on_time_completed += 1
                else:
                    unit_on_time_completed += 1

            elif task.due_date and task.due_date < now and _is_open_task(task):
                unit_overdue += 1
                unit_delay_values.append(_date_diff_days_late(task.due_date, now))

        on_time_rate = round((unit_on_time_completed / unit_total_completed) * 100, 1) if unit_total_completed > 0 else 0
        avg_unit_delay = round(sum(unit_delay_values) / len(unit_delay_values), 1) if unit_delay_values else 0

        delay_by_unit.append({
            "unit_name": unit,
            "overdue": unit_overdue,
            "late_completed": unit_late_completed,
            "avg_delay_days": avg_unit_delay,
            "on_time_rate": on_time_rate,
        })

    delay_by_unit = sorted(delay_by_unit, key=lambda x: (-x["overdue"], -x["late_completed"], -x["avg_delay_days"], x["unit_name"]))

    top_performers = (
        all_assignments
        .values(
            "assigned_to__id",
            "assigned_to__username",
            "assigned_to__first_name",
            "assigned_to__last_name",
            "assigned_to__profile__unit_name",
        )
        .annotate(
            total_assignments=Count("id"),
            completed_assignments=Count("id", filter=Q(status="COMPLETED")),
            avg_progress=Avg("progress_percent"),
        )
        .order_by("-completed_assignments", "-avg_progress", "-total_assignments")[:10]
    )

    top_performers = [
        {
            "user_id": row["assigned_to__id"],
            "name": (
                f"{row['assigned_to__first_name']} {row['assigned_to__last_name']}".strip()
                or row["assigned_to__username"]
            ),
            "unit_name": row["assigned_to__profile__unit_name"] or "-",
            "total_assignments": row["total_assignments"],
            "completed_assignments": row["completed_assignments"],
            "avg_progress": round(row["avg_progress"] or 0, 1),
            "completion_rate": round(
                (row["completed_assignments"] / row["total_assignments"]) * 100, 1
            ) if row["total_assignments"] > 0 else 0,
        }
        for row in top_performers
    ]

    staff_delay_map = {}

    for assignment in all_assignments:
        user = assignment.assigned_to
        if not user:
            continue

        unit_name = getattr(getattr(user, "profile", None), "unit_name", "") or "-"
        key = user.id

        if key not in staff_delay_map:
            staff_delay_map[key] = {
                "user_id": user.id,
                "name": _safe_full_name(user),
                "unit_name": unit_name,
                "open_overdue": 0,
                "late_completed": 0,
                "avg_delay_days": 0,
                "on_time_rate": 0,
                "_completed_total": 0,
                "_completed_on_time": 0,
                "_delay_values": [],
            }

        row = staff_delay_map[key]
        due_date = assignment.task.due_date

        if assignment.status == "COMPLETED":
            row["_completed_total"] += 1
            if due_date and assignment.completed_at:
                days_late = _date_diff_days_late(due_date, assignment.completed_at)
                if days_late > 0:
                    row["late_completed"] += 1
                    row["_delay_values"].append(days_late)
                else:
                    row["_completed_on_time"] += 1
            else:
                row["_completed_on_time"] += 1

        elif due_date and due_date < now and assignment.status in ["ASSIGNED", "ACCEPTED", "IN_PROGRESS"]:
            row["open_overdue"] += 1
            row["_delay_values"].append(_date_diff_days_late(due_date, now))

    delayed_assignments_by_staff = []
    for _, row in staff_delay_map.items():
        completed_total = row["_completed_total"]
        completed_on_time = row["_completed_on_time"]
        row["on_time_rate"] = round((completed_on_time / completed_total) * 100, 1) if completed_total > 0 else 0
        row["avg_delay_days"] = round(sum(row["_delay_values"]) / len(row["_delay_values"]), 1) if row["_delay_values"] else 0

        delayed_assignments_by_staff.append({
            "user_id": row["user_id"],
            "name": row["name"],
            "unit_name": row["unit_name"],
            "open_overdue": row["open_overdue"],
            "late_completed": row["late_completed"],
            "avg_delay_days": row["avg_delay_days"],
            "on_time_rate": row["on_time_rate"],
        })

    delayed_assignments_by_staff = sorted(
        delayed_assignments_by_staff,
        key=lambda x: (-x["open_overdue"], -x["late_completed"], -x["avg_delay_days"], x["name"])
    )[:10]

    stalled_tasks = []

    for task in all_tasks:
        if not task.due_date:
            continue
        if not _is_open_task(task):
            continue

        days_remaining = _date_diff_days_remaining(now, task.due_date)
        days_late = _date_diff_days_late(task.due_date, now) if task.due_date < now else 0
        progress = task.progress_percent or 0

        is_stalled = False
        if days_late > 0:
            is_stalled = True
        elif progress < 50 and days_remaining is not None and days_remaining <= 3:
            is_stalled = True
        elif progress == 0 and days_remaining is not None and days_remaining <= 7:
            is_stalled = True

        if is_stalled:
            latest_assignment = _latest_assignment_for_task(task)
            stalled_tasks.append({
                "task_id": task.id,
                "title": task.title,
                "unit_name": task.unit_name or "-",
                "assignee_name": _safe_full_name(latest_assignment.assigned_to) if latest_assignment else "-",
                "progress": progress,
                "due_date": timezone.localtime(task.due_date).strftime("%Y-%m-%d %H:%M") if timezone.is_aware(task.due_date) else task.due_date.strftime("%Y-%m-%d %H:%M"),
                "days_late": days_late,
            })

    stalled_tasks = sorted(
        stalled_tasks,
        key=lambda x: (-x["days_late"], x["progress"], x["due_date"])
    )[:10]

    context = {
        "profile": profile,
        "total_tasks": total_tasks,
        "completed_tasks": completed_tasks,
        "overdue_tasks": overdue_tasks,
        "completion_rate": completion_rate,
        "on_time_completed": on_time_completed,
        "late_completed": late_completed,
        "avg_delay_days": avg_delay_days,
        "due_soon_tasks": due_soon_tasks,
        "overdue_1_3": overdue_1_3,
        "overdue_4_7": overdue_4_7,
        "overdue_8_14": overdue_8_14,
        "overdue_15_plus": overdue_15_plus,
        "tasks_per_unit": tasks_per_unit,
        "completion_rate_by_unit": completion_rate_by_unit,
        "delay_by_unit": delay_by_unit,
        "top_performers": top_performers,
        "delayed_assignments_by_staff": delayed_assignments_by_staff,
        "due_soon_list": due_soon_list,
        "stalled_tasks": stalled_tasks,
        "allow_task_export": _get_system_setting().allow_task_export,
    }

    return render(request, "tasks/task_analytics.html", context)


# --------------------------------------------------
# Overdue Detail View (for Aging Table Action)
# --------------------------------------------------
@login_required
def task_analytics_overdue_detail(request, range_key):
    profile = _get_profile(request.user)
    role = _get_user_role(request.user)

    if role not in [
        "ADMIN",
        "DIRECTOR",
        "ASSISTANT_DIRECTOR",
        "HEAD_OF_UNIT",
    ] and role not in EXECUTIVE_TASK_ROLES:
        return HttpResponseForbidden(
            "You are not allowed to view analytics."
        )

    now = timezone.now()

    all_tasks, _assignments = _get_visible_task_scope(
        request.user, strict_department=True
    )

    filtered_tasks = []

    for task in all_tasks:
        if not task.due_date:
            continue
        if not _is_open_task(task):
            continue
        if task.due_date >= now:
            continue

        days_late = _date_diff_days_late(task.due_date, now)

        match = False
        if range_key == "1-3" and 1 <= days_late <= 3:
            match = True
        elif range_key == "4-7" and 4 <= days_late <= 7:
            match = True
        elif range_key == "8-14" and 8 <= days_late <= 14:
            match = True
        elif range_key == "15-plus" and days_late >= 15:
            match = True

        if match:
            latest_assignment = _latest_assignment_for_task(task)

            filtered_tasks.append({
                "task_id": task.id,
                "title": task.title,
                "unit_name": task.unit_name or "-",
                "assignee_name": _safe_full_name(latest_assignment.assigned_to) if latest_assignment else "-",
                "progress": task.progress_percent or 0,
                "due_date": timezone.localtime(task.due_date).strftime("%Y-%m-%d %H:%M")
                if task.due_date and timezone.is_aware(task.due_date)
                else (task.due_date.strftime("%Y-%m-%d %H:%M") if task.due_date else "-"),
                "days_late": days_late,
            })

    filtered_tasks = sorted(filtered_tasks, key=lambda x: -x["days_late"])

    return render(request, "tasks/overdue_detail.html", {
        "tasks": filtered_tasks,
        "range_key": range_key,
    })


# --------------------------------------------------
# Completion-by-Unit Detail View
# --------------------------------------------------
@login_required
def task_analytics_completion_unit_detail(request, unit_name):
    profile = _get_profile(request.user)

    if not (
        request.user.is_superuser
        or _get_user_role(request.user) == "DIRECTOR"
        or _get_user_role(request.user) in EXECUTIVE_TASK_ROLES
    ):
        return HttpResponseForbidden("You are not allowed to view this.")

    unit_name = (unit_name or "").strip()
    if not unit_name:
        messages.error(request, "Unit name is required.")
        return redirect("task_analytics")

    tasks, _assignments = _get_visible_task_scope(
        request.user, strict_department=True
    )
    tasks = tasks.filter(unit_name=unit_name).order_by("-created_at")

    detail_rows = []
    for task in tasks:
        latest_assignment = _latest_assignment_for_task(task)
        detail_rows.append({
            "task_id": task.id,
            "title": task.title,
            "status": task.get_status_display(),
            "assignee_name": _safe_full_name(latest_assignment.assigned_to) if latest_assignment else "-",
            "progress": task.progress_percent or 0,
            "due_date": timezone.localtime(task.due_date).strftime("%Y-%m-%d %H:%M")
            if task.due_date and timezone.is_aware(task.due_date)
            else (task.due_date.strftime("%Y-%m-%d %H:%M") if task.due_date else "-"),
        })

    return render(request, "tasks/analytics_unit_tasks_detail.html", {
        "page_title": f"Completion Details - {unit_name}",
        "unit_name": unit_name,
        "detail_rows": detail_rows,
        "detail_type": "completion",
    })


# --------------------------------------------------
# Delay-by-Unit Detail View
# --------------------------------------------------
@login_required
def task_analytics_delay_unit_detail(request, unit_name):
    profile = _get_profile(request.user)

    if not (
        request.user.is_superuser
        or _get_user_role(request.user) == "DIRECTOR"
        or _get_user_role(request.user) in EXECUTIVE_TASK_ROLES
    ):
        return HttpResponseForbidden("You are not allowed to view this.")

    unit_name = (unit_name or "").strip()
    if not unit_name:
        messages.error(request, "Unit name is required.")
        return redirect("task_analytics")

    now = timezone.now()

    tasks, _assignments = _get_visible_task_scope(
        request.user, strict_department=True
    )
    tasks = tasks.filter(unit_name=unit_name).order_by("-created_at")

    detail_rows = []
    for task in tasks:
        days_late = 0
        is_late_completed = False
        is_open_overdue = False

        if task.status == "COMPLETED" and task.due_date and task.completed_at:
            days_late = _date_diff_days_late(task.due_date, task.completed_at)
            is_late_completed = days_late > 0
        elif task.due_date and task.due_date < now and _is_open_task(task):
            days_late = _date_diff_days_late(task.due_date, now)
            is_open_overdue = days_late > 0

        if not (is_late_completed or is_open_overdue):
            continue

        latest_assignment = _latest_assignment_for_task(task)
        detail_rows.append({
            "task_id": task.id,
            "title": task.title,
            "status": "Open Overdue" if is_open_overdue else "Late Completed",
            "assignee_name": _safe_full_name(latest_assignment.assigned_to) if latest_assignment else "-",
            "progress": task.progress_percent or 0,
            "due_date": timezone.localtime(task.due_date).strftime("%Y-%m-%d %H:%M")
            if task.due_date and timezone.is_aware(task.due_date)
            else (task.due_date.strftime("%Y-%m-%d %H:%M") if task.due_date else "-"),
            "delay_days": days_late,
        })

    detail_rows = sorted(detail_rows, key=lambda x: (-x["delay_days"], x["title"]))

    return render(request, "tasks/analytics_unit_tasks_detail.html", {
        "page_title": f"Delay Details - {unit_name}",
        "unit_name": unit_name,
        "detail_rows": detail_rows,
        "detail_type": "delay",
    })


# --------------------------------------------------
# Staff Detail View (Top Performers / Delayed Staff)
# --------------------------------------------------
@login_required
def task_analytics_staff_detail(request, user_id):
    profile = _get_profile(request.user)

    if not (
        request.user.is_superuser
        or _get_user_role(request.user) == "DIRECTOR"
        or _get_user_role(request.user) in EXECUTIVE_TASK_ROLES
    ):
        return HttpResponseForbidden("You are not allowed to view this.")

    staff_user = get_object_or_404(User.objects.select_related("profile"), pk=user_id)

    _tasks, assignments = _get_visible_task_scope(
        request.user, strict_department=True
    )
    assignments = assignments.filter(assigned_to_id=user_id).order_by("-assigned_at")
    if not assignments.exists():
        return HttpResponseForbidden("This official is outside your analytics scope.")

    detail_rows = []
    for assignment in assignments:
        detail_rows.append({
            "task_id": assignment.task.id,
            "task_title": assignment.task.title,
            "unit_name": assignment.task.unit_name or "-",
            "assignment_status": assignment.get_status_display(),
            "task_status": assignment.task.get_status_display(),
            "progress": assignment.progress_percent or 0,
            "due_date": timezone.localtime(assignment.task.due_date).strftime("%Y-%m-%d %H:%M")
            if assignment.task.due_date and timezone.is_aware(assignment.task.due_date)
            else (assignment.task.due_date.strftime("%Y-%m-%d %H:%M") if assignment.task.due_date else "-"),
        })

    return render(request, "tasks/analytics_staff_detail.html", {
        "page_title": f"Staff Assignment Details - {_safe_full_name(staff_user)}",
        "staff_name": _safe_full_name(staff_user),
        "staff_unit": getattr(getattr(staff_user, "profile", None), "unit_name", "") or "-",
        "detail_rows": detail_rows,
    })

# --------------------------------------------------
# Due soon tasks
# --------------------------------------------------
@login_required
def task_analytics_due_soon_detail(request):
    profile = _get_profile(request.user)

    if not (
        request.user.is_superuser
        or _get_user_role(request.user) == "DIRECTOR"
        or _get_user_role(request.user) in EXECUTIVE_TASK_ROLES
    ):
        return HttpResponseForbidden("You are not allowed to view this.")

    now = timezone.now()

    all_tasks, _assignments = _get_visible_task_scope(
        request.user, strict_department=True
    )

    detail_rows = []

    for task in all_tasks:
        if not task.due_date:
            continue
        if not _is_open_task(task):
            continue

        days_remaining = _date_diff_days_remaining(now, task.due_date)
        if days_remaining is None:
            continue

        if 0 <= days_remaining <= 3:
            latest_assignment = _latest_assignment_for_task(task)
            detail_rows.append({
                "task_id": task.id,
                "title": task.title,
                "unit_name": task.unit_name or "-",
                "assignee_name": _safe_full_name(latest_assignment.assigned_to) if latest_assignment else "-",
                "progress": task.progress_percent or 0,
                "due_date": timezone.localtime(task.due_date).strftime("%Y-%m-%d %H:%M")
                if timezone.is_aware(task.due_date)
                else task.due_date.strftime("%Y-%m-%d %H:%M"),
                "days_remaining": days_remaining,
            })

    detail_rows = sorted(detail_rows, key=lambda x: (x["days_remaining"], x["progress"], x["title"]))

    return render(request, "tasks/analytics_due_soon_detail.html", {
        "page_title": "Due Soon Tasks Details",
        "detail_rows": detail_rows,
    })

# --------------------------------------------------
# Task_analytics_stalled_detail
# --------------------------------------------------
@login_required
def task_analytics_stalled_detail(request):
    profile = _get_profile(request.user)

    if not (
        request.user.is_superuser
        or _get_user_role(request.user) == "DIRECTOR"
        or _get_user_role(request.user) in EXECUTIVE_TASK_ROLES
    ):
        return HttpResponseForbidden("You are not allowed to view this.")

    now = timezone.now()

    all_tasks, _assignments = _get_visible_task_scope(
        request.user, strict_department=True
    )

    detail_rows = []

    for task in all_tasks:
        if not task.due_date:
            continue
        if not _is_open_task(task):
            continue

        days_remaining = _date_diff_days_remaining(now, task.due_date)
        days_late = _date_diff_days_late(task.due_date, now) if task.due_date < now else 0
        progress = task.progress_percent or 0

        is_stalled = False
        status_label = "At Risk"

        if days_late > 0:
            is_stalled = True
            status_label = "Overdue"
        elif progress < 50 and days_remaining is not None and days_remaining <= 3:
            is_stalled = True
            status_label = "Slow Progress"
        elif progress == 0 and days_remaining is not None and days_remaining <= 7:
            is_stalled = True
            status_label = "At Risk"

        if is_stalled:
            latest_assignment = _latest_assignment_for_task(task)
            detail_rows.append({
                "task_id": task.id,
                "title": task.title,
                "unit_name": task.unit_name or "-",
                "assignee_name": _safe_full_name(latest_assignment.assigned_to) if latest_assignment else "-",
                "progress": progress,
                "due_date": timezone.localtime(task.due_date).strftime("%Y-%m-%d %H:%M")
                if timezone.is_aware(task.due_date)
                else task.due_date.strftime("%Y-%m-%d %H:%M"),
                "days_late": days_late,
                "status_label": status_label,
            })

    detail_rows = sorted(detail_rows, key=lambda x: (-x["days_late"], x["progress"], x["title"]))

    return render(request, "tasks/analytics_stalled_detail.html", {
        "page_title": "High-Risk / Stalled Tasks Details",
        "detail_rows": detail_rows,
    })


# --------------------------------------------------
# Export Analytics to Excel (Exports Active/Selected Table Only)
# --------------------------------------------------
@login_required
def task_analytics_export_excel(request):
    profile = _get_profile(request.user)

    if not (
        request.user.is_superuser
        or _get_user_role(request.user) in ["ADMIN", "DIRECTOR", "ADRD", "ADSTI", "HEAD_OF_UNIT"]
        or _get_user_role(request.user) in EXECUTIVE_TASK_ROLES
    ):
        return HttpResponseForbidden("You are not allowed to export this.")

    selected_table = request.GET.get("table", "all_tables").strip()
    export_format = request.GET.get("format", "").strip()

    now = timezone.now()

    all_tasks, all_assignments = (
        _get_visible_task_scope(
            request.user,
            strict_department=True,
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    def style_title(row_num, text, merge_to_col=6):
        ws.cell(row=row_num, column=1, value=text)
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=14, color="FFFFFF")
        ws.cell(row=row_num, column=1).fill = title_fill
        ws.cell(row=row_num, column=1).alignment = center
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=merge_to_col)

        for col in range(1, merge_to_col + 1):
            ws.cell(row=row_num, column=col).fill = title_fill
            ws.cell(row=row_num, column=col).border = border

    #----------------------------------------------------------------------------------------
    def style_header(row_num, headers):
        ws._header_map = {}  # store header → column mapping
        left_headers = {"Name", "Task", "Unit", "Assignee", "Status", "Range", "Due Date"}
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            ws._header_map[col_num] = header  # store mapping
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = border
            # ✅ Apply alignment ALSO to header row
            if header in left_headers:
                cell.alignment = left
            else:
                    cell.alignment = center
    #----------------------------------------------------------------------------------------

    def style_data_row(row_num, values):
        left_headers = {"Name", "Task", "Unit", "Assignee", "Status", "Range", "Due Date"}
        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            header_text = ws._header_map.get(col_num)
            if header_text in left_headers:
                cell.alignment = left
            else:
                cell.alignment = center
            cell.border = border

    def auto_fit_columns():
        fixed_widths = {
            "A": 6,
        }

        header_based_max = {
            "#": 6,
            "Name": 22,
            "Task": 28,
            "Unit": 18,
            "Assignee": 22,
            "Status": 16,
            "Due Date": 20,
            "Days Late": 12,
            "Days Remaining": 14,
            "Progress (%)": 14,
            "Rate (%)": 12,
            "On-Time Rate (%)": 16,
            "Avg Delay Days": 14,
            "Avg Progress (%)": 16,
            "Completed": 12,
            "Total": 12,
            "Tasks": 10,
            "Range": 18,
        }

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)

            if col_letter in fixed_widths:
                ws.column_dimensions[col_letter].width = fixed_widths[col_letter]
                continue

            max_length = 0
            header_value = None

            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if hasattr(cell, "coordinate") and cell.coordinate in ws.merged_cells:
                    continue

                value = cell.value
                if value is None:
                    continue

                text = str(value).strip()
                if not text:
                    continue

                if header_value is None and row_idx > 1:
                    header_value = text

                text_length = len(text)

                if text_length > 30:
                    text_length = 30 + int((text_length - 30) * 0.35)

                if text_length > max_length:
                    max_length = text_length

            width = max_length + 2

            if header_value in header_based_max:
                width = min(width, header_based_max[header_value])
            else:
                width = min(width, 24)

            width = max(width, 8)
            ws.column_dimensions[col_letter].width = width

    def add_blank_line(row_num):
        return row_num + 1

    current_row = 1

    def build_overdue_aging(row_num):
        overdue_1_3 = 0
        overdue_4_7 = 0
        overdue_8_14 = 0
        overdue_15_plus = 0

        for task in all_tasks:
            if not task.due_date:
                continue
            if not _is_open_task(task):
                continue
            if task.due_date >= now:
                continue

            days_late = _date_diff_days_late(task.due_date, now)

            if 1 <= days_late <= 3:
                overdue_1_3 += 1
            elif 4 <= days_late <= 7:
                overdue_4_7 += 1
            elif 8 <= days_late <= 14:
                overdue_8_14 += 1
            elif days_late >= 15:
                overdue_15_plus += 1

        style_title(row_num, "Overdue Aging Analysis", merge_to_col=2)
        row_num += 1
        style_header(row_num, ["Range", "Tasks"])
        row_num += 1

        rows = [
            ["1–3 Days", overdue_1_3],
            ["4–7 Days", overdue_4_7],
            ["8–14 Days", overdue_8_14],
            ["15+ Days", overdue_15_plus],
        ]
        for row in rows:
            style_data_row(row_num, row)
            row_num += 1

        return add_blank_line(row_num)

    def build_completion_rate_by_unit(row_num):
        units = (
            all_tasks
            .exclude(unit_name__isnull=True)
            .exclude(unit_name="")
            .values_list("unit_name", flat=True)
            .distinct()
            .order_by("unit_name")
        )

        task_counts = (
            all_tasks
            .exclude(unit_name__isnull=True)
            .exclude(unit_name="")
            .values("unit_name")
            .annotate(
                total=Count("id"),
                completed=Count("id", filter=Q(status="COMPLETED")),
            )
        )

        task_map = {row["unit_name"]: row for row in task_counts}

        style_title(row_num, "Completion Rate by Unit", merge_to_col=4)
        row_num += 1
        style_header(row_num, ["Unit", "Completed", "Total", "Rate (%)"])
        row_num += 1

        for unit in units:
            row = task_map.get(unit)
            total = row["total"] if row else 0
            completed = row["completed"] if row else 0
            rate = round((completed / total) * 100, 1) if total > 0 else 0
            style_data_row(row_num, [unit, completed, total, rate])
            row_num += 1

        return add_blank_line(row_num)

    def build_delay_by_unit(row_num):
        units = (
            all_tasks
            .exclude(unit_name__isnull=True)
            .exclude(unit_name="")
            .values_list("unit_name", flat=True)
            .distinct()
            .order_by("unit_name")
        )

        delay_by_unit = []

        for unit in units:
            unit_tasks = [task for task in all_tasks if (task.unit_name or "") == unit]
            unit_total_completed = 0
            unit_on_time_completed = 0
            unit_late_completed = 0
            unit_overdue = 0
            unit_delay_values = []

            for task in unit_tasks:
                if task.status == "COMPLETED":
                    unit_total_completed += 1
                    if task.due_date and task.completed_at:
                        days_late = _date_diff_days_late(task.due_date, task.completed_at)
                        if days_late > 0:
                            unit_late_completed += 1
                            unit_delay_values.append(days_late)
                        else:
                            unit_on_time_completed += 1
                    else:
                        unit_on_time_completed += 1

                elif task.due_date and task.due_date < now and _is_open_task(task):
                    unit_overdue += 1
                    unit_delay_values.append(_date_diff_days_late(task.due_date, now))

            on_time_rate = round((unit_on_time_completed / unit_total_completed) * 100, 1) if unit_total_completed > 0 else 0
            avg_unit_delay = round(sum(unit_delay_values) / len(unit_delay_values), 1) if unit_delay_values else 0

            delay_by_unit.append({
                "unit_name": unit,
                "overdue": unit_overdue,
                "late_completed": unit_late_completed,
                "avg_delay_days": avg_unit_delay,
                "on_time_rate": on_time_rate,
            })

        delay_by_unit = sorted(
            delay_by_unit,
            key=lambda x: (-x["overdue"], -x["late_completed"], -x["avg_delay_days"], x["unit_name"])
        )

        style_title(row_num, "Delay Performance by Unit", merge_to_col=5)
        row_num += 1
        style_header(row_num, ["Unit", "Overdue Tasks", "Late Completed", "Avg Delay Days", "On-Time Rate (%)"])
        row_num += 1

        for row in delay_by_unit:
            style_data_row(row_num, [
                row["unit_name"],
                row["overdue"],
                row["late_completed"],
                row["avg_delay_days"],
                row["on_time_rate"],
            ])
            row_num += 1

        return add_blank_line(row_num)

    def build_top_performers(row_num):
        top_performers = (
            all_assignments
            .values(
                "assigned_to__id",
                "assigned_to__username",
                "assigned_to__first_name",
                "assigned_to__last_name",
                "assigned_to__profile__unit_name",
            )
            .annotate(
                total_assignments=Count("id"),
                completed_assignments=Count("id", filter=Q(status="COMPLETED")),
                avg_progress=Avg("progress_percent"),
            )
            .order_by("-completed_assignments", "-avg_progress", "-total_assignments")[:10]
        )

        style_title(row_num, "Top Performers", merge_to_col=6)
        row_num += 1
        style_header(row_num, ["#", "Name", "Unit", "Completed Assignments", "Total Assignments", "Avg Progress (%)"])
        row_num += 1

        counter = 1
        for row in top_performers:
            name = (
                f"{row['assigned_to__first_name']} {row['assigned_to__last_name']}".strip()
                or row["assigned_to__username"]
            )
            total_assignments = row["total_assignments"]
            completed_assignments = row["completed_assignments"]
            avg_progress = round(row["avg_progress"] or 0, 1)

            style_data_row(row_num, [
                counter,
                name,
                row["assigned_to__profile__unit_name"] or "-",
                completed_assignments,
                total_assignments,
                avg_progress,
            ])
            row_num += 1
            counter += 1

        return add_blank_line(row_num)

    def build_delayed_assignments_by_staff(row_num):
        staff_delay_map = {}

        for assignment in all_assignments:
            user = assignment.assigned_to
            if not user:
                continue

            unit_name = getattr(getattr(user, "profile", None), "unit_name", "") or "-"
            key = user.id

            if key not in staff_delay_map:
                staff_delay_map[key] = {
                    "name": _safe_full_name(user),
                    "unit_name": unit_name,
                    "open_overdue": 0,
                    "late_completed": 0,
                    "_completed_total": 0,
                    "_completed_on_time": 0,
                    "_delay_values": [],
                }

            row = staff_delay_map[key]
            due_date = assignment.task.due_date

            if assignment.status == "COMPLETED":
                row["_completed_total"] += 1
                if due_date and assignment.completed_at:
                    days_late = _date_diff_days_late(due_date, assignment.completed_at)
                    if days_late > 0:
                        row["late_completed"] += 1
                        row["_delay_values"].append(days_late)
                    else:
                        row["_completed_on_time"] += 1
                else:
                    row["_completed_on_time"] += 1

            elif due_date and due_date < now and assignment.status in ["ASSIGNED", "ACCEPTED", "IN_PROGRESS"]:
                row["open_overdue"] += 1
                row["_delay_values"].append(_date_diff_days_late(due_date, now))

        delayed_assignments_by_staff = []
        for row in staff_delay_map.values():
            avg_delay_days = round(sum(row["_delay_values"]) / len(row["_delay_values"]), 1) if row["_delay_values"] else 0
            on_time_rate = round((row["_completed_on_time"] / row["_completed_total"]) * 100, 1) if row["_completed_total"] > 0 else 0

            delayed_assignments_by_staff.append({
                "name": row["name"],
                "unit_name": row["unit_name"],
                "open_overdue": row["open_overdue"],
                "late_completed": row["late_completed"],
                "avg_delay_days": avg_delay_days,
                "on_time_rate": on_time_rate,
            })

        delayed_assignments_by_staff = sorted(
            delayed_assignments_by_staff,
            key=lambda x: (-x["open_overdue"], -x["late_completed"], -x["avg_delay_days"], x["name"])
        )[:10]

        style_title(row_num, "Delayed Assignments by Staff", merge_to_col=7)
        row_num += 1
        style_header(row_num, ["#", "Name", "Unit", "Open Overdue", "Late Completed", "Avg Delay Days", "On-Time Rate (%)"])
        row_num += 1

        counter = 1
        for row in delayed_assignments_by_staff:
            style_data_row(row_num, [
                counter,
                row["name"],
                row["unit_name"],
                row["open_overdue"],
                row["late_completed"],
                row["avg_delay_days"],
                row["on_time_rate"],
            ])
            row_num += 1
            counter += 1

        return add_blank_line(row_num)

    def build_due_soon_list(row_num):
        due_soon_list = []

        for task in all_tasks:
            if not task.due_date:
                continue
            if not _is_open_task(task):
                continue

            days_remaining = _date_diff_days_remaining(now, task.due_date)
            if days_remaining is None:
                continue

            if 0 <= days_remaining <= 3:
                latest_assignment = _latest_assignment_for_task(task)
                due_soon_list.append({
                    "title": task.title,
                    "unit_name": task.unit_name or "-",
                    "assignee_name": _safe_full_name(latest_assignment.assigned_to) if latest_assignment else "-",
                    "progress": task.progress_percent or 0,
                    "due_date": timezone.localtime(task.due_date).strftime("%Y-%m-%d %H:%M") if timezone.is_aware(task.due_date) else task.due_date.strftime("%Y-%m-%d %H:%M"),
                    "days_remaining": days_remaining,
                })

        due_soon_list = sorted(due_soon_list, key=lambda x: (x["days_remaining"], x["progress"]))[:10]

        style_title(row_num, "Due Soon Tasks", merge_to_col=6)
        row_num += 1
        style_header(row_num, ["Task", "Unit", "Assignee", "Progress (%)", "Due Date", "Days Remaining"])
        row_num += 1

        for row in due_soon_list:
            style_data_row(row_num, [
                row["title"],
                row["unit_name"],
                row["assignee_name"],
                row["progress"],
                row["due_date"],
                row["days_remaining"],
            ])
            row_num += 1

        return add_blank_line(row_num)

    def build_stalled_tasks(row_num):
        stalled_tasks = []

        for task in all_tasks:
            if not task.due_date:
                continue
            if not _is_open_task(task):
                continue

            days_remaining = _date_diff_days_remaining(now, task.due_date)
            days_late = _date_diff_days_late(task.due_date, now) if task.due_date < now else 0
            progress = task.progress_percent or 0

            is_stalled = False
            if days_late > 0:
                is_stalled = True
            elif progress < 50 and days_remaining is not None and days_remaining <= 3:
                is_stalled = True
            elif progress == 0 and days_remaining is not None and days_remaining <= 7:
                is_stalled = True

            if is_stalled:
                latest_assignment = _latest_assignment_for_task(task)
                stalled_tasks.append({
                    "title": task.title,
                    "unit_name": task.unit_name or "-",
                    "assignee_name": _safe_full_name(latest_assignment.assigned_to) if latest_assignment else "-",
                    "progress": progress,
                    "due_date": timezone.localtime(task.due_date).strftime("%Y-%m-%d %H:%M") if timezone.is_aware(task.due_date) else task.due_date.strftime("%Y-%m-%d %H:%M"),
                    "days_late": days_late,
                    "status": "Overdue" if days_late > 0 else ("Slow Progress" if progress < 50 else "At Risk"),
                })

        stalled_tasks = sorted(
            stalled_tasks,
            key=lambda x: (-x["days_late"], x["progress"], x["due_date"])
        )[:10]

        style_title(row_num, "High-Risk / Stalled Tasks", merge_to_col=7)
        row_num += 1
        style_header(row_num, ["Task", "Unit", "Assignee", "Progress (%)", "Due Date", "Days Late", "Status"])
        row_num += 1

        for row in stalled_tasks:
            style_data_row(row_num, [
                row["title"],
                row["unit_name"],
                row["assignee_name"],
                row["progress"],
                row["due_date"],
                row["days_late"],
                row["status"],
            ])
            row_num += 1

        return add_blank_line(row_num)

    if selected_table == "all_tables":
        current_row = build_completion_rate_by_unit(current_row)
        current_row = build_overdue_aging(current_row)
        current_row = build_delay_by_unit(current_row)
        current_row = build_top_performers(current_row)
        current_row = build_delayed_assignments_by_staff(current_row)
        current_row = build_due_soon_list(current_row)
        current_row = build_stalled_tasks(current_row)

    elif selected_table == "completion_rate_by_unit":
        current_row = build_completion_rate_by_unit(current_row)

    elif selected_table == "overdue_aging":
        current_row = build_overdue_aging(current_row)

    elif selected_table == "delay_by_unit":
        current_row = build_delay_by_unit(current_row)

    elif selected_table == "top_performers":
        current_row = build_top_performers(current_row)

    elif selected_table == "delayed_assignments_by_staff":
        current_row = build_delayed_assignments_by_staff(current_row)

    elif selected_table == "due_soon_list":
        current_row = build_due_soon_list(current_row)

    elif selected_table == "stalled_tasks":
        current_row = build_stalled_tasks(current_row)

    else:
        style_title(current_row, "Task Analytics", merge_to_col=2)
        current_row += 1
        style_header(current_row, ["Message", "Value"])
        current_row += 1
        style_data_row(current_row, ["Invalid table selected", selected_table])

    auto_fit_columns()
    ws.freeze_panes = "A3"

    if export_format == "a4":
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    filename = f"{selected_table}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
