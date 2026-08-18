from collections import OrderedDict
from dataclasses import dataclass

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone

from .models import (
    SpecialEventParticipant,
    SpecialEventPublication,
    special_event_researcher_identity,
)


EXPECTED_HEADERS = (
    "NA",
    "NAME",
    "INSTITUTION",
    "RESEARCH TITLE",
    "AWARD CATEGORY",
    "YEAR",
)


@dataclass(frozen=True)
class PublicationRow:
    source_sheet: str
    source_number: str
    source_row_index: int
    full_name: str
    institution: str
    research_title: str
    award_category: str
    award_year: str


@dataclass(frozen=True)
class ImportResult:
    researchers_created: int
    researchers_updated: int
    publications_created: int
    publications_updated: int
    skipped: int
    sheets: int

    @property
    def created(self):
        return self.researchers_created

    @property
    def updated(self):
        return self.researchers_updated


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalized_header(value):
    return " ".join(_text(value).upper().split()).rstrip(".")


def _read_publication_rows(uploaded_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError(
            "Excel import is unavailable because openpyxl is not installed."
        ) from exc

    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError(
            "The uploaded file could not be read as an Excel workbook."
        ) from exc

    publication_rows = []
    skipped = matching_sheets = 0
    seen_source_rows = set()
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue
            actual = tuple(_normalized_header(value) for value in header[:6])
            if actual != EXPECTED_HEADERS:
                continue

            matching_sheets += 1
            source_sheet = worksheet.title.strip()
            for row_index, row in enumerate(rows, start=2):
                values = [_text(value) for value in row[:6]]
                if not any(values):
                    skipped += 1
                    continue

                (
                    source_number,
                    full_name,
                    institution,
                    research_title,
                    award_category,
                    award_year,
                ) = values
                required_values = {
                    "Na": source_number,
                    "NAME": full_name,
                    "INSTITUTION": institution,
                    "RESEARCH TITLE": research_title,
                    "AWARD CATEGORY": award_category,
                    "YEAR": award_year,
                }
                missing = [label for label, value in required_values.items() if not value]
                if missing:
                    raise ValidationError(
                        f"Sheet {source_sheet} row {row_index} is missing: "
                        f"{', '.join(missing)}."
                    )

                source_key = (source_sheet.casefold(), source_number.casefold())
                if source_key in seen_source_rows:
                    raise ValidationError(
                        f"Sheet {source_sheet} contains duplicate row number "
                        f"{source_number}."
                    )
                seen_source_rows.add(source_key)
                publication_rows.append(PublicationRow(
                    source_sheet=source_sheet,
                    source_number=source_number,
                    source_row_index=row_index,
                    full_name=full_name,
                    institution=institution,
                    research_title=research_title,
                    award_category=award_category,
                    award_year=award_year,
                ))
    finally:
        workbook.close()

    if not matching_sheets:
        raise ValidationError(
            "No worksheet has the required columns: Na, NAME, INSTITUTION, "
            "RESEARCH TITLE, AWARD CATEGORY and YEAR."
        )
    if not publication_rows:
        raise ValidationError("No publication rows were found in the Excel file.")
    return publication_rows, skipped, matching_sheets


@transaction.atomic
def import_special_event_participants(*, event, uploaded_file, user):
    """Synchronize publications and preserve one QR token per researcher."""
    if not event.category.is_special_event:
        raise ValidationError("Select an event in the Special Event category.")

    publication_rows, skipped, matching_sheets = _read_publication_rows(uploaded_file)
    grouped_rows = OrderedDict()
    for row in publication_rows:
        identity_key = special_event_researcher_identity(
            row.full_name,
            row.institution,
        )
        grouped_rows.setdefault(identity_key, []).append(row)

    now = timezone.now()
    SpecialEventParticipant.objects.filter(
        event=event,
        is_active=True,
    ).update(is_active=False, updated_by=user, updated_at=now)
    SpecialEventPublication.objects.filter(
        participant__event=event,
        is_active=True,
    ).update(is_active=False, updated_by=user, updated_at=now)

    researchers_created = researchers_updated = 0
    publications_created = publications_updated = 0
    for identity_key, researcher_rows in grouped_rows.items():
        first_row = researcher_rows[0]
        existing_researcher = SpecialEventParticipant.objects.filter(
            event=event,
            identity_key=identity_key,
        ).first()
        participant, was_created = SpecialEventParticipant.objects.update_or_create(
            event=event,
            identity_key=identity_key,
            defaults={
                "full_name": first_row.full_name,
                "institution": first_row.institution,
                "is_active": True,
                "created_by": (
                    existing_researcher.created_by if existing_researcher else user
                ),
                "updated_by": user,
            },
        )
        researchers_created += int(was_created)
        researchers_updated += int(not was_created)

        for row in researcher_rows:
            existing_publication = SpecialEventPublication.objects.filter(
                participant=participant,
                source_sheet=row.source_sheet,
                source_number=row.source_number,
            ).first()
            _, publication_created = SpecialEventPublication.objects.update_or_create(
                participant=participant,
                source_sheet=row.source_sheet,
                source_number=row.source_number,
                defaults={
                    "source_row_index": row.source_row_index,
                    "research_title": row.research_title,
                    "award_category": row.award_category,
                    "award_year": row.award_year,
                    "is_active": True,
                    "created_by": (
                        existing_publication.created_by if existing_publication else user
                    ),
                    "updated_by": user,
                },
            )
            publications_created += int(publication_created)
            publications_updated += int(not publication_created)

    return ImportResult(
        researchers_created=researchers_created,
        researchers_updated=researchers_updated,
        publications_created=publications_created,
        publications_updated=publications_updated,
        skipped=skipped,
        sheets=matching_sheets,
    )

