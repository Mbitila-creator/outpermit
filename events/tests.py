from datetime import timedelta
from io import BytesIO
import shutil
import tempfile
from uuid import uuid4

from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.messages.middleware import MessageMiddleware
from django.contrib.sessions.middleware import SessionMiddleware
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.test import RequestFactory
from django.urls import reverse
from django.utils import timezone
from pypdf import PdfReader
from openpyxl import load_workbook

from core.models import Council, Country, Region
from permits.models import Department

from .access import events_visible_to
from .auth import EventRole
from .models import Event, EventCategory, EventTimetable, Venue
from forms_builder.models import (
    CertificateRecord,
    EventForm,
    FormAnswer,
    FormQuestion,
    FormSection,
    FormSubmission,
    QuestionOption,
)
from checkin.models import ParticipantCheckIn
from conferences.views import _conference_registration_forms
from conferences.models import ConferenceFeedback
from permits.views import system_home


class DepartmentEventAccessTests(TestCase):
    def setUp(self):
        self.dsti = Department.objects.create(
            code="DSTI", name="Department of Science Technology and Innovation"
        )
        self.dhe = Department.objects.create(
            code="DHE", name="Department of Higher Education"
        )
        self.category = EventCategory.objects.create(
            code="CONFERENCE", name_sw="Kongamano", name_en="Conference"
        )
        now = timezone.now()
        self.dsti_event = Event.objects.create(
            owning_department=self.dsti,
            category=self.category,
            code="NESIF-2026",
            title_sw="NESIF",
            title_en="NESIF",
            starts_at=now,
            ends_at=now + timedelta(days=1),
        )
        self.dhe_event = Event.objects.create(
            owning_department=self.dhe,
            category=self.category,
            code="DHE-2026",
            title_sw="DHE Event",
            title_en="DHE Event",
            starts_at=now,
            ends_at=now + timedelta(days=1),
        )

    def _staff(self, username, department):
        user = User.objects.create_user(username=username, password="safe-password")
        user.profile.department = department
        user.profile.save(update_fields=["department"])
        return user

    def test_department_staff_only_see_their_department_events(self):
        dsti_user = self._staff("dsti-user", self.dsti)
        self.assertEqual(list(events_visible_to(dsti_user)), [self.dsti_event])

    def test_other_department_starts_with_empty_workspace(self):
        empty_department = Department.objects.create(code="DPP", name="Policy and Planning")
        user = self._staff("dpp-user", empty_department)
        self.client.force_login(user)
        response = self.client.get(reverse("events:department_event_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No events created")
        self.assertNotContains(response, "NESIF-2026")

    def test_system_administrator_can_see_all_departments(self):
        admin = User.objects.create_superuser(
            username="system-admin", email="admin@example.test", password="safe-password"
        )
        self.assertEqual(events_visible_to(admin).count(), 2)

    def test_public_event_home_opens_without_a_selected_event(self):
        self.dsti_event.is_public = True
        self.dsti_event.save(update_fields=["is_public", "updated_at"])

        response = self.client.get(reverse("events:home"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.dsti_event.title_en)

    def test_conference_registration_workspace_is_department_scoped(self):
        dsti_form = EventForm.objects.create(
            event=self.dsti_event,
            name_sw="Usajili wa DSTI",
            name_en="DSTI registration",
            form_type=EventForm.FormType.REGISTRATION,
        )
        EventForm.objects.create(
            event=self.dhe_event,
            name_sw="Usajili wa DHE",
            name_en="DHE registration",
            form_type=EventForm.FormType.REGISTRATION,
        )
        dsti_user = self._staff("dsti-conference-user", self.dsti)
        self.assertEqual(
            list(_conference_registration_forms(dsti_user)),
            [dsti_form],
        )

    def test_event_administrator_builds_and_publishes_questionnaire(self):
        user = self._staff("questionnaire-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)

        response = self.client.post(
            reverse("forms_builder:questionnaire_create", args=[self.dsti_event.slug]),
            {
                "name_en": "Participant evaluation",
                "name_sw": "Tathmini ya mshiriki",
                "form_type": EventForm.FormType.EVALUATION,
                "introduction_en": "Tell us about the event.",
                "introduction_sw": "Tueleze kuhusu tukio.",
                "success_message_en": "Thank you.",
                "success_message_sw": "Asante.",
                "show_event_summary": "on",
            },
        )
        questionnaire = EventForm.objects.get(name_en="Participant evaluation")
        self.assertRedirects(
            response,
            reverse(
                "forms_builder:questionnaire_builder",
                args=[self.dsti_event.slug, questionnaire.pk],
            ),
        )
        self.assertFalse(questionnaire.is_published)
        self.assertEqual(questionnaire.created_by, user)

        self.client.post(
            reverse(
                "forms_builder:section_create",
                args=[self.dsti_event.slug, questionnaire.pk],
            ),
            {"title_en": "Experience", "title_sw": "Uzoefu"},
        )
        section = questionnaire.sections.get()
        self.client.post(
            reverse(
                "forms_builder:question_create",
                args=[self.dsti_event.slug, questionnaire.pk, section.pk],
            ),
            {
                "label_en": "Were you satisfied?",
                "label_sw": "Umeridhika?",
                "question_type": FormQuestion.QuestionType.SINGLE_CHOICE,
                "is_required": "on",
            },
        )
        question = section.questions.get()
        self.client.post(
            reverse(
                "forms_builder:option_create",
                args=[self.dsti_event.slug, questionnaire.pk, question.pk],
            ),
            {"label_en": "Yes", "label_sw": "Ndiyo", "value": "YES"},
        )
        self.assertTrue(question.options.filter(value="YES").exists())

        submission = FormSubmission.objects.create(
            event_form=questionnaire,
            submitter_email="tracked@example.com",
            badge_name="Tracked Participant",
            is_complete=True,
        )
        tracking_question = section.questions.create(
            label_en="Tracking note",
            label_sw="Dokezo la ufuatiliaji",
            question_type=FormQuestion.QuestionType.SHORT_TEXT,
            display_order=2,
        )
        FormAnswer.objects.create(
            submission=submission,
            question=tracking_question,
            text_value="Distinctive tracked answer",
        )
        submission_list_url = reverse(
            "forms_builder:event_submission_list", args=[self.dsti_event.slug]
        )
        tracker_response = self.client.get(
            submission_list_url,
            {"form": questionnaire.pk, "q": "Distinctive tracked"},
        )
        self.assertContains(tracker_response, submission.reference_number)
        self.assertContains(tracker_response, "Tracked Participant")
        detail_response = self.client.get(reverse(
            "forms_builder:event_submission_detail",
            args=[self.dsti_event.slug, submission.pk],
        ))
        self.assertContains(detail_response, "Distinctive tracked answer")
        self.assertNotContains(detail_response, "Open in Django Admin")
        self.assertContains(detail_response, "Edit")
        self.assertContains(detail_response, "Delete")

        edit_response = self.client.post(reverse(
            "forms_builder:event_submission_edit",
            args=[self.dsti_event.slug, submission.pk],
        ), {
            "badge_name": "Updated Participant",
            "badge_organization": "Updated Organization",
            "badge_title": "Coordinator",
            "submitter_email": "updated@example.com",
            "submitter_phone": "0712345678",
            "review_status": FormSubmission.ReviewStatus.APPROVED,
            "review_notes": "Verified in the event workspace.",
        })
        self.assertRedirects(edit_response, reverse(
            "forms_builder:event_submission_detail",
            args=[self.dsti_event.slug, submission.pk],
        ))
        submission.refresh_from_db()
        self.assertEqual(submission.badge_name, "Updated Participant")
        self.assertEqual(submission.review_status, FormSubmission.ReviewStatus.APPROVED)
        self.assertEqual(submission.reviewed_by, user)

        delete_response = self.client.post(reverse(
            "forms_builder:event_submission_delete",
            args=[self.dsti_event.slug, submission.pk],
        ))
        self.assertRedirects(delete_response, submission_list_url)
        submission.refresh_from_db()
        self.assertFalse(submission.is_active)

        response = self.client.post(reverse(
            "forms_builder:questionnaire_publish",
            args=[self.dsti_event.slug, questionnaire.pk],
        ))
        self.assertRedirects(
            response,
            reverse(
                "forms_builder:questionnaire_builder",
                args=[self.dsti_event.slug, questionnaire.pk],
            ),
        )
        questionnaire.refresh_from_db()
        self.assertTrue(questionnaire.is_published)
        list_response = self.client.get(reverse(
            "forms_builder:questionnaire_list", args=[self.dsti_event.slug]
        ))
        self.assertContains(list_response, "Participant evaluation")
        self.assertContains(list_response, "1 section")
        self.assertContains(list_response, "Form QR code")
        qr_page_response = self.client.get(reverse(
            "forms_builder:questionnaire_qr",
            args=[self.dsti_event.slug, questionnaire.pk],
        ))
        public_form_url = reverse(
            "forms_builder:public_event_form",
            args=[self.dsti_event.slug, questionnaire.slug],
        )
        self.assertContains(qr_page_response, public_form_url)
        self.assertNotContains(qr_page_response, f"{public_form_url}?preview=1")
        self.assertContains(qr_page_response, "Open public form")
        self.assertContains(qr_page_response, "View QR code")
        self.assertContains(qr_page_response, "Download QR code")
        self.assertContains(qr_page_response, "Print QR code")
        qr_image_response = self.client.get(reverse(
            "forms_builder:questionnaire_qr_image",
            args=[self.dsti_event.slug, questionnaire.pk],
        ))
        self.assertEqual(qr_image_response.status_code, 200)
        self.assertEqual(qr_image_response["Content-Type"], "image/png")
        qr_download_response = self.client.get(
            reverse(
                "forms_builder:questionnaire_qr_image",
                args=[self.dsti_event.slug, questionnaire.pk],
            ),
            {"download": "1"},
        )
        self.assertIn("attachment", qr_download_response["Content-Disposition"])
        qr_print_response = self.client.get(reverse(
            "forms_builder:questionnaire_qr_print",
            args=[self.dsti_event.slug, questionnaire.pk],
        ))
        self.assertContains(qr_print_response, "Print / Save as PDF")
        builder_response = self.client.get(reverse(
            "forms_builder:questionnaire_builder",
            args=[self.dsti_event.slug, questionnaire.pk],
        ))
        self.assertContains(builder_response, "Were you satisfied?")
        self.assertContains(builder_response, "Yes / Ndiyo")
        self.assertContains(builder_response, "Print / Save as PDF")

        print_response = self.client.get(reverse(
            "forms_builder:questionnaire_print",
            args=[self.dsti_event.slug, questionnaire.pk],
        ))
        self.assertEqual(print_response.status_code, 200)
        self.assertContains(print_response, "Were you satisfied?")
        self.assertNotContains(print_response, "Umeridhika?")
        self.assertContains(print_response, "Yes")
        self.assertNotContains(print_response, "Ndiyo")
        self.assertContains(print_response, "window.print()")

        sw_print_response = self.client.get(
            reverse(
                "forms_builder:questionnaire_print",
                args=[self.dsti_event.slug, questionnaire.pk],
            ),
            {"language": "sw"},
        )
        self.assertContains(sw_print_response, "Umeridhika?")
        self.assertContains(sw_print_response, "Ndiyo")
        self.assertNotContains(sw_print_response, "Were you satisfied?")

        event_response = self.client.get(reverse(
            "events:department_event_detail", args=[self.dsti_event.slug]
        ))
        self.assertContains(event_response, "Questionnaires and forms")

    def test_event_administrator_can_edit_and_archive_questionnaire_from_list(self):
        user = self._staff("questionnaire-list-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)
        questionnaire = EventForm.objects.create(
            event=self.dsti_event,
            name_en="Form list actions",
            name_sw="Vitendo vya orodha ya fomu",
            is_published=True,
        )
        submission = FormSubmission.objects.create(
            event_form=questionnaire,
            badge_name="Retained participant",
        )

        list_url = reverse(
            "forms_builder:questionnaire_list", args=[self.dsti_event.slug]
        )
        list_response = self.client.get(list_url)
        self.assertContains(list_response, reverse(
            "forms_builder:questionnaire_edit",
            args=[self.dsti_event.slug, questionnaire.pk],
        ))
        self.assertContains(list_response, reverse(
            "forms_builder:questionnaire_delete",
            args=[self.dsti_event.slug, questionnaire.pk],
        ))

        delete_url = reverse(
            "forms_builder:questionnaire_delete",
            args=[self.dsti_event.slug, questionnaire.pk],
        )
        confirmation = self.client.get(delete_url)
        self.assertContains(confirmation, "1 submission")
        response = self.client.post(delete_url)
        self.assertRedirects(response, list_url)
        questionnaire.refresh_from_db()
        self.assertFalse(questionnaire.is_active)
        self.assertFalse(questionnaire.is_published)
        self.assertTrue(FormSubmission.objects.filter(pk=submission.pk).exists())
        self.assertNotContains(self.client.get(list_url), "Form list actions")

    def test_event_administrator_can_preview_an_unpublished_questionnaire(self):
        user = self._staff("draft-preview-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)
        questionnaire = EventForm.objects.create(
            event=self.dsti_event,
            name_en="Draft registration",
            name_sw="Usajili wa rasimu",
            form_type=EventForm.FormType.REGISTRATION,
            is_published=False,
        )
        preview_url = reverse(
            "forms_builder:public_event_form",
            kwargs={
                "event_slug": self.dsti_event.slug,
                "form_slug": questionnaire.slug,
            },
        )

        response = self.client.get(f"{preview_url}?preview=1")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Draft registration")
        self.client.logout()
        self.assertEqual(self.client.get(preview_url).status_code, 404)

    def test_event_administrator_imports_form_aware_excel_records_atomically(self):
        user = self._staff("excel-form-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)
        questionnaire = EventForm.objects.create(
            event=self.dsti_event,
            name_en="Excel individual records",
            name_sw="Rekodi za Excel",
            qr_record_enabled=True,
        )
        section = FormSection.objects.create(
            event_form=questionnaire,
            title_en="Individual",
            title_sw="Mtu binafsi",
        )
        name_question = FormQuestion.objects.create(
            section=section,
            label_en="Full name",
            label_sw="Jina kamili",
            question_type=FormQuestion.QuestionType.SHORT_TEXT,
            display_order=1,
        )
        category_question = FormQuestion.objects.create(
            section=section,
            label_en="Category",
            label_sw="Aina",
            question_type=FormQuestion.QuestionType.DROPDOWN,
            display_order=2,
        )
        gold = QuestionOption.objects.create(
            question=category_question,
            value="GOLD",
            label_en="Gold",
            label_sw="Dhahabu",
        )
        template_url = reverse(
            "forms_builder:questionnaire_excel_template",
            args=[self.dsti_event.slug, questionnaire.pk],
        )
        template_response = self.client.get(template_url)
        self.assertEqual(template_response.status_code, 200)
        self.assertIn("spreadsheetml", template_response["Content-Type"])
        workbook = load_workbook(BytesIO(template_response.content))
        self.assertIn("Records", workbook.sheetnames)
        self.assertIn("Choices", workbook.sheetnames)
        sheet = workbook["Records"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        sheet.cell(2, headers["Badge name"], "Imported Individual")
        sheet.cell(2, headers[f"Full name [q{name_question.pk}]"], "Imported Individual")
        sheet.cell(2, headers[f"Category [q{category_question.pk}]"], "GOLD")
        sheet.cell(3, headers["Badge name"], "Invalid Individual")
        sheet.cell(3, headers[f"Category [q{category_question.pk}]"], "UNKNOWN")
        invalid_file = BytesIO()
        workbook.save(invalid_file)
        import_url = reverse(
            "forms_builder:questionnaire_excel_import",
            args=[self.dsti_event.slug, questionnaire.pk],
        )
        invalid_response = self.client.post(import_url, {
            "excel_file": SimpleUploadedFile(
                "records.xlsx", invalid_file.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        })
        self.assertContains(invalid_response, "Row 3")
        self.assertEqual(questionnaire.submissions.count(), 0)

        sheet.delete_rows(3)
        valid_file = BytesIO()
        workbook.save(valid_file)
        response = self.client.post(import_url, {
            "excel_file": SimpleUploadedFile(
                "records.xlsx", valid_file.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        })
        self.assertRedirects(response, reverse(
            "forms_builder:event_submission_list", args=[self.dsti_event.slug]
        ))
        submission = questionnaire.submissions.get()
        self.assertTrue(submission.is_complete)
        self.assertEqual(submission.badge_name, "Imported Individual")
        self.assertEqual(
            submission.answers.get(question=name_question).text_value,
            "Imported Individual",
        )
        self.assertEqual(
            list(submission.answers.get(question=category_question).selected_options.all()),
            [gold],
        )

    def test_builder_configures_question_skip_logic_with_readable_answer(self):
        user = self._staff("logic-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)
        questionnaire = EventForm.objects.create(
            event=self.dsti_event,
            name_en="Logic form",
            name_sw="Fomu ya mantiki",
        )
        section = FormSection.objects.create(
            event_form=questionnaire,
            title_en="Details",
            title_sw="Maelezo",
            display_order=1,
        )
        controller = FormQuestion.objects.create(
            section=section,
            label_en="Choose type",
            label_sw="Chagua aina",
            question_type=FormQuestion.QuestionType.SINGLE_CHOICE,
            display_order=1,
        )
        QuestionOption.objects.create(
            question=controller,
            value="OTHER",
            label_en="Other",
            label_sw="Nyingine",
        )

        response = self.client.post(
            reverse(
                "forms_builder:question_create",
                args=[self.dsti_event.slug, questionnaire.pk, section.pk],
            ),
            {
                "label_en": "Specify other",
                "label_sw": "Taja nyingine",
                "question_type": FormQuestion.QuestionType.SHORT_TEXT,
                "is_required": "on",
                "condition_question": controller.pk,
                "condition_value": "OTHER",
            },
        )
        self.assertEqual(response.status_code, 302)
        dependent = section.questions.get(label_en="Specify other")
        self.assertEqual(dependent.condition_question, controller)
        self.assertEqual(dependent.condition_value, "OTHER")

    def test_non_event_manager_cannot_open_questionnaire_builder(self):
        user = self._staff("ordinary-department-user", self.dsti)
        self.client.force_login(user)
        response = self.client.get(reverse(
            "forms_builder:questionnaire_list", args=[self.dsti_event.slug]
        ))
        self.assertEqual(response.status_code, 403)

    def test_department_head_creates_event_for_own_department(self):
        user = self._staff("dhe-head", self.dhe)
        user.profile.role = "HEAD_OF_UNIT"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)
        now = timezone.now()
        response = self.client.post(
            reverse("events:department_event_create"),
            {
                "category": self.category.pk,
                "code": "DHE-FORUM-2027",
                "title_sw": "Jukwaa la DHE",
                "title_en": "DHE Forum",
                "starts_at": now.strftime("%Y-%m-%dT%H:%M"),
                "ends_at": (now + timedelta(days=1)).strftime("%Y-%m-%dT%H:%M"),
                "status": Event.Status.DRAFT,
                "payment_currency": "TZS",
            },
        )
        self.assertRedirects(response, reverse("events:department_event_list"))
        created = Event.objects.get(code="DHE-FORUM-2027")
        self.assertEqual(created.owning_department, self.dhe)
        self.assertNotIn(created, events_visible_to(self._staff("dsti-other", self.dsti)))

    def test_event_administrator_can_create_event_with_manually_entered_venue(self):
        user = self._staff("manual-venue-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)
        now = timezone.now()
        country = Country.objects.create(
            name_sw="Tanzania",
            name_en="Tanzania",
            code="TZA",
        )
        region = Region.objects.create(
            country=country,
            name_sw="Dodoma",
            name_en="Dodoma",
            code="DOM",
        )
        council = Council.objects.create(
            region=region,
            name_sw="Jiji la Dodoma",
            name_en="Dodoma City",
            code="DODOMA-CITY",
            council_type=Council.CouncilType.CITY,
        )

        response = self.client.post(
            reverse("events:department_event_create"),
            {
                "category": self.category.pk,
                "venue_mode": "NEW",
                "new_venue_name": "  Ministry   Conference Hall  ",
                "new_venue_country": country.pk,
                "new_venue_region": region.pk,
                "new_venue_council": council.pk,
                "code": "DSTI-HALL-2027",
                "title_sw": "Tukio la DSTI",
                "title_en": "DSTI Hall Event",
                "starts_at": now.strftime("%Y-%m-%dT%H:%M"),
                "ends_at": (now + timedelta(days=1)).strftime("%Y-%m-%dT%H:%M"),
                "status": Event.Status.DRAFT,
                "payment_currency": "TZS",
            },
        )

        self.assertRedirects(response, reverse("events:department_event_list"))
        created = Event.objects.get(code="DSTI-HALL-2027")
        self.assertEqual(created.venue.name, "Ministry Conference Hall")
        self.assertEqual(created.venue.council, council)
        self.assertEqual(created.venue.council.region, region)
        self.assertEqual(created.venue.council.region.country, country)
        self.assertEqual(created.venue.created_by, user)

    def test_event_creation_rejects_existing_and_manual_venue_together(self):
        venue = Venue.objects.create(name="Existing Hall")
        user = self._staff("ambiguous-venue-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)
        now = timezone.now()

        response = self.client.post(
            reverse("events:department_event_create"),
            {
                "category": self.category.pk,
                "venue_mode": "NEW",
                "venue": venue.pk,
                "new_venue_name": "Another Hall",
                "code": "DSTI-AMBIGUOUS-2027",
                "title_sw": "Tukio la DSTI",
                "title_en": "DSTI Ambiguous Event",
                "starts_at": now.strftime("%Y-%m-%dT%H:%M"),
                "ends_at": (now + timedelta(days=1)).strftime("%Y-%m-%dT%H:%M"),
                "status": Event.Status.DRAFT,
                "payment_currency": "TZS",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "Clear the saved venue when creating a new one.",
        )
        self.assertFalse(Event.objects.filter(code="DSTI-AMBIGUOUS-2027").exists())

    def test_direct_staff_url_cannot_bypass_department_ownership(self):
        dsti_form = EventForm.objects.create(
            event=self.dsti_event,
            name_sw="Usajili wa DSTI",
            name_en="DSTI registration",
            form_type=EventForm.FormType.REGISTRATION,
        )
        dhe_user = self._staff("dhe-direct-url", self.dhe)
        dhe_user.profile.role = "DIRECTOR"
        dhe_user.profile.save(update_fields=["role"])
        self.client.force_login(dhe_user)

        response = self.client.get(
            reverse("conferences:conference_detail", kwargs={"form_id": dsti_form.pk})
        )

        self.assertEqual(response.status_code, 403)

    def test_event_management_roles_are_available_to_user_administration(self):
        role_codes = dict(self.dsti.user_profiles.model.ROLE_CHOICES)
        self.assertIn("EVENT_ADMIN", role_codes)
        self.assertIn("REGISTRATION_OFFICER", role_codes)
        self.assertIn("ATTENDANCE_OFFICER", role_codes)
        self.assertIn("REPORT_OFFICER", role_codes)

    def test_registration_officer_maps_to_event_registration_access(self):
        user = self._staff("registration-officer", self.dsti)
        user.profile.role = "REGISTRATION_OFFICER"
        user.profile.save(update_fields=["role"])
        self.assertEqual(user.role, EventRole.REGISTRATION_OFFICER)

    def test_system_home_discards_messages_from_unrelated_workflows(self):
        admin = User.objects.create_superuser(
            username="message-admin",
            email="messages@example.test",
            password="safe-password",
        )
        request = RequestFactory().get("/")
        SessionMiddleware(lambda response: response).process_request(request)
        request.session.save()
        MessageMiddleware(lambda response: response).process_request(request)
        request.user = admin
        messages.success(request, "User account updated successfully.")

        response = system_home(request)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(request._messages.used)
        self.assertNotIn(
            b"User account updated successfully.",
            response.content,
        )

    def test_old_badge_qr_url_redirects_to_integrated_check_in(self):
        participant_token = uuid4()

        response = self.client.get(
            f"/en/check-in/{participant_token}/?auto=1",
        )

        self.assertEqual(response.status_code, 301)
        self.assertEqual(
            response["Location"],
            f"/event-management/check-in/{participant_token}/?auto=1",
        )

    def test_old_participant_portal_url_redirects_with_same_token(self):
        participant_token = uuid4()

        response = self.client.get(f"/sw/participants/{participant_token}/")

        self.assertEqual(response.status_code, 301)
        self.assertEqual(
            response["Location"],
            f"/event-management/participants/{participant_token}/",
        )

    def test_language_prefixed_event_page_redirects_to_public_event(self):
        response = self.client.get(f"/sw/events/{self.dsti_event.slug}/")

        self.assertEqual(response.status_code, 301)
        self.assertEqual(
            response["Location"],
            f"/event-management/events/{self.dsti_event.slug}/",
        )

    def test_public_event_language_change_keeps_canonical_url(self):
        self.dsti_event.is_public = True
        self.dsti_event.title_en = "English public event"
        self.dsti_event.title_sw = "Tukio la umma la Kiswahili"
        self.dsti_event.contact_person = "Afisa wa Tukio"
        self.dsti_event.contact_phone = "0712345678"
        self.dsti_event.status = Event.Status.PUBLISHED
        self.dsti_event.registration_enabled = True
        self.dsti_event.save(update_fields=(
            "is_public",
            "title_en",
            "title_sw",
            "contact_person",
            "contact_phone",
            "status",
            "registration_enabled",
            "updated_at",
        ))
        EventForm.objects.create(
            event=self.dsti_event,
            name_sw="Fomu ya Usajili",
            name_en="Registration Form",
            form_type=EventForm.FormType.REGISTRATION,
            is_active=True,
            is_published=True,
            requires_participant_registration=True,
        )
        event_url = reverse("events:event_detail", args=(self.dsti_event.slug,))

        response = self.client.post(
            reverse("set_language"),
            {"language": "sw", "next": event_url},
            follow=True,
        )

        self.assertRedirects(response, event_url)
        self.assertContains(response, "Tukio la umma la Kiswahili")
        self.assertNotContains(response, "English public event")
        for translated_label in (
            "Kuhusu Tukio",
            "Fomu ya Tukio",
            "Taarifa muhimu",
            "Mtu wa kuwasiliana naye",
            "Simu",
            "Jisajili Sasa",
            "Hali ya Usajili",
            "Ingia kama mtumishi",
        ):
            self.assertContains(response, translated_label)
        for english_label in (
            "About the event",
            "Event forms",
            "Preview questions",
            "Important information",
            "Contact person",
            "Register now",
            "Registration Status",
            "Staff login",
        ):
            self.assertNotContains(response, english_label)

        user = self._staff("public-header-user", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)
        authenticated_response = self.client.get(event_url)
        self.assertContains(authenticated_response, "Angalia maswali")
        self.assertContains(authenticated_response, "Badilisha nywila")
        self.assertContains(authenticated_response, "Eneo langu la kazi")
        self.assertContains(authenticated_response, "Toka kwenye akaunti")

    def test_conference_evaluation_is_available_in_participant_portal(self):
        registration_form = EventForm.objects.create(
            event=self.dsti_event,
            name_sw="Usajili wa NESIF",
            name_en="NESIF registration",
            form_type=EventForm.FormType.REGISTRATION,
            is_active=True,
        )
        submission = FormSubmission.objects.create(
            event_form=registration_form,
            submitter_email="participant@example.test",
            is_complete=True,
            is_active=True,
        )

        response = self.client.get(
            reverse(
                "forms_builder:participant_portal",
                kwargs={"participant_token": submission.participant_token},
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            reverse(
                "conferences:feedback_submit",
                kwargs={"event_slug": self.dsti_event.slug},
            ),
        )
        self.assertContains(response, "Conference feedback and evaluation")

    def test_conference_evaluation_dashboard_uses_statistical_question_analysis(self):
        registration_form = EventForm.objects.create(
            event=self.dsti_event,
            name_sw="Usajili wa kongamano",
            name_en="Conference registration",
            form_type=EventForm.FormType.REGISTRATION,
            is_active=True,
        )
        for rating, recommend in ((5, True), (3, False)):
            ConferenceFeedback.objects.create(
                event=self.dsti_event,
                overall_rating=rating,
                content_rating=rating,
                speakers_rating=rating,
                organization_rating=rating,
                venue_rating=rating,
                would_recommend=recommend,
            )
        administrator = User.objects.create_superuser(
            username="conference-analysis-admin",
            email="conference-analysis@example.test",
            password="safe-password",
        )
        self.client.force_login(administrator)

        response = self.client.get(reverse(
            "conferences:feedback_dashboard",
            kwargs={"form_id": registration_form.pk},
        ))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["evaluation_statistics"]), 6)
        self.assertEqual(
            response.context["evaluation_statistics"][0]["rows"][4]["count"],
            1,
        )
        self.assertContains(response, "Conference event evaluation")
        self.assertContains(response, "Analysis by evaluation question")
        self.assertContains(response, "Evaluation responses")
        self.assertNotContains(response, "Feedback responses")

    def test_event_without_evaluation_offers_evaluation_setup(self):
        seminar_category = EventCategory.objects.create(
            code="SEMINAR",
            name_sw="Semina",
            name_en="Seminar",
        )
        event = Event.objects.create(
            owning_department=self.dsti,
            category=seminar_category,
            code="SEM-EVAL-2026",
            title_sw="Semina ya tathmini",
            title_en="Evaluation seminar",
            starts_at=timezone.now(),
            ends_at=timezone.now() + timedelta(days=1),
        )
        user = self._staff("seminar-evaluation-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)

        response = self.client.get(reverse(
            "events:department_event_detail",
            kwargs={"event_slug": event.slug},
        ))

        self.assertContains(response, "Set up event evaluation")
        self.assertContains(
            response,
            reverse("forms_builder:questionnaire_list", kwargs={"event_slug": event.slug}),
        )

    def test_registration_remains_available_while_event_is_ongoing(self):
        now = timezone.now()
        self.dsti_event.starts_at = now - timedelta(hours=1)
        self.dsti_event.ends_at = now + timedelta(hours=2)
        self.dsti_event.status = Event.Status.ONGOING
        self.dsti_event.is_public = True
        self.dsti_event.registration_enabled = True
        self.dsti_event.save()
        registration_form = EventForm.objects.create(
            event=self.dsti_event,
            name_sw="Usajili unaoendelea",
            name_en="Ongoing registration",
            form_type=EventForm.FormType.REGISTRATION,
            is_active=True,
            is_published=True,
        )

        response = self.client.get(reverse(
            "events:event_detail",
            kwargs={"event_slug": self.dsti_event.slug},
        ))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["registration_available"])
        self.assertContains(
            response,
            reverse(
                "forms_builder:public_event_form",
                kwargs={
                    "event_slug": self.dsti_event.slug,
                    "form_slug": registration_form.slug,
                },
            ),
        )

    def test_registration_closes_when_event_ends(self):
        now = timezone.now()
        self.dsti_event.starts_at = now - timedelta(hours=2)
        self.dsti_event.ends_at = now - timedelta(minutes=1)
        self.dsti_event.status = Event.Status.ONGOING
        self.dsti_event.is_public = True
        self.dsti_event.registration_enabled = True
        self.dsti_event.save()
        registration_form = EventForm.objects.create(
            event=self.dsti_event,
            name_sw="Usajili uliofungwa",
            name_en="Ended registration",
            form_type=EventForm.FormType.REGISTRATION,
            is_active=True,
            is_published=True,
        )

        detail_response = self.client.get(reverse(
            "events:event_detail",
            kwargs={"event_slug": self.dsti_event.slug},
        ))
        form_response = self.client.post(reverse(
            "forms_builder:public_event_form",
            kwargs={
                "event_slug": self.dsti_event.slug,
                "form_slug": registration_form.slug,
            },
        ))

        self.assertFalse(detail_response.context["registration_available"])
        self.assertTrue(detail_response.context["registration_closed"])
        self.assertEqual(form_response.status_code, 400)
        self.assertEqual(
            form_response.json()["message"],
            "The submission period for this form has ended.",
        )

    def test_registration_can_be_configured_to_close_after_event_starts(self):
        now = timezone.now()
        self.dsti_event.starts_at = now
        self.dsti_event.ends_at = now + timedelta(hours=3)
        self.dsti_event.registration_closes_at = now + timedelta(hours=2)

        self.dsti_event.full_clean()

    def test_event_operations_are_inside_selected_event_workspace(self):
        user = self._staff("event-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        registration_form = EventForm.objects.create(
            event=self.dsti_event,
            name_sw="Usajili wa NESIF",
            name_en="NESIF registration",
            form_type=EventForm.FormType.REGISTRATION,
            is_active=True,
        )
        evaluation_form = EventForm.objects.create(
            event=self.dsti_event,
            name_sw="Tathmini ya NESIF",
            name_en="NESIF evaluation",
            form_type=EventForm.FormType.EVALUATION,
            is_active=True,
        )
        self.client.force_login(user)

        list_response = self.client.get(reverse("events:department_event_list"))

        self.assertEqual(list_response.status_code, 200)
        self.assertNotContains(list_response, "Your event operations")
        self.assertNotContains(list_response, reverse("checkin:lookup"))
        self.assertContains(
            list_response,
            reverse(
                "events:department_event_detail",
                kwargs={"event_slug": self.dsti_event.slug},
            ),
        )

        response = self.client.get(reverse(
            "events:department_event_detail",
            kwargs={"event_slug": self.dsti_event.slug},
        ))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.dsti_event.code)
        self.assertContains(
            response,
            reverse(
                "conferences:conference_detail",
                kwargs={"form_id": registration_form.pk},
            ),
        )
        self.assertContains(response, f"{reverse('checkin:lookup')}?event={self.dsti_event.pk}")
        self.assertContains(response, f"{reverse('checkin:reports')}?event={self.dsti_event.pk}")
        self.assertContains(
            response,
            f"{reverse('forms_builder:evaluation_reports')}?form={evaluation_form.pk}",
        )
        self.assertNotContains(response, reverse("meetings:meeting_list"))

    def test_event_admin_uploads_and_shares_stable_public_timetable(self):
        user = self._staff("timetable-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.dsti_event.is_public = True
        self.dsti_event.save(update_fields=["is_public", "updated_at"])
        self.client.force_login(user)
        media_root = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, media_root, True)

        with override_settings(MEDIA_ROOT=media_root, PUBLIC_BASE_URL="https://events.test"):
            upload_response = self.client.post(
                reverse("events:department_event_timetable", kwargs={
                    "event_slug": self.dsti_event.slug,
                }),
                {
                    "title_sw": "Ratiba ya kilele",
                    "title_en": "Climax timetable",
                    "pdf_file": SimpleUploadedFile(
                        "ratiba.pdf", b"%PDF-1.4\n%%EOF", "application/pdf"
                    ),
                    "is_published": "on",
                },
            )
            timetable = EventTimetable.objects.get(event=self.dsti_event)
            public_url = reverse("events:public_event_timetable", kwargs={
                "event_slug": self.dsti_event.slug,
                "public_token": timetable.public_token,
            })
            download_url = reverse("events:public_event_timetable_download", kwargs={
                "event_slug": self.dsti_event.slug,
                "public_token": timetable.public_token,
            })

            self.assertRedirects(
                upload_response,
                reverse("events:department_event_timetable", kwargs={
                    "event_slug": self.dsti_event.slug,
                }),
            )
            public_response = self.client.get(public_url)
            self.assertContains(public_response, "Climax timetable")
            self.assertContains(public_response, "View timetable")
            self.assertContains(public_response, "Download timetable")
            self.assertNotContains(public_response, "Registration Status")
            self.assertNotContains(public_response, "Event information")
            self.assertEqual(self.client.get(download_url).status_code, 200)
            qr_response = self.client.get(reverse(
                "events:public_event_timetable_qr",
                kwargs={
                    "event_slug": self.dsti_event.slug,
                    "public_token": timetable.public_token,
                },
            ))
            self.assertEqual(qr_response.status_code, 200)
            self.assertEqual(qr_response["Content-Type"], "image/png")

    def test_event_workspace_cannot_bypass_department_ownership(self):
        user = self._staff("dsti-workspace-user", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)

        response = self.client.get(reverse(
            "events:department_event_detail",
            kwargs={"event_slug": self.dhe_event.slug},
        ))

        self.assertEqual(response.status_code, 404)

    def test_event_administrator_can_open_safe_elimu_certificate_preview(self):
        self.dsti_event.code = "WEUUTz-2026"
        self.dsti_event.title_en = "Education and Innovation Week"
        self.dsti_event.save()
        self.assertEqual(self.dsti_event.code, "WEUUTz-2026")
        user = self._staff("elimu-preview-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)

        detail_response = self.client.get(reverse(
            "events:department_event_detail",
            kwargs={"event_slug": self.dsti_event.slug},
        ))
        preview_url = reverse(
            "events:department_certificate_preview",
            kwargs={"event_slug": self.dsti_event.slug},
        )
        preview_response = self.client.get(preview_url)

        self.assertContains(detail_response, preview_url)
        self.assertEqual(preview_response.status_code, 200)
        self.assertContains(preview_response, "SAMPLE · NOT VALID")
        self.assertContains(preview_response, "Sample Participating Institution")
        self.assertContains(preview_response, "Prof. Carolyne I. Nombo")
        self.assertContains(
            preview_response,
            "images/weuutz-permanent-secretary-signature.png",
        )
        self.assertContains(preview_response, "CERTIFICATION OF PARTICIPATION")
        self.assertContains(preview_response, "SAMPLE QR - NOT VALID")
        self.assertContains(preview_response, "data:image/png;base64,")
        self.assertNotContains(preview_response, "Download PDF certificate")

    def test_exhibition_workspace_links_participants_and_certificate_review(self):
        self.category.code = "EXHIBITION"
        self.category.name_en = "Exhibition"
        self.category.name_sw = "Maonesho"
        self.category.slug = "exhibition"
        self.category.save(update_fields=[
            "code", "name_en", "name_sw", "slug", "updated_at",
        ])
        self.dsti_event.certificate_enabled = True
        self.dsti_event.badge_enabled = True
        self.dsti_event.save(update_fields=[
            "certificate_enabled", "badge_enabled", "updated_at",
        ])
        EventForm.objects.create(
            event=self.dsti_event,
            name_sw="Usajili wa Waoneshaji",
            name_en="Exhibitor registration",
            form_type=EventForm.FormType.EXHIBITOR,
            is_active=True,
        )
        user = self._staff("exhibition-event-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)

        response = self.client.get(reverse(
            "events:department_event_detail",
            kwargs={"event_slug": self.dsti_event.slug},
        ))

        reports_url = reverse("checkin:reports")
        self.assertContains(response, "Registration and Certificate approval")
        self.assertContains(
            response,
            f"{reports_url}?event={self.dsti_event.pk}&amp;filter=registration_certificates",
        )
        self.assertContains(response, "Certificate list")
        self.assertContains(
            response,
            f"{reports_url}?event={self.dsti_event.pk}&amp;filter=certificate_list",
        )

    def test_certificate_operations_are_available_without_a_registration_form(self):
        self.dsti_event.certificate_enabled = True
        self.dsti_event.save(update_fields=["certificate_enabled", "updated_at"])
        user = self._staff("generic-certificate-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)

        response = self.client.get(reverse(
            "events:department_event_detail",
            kwargs={"event_slug": self.dsti_event.slug},
        ))

        reports_url = reverse("checkin:reports")
        self.assertContains(response, "Registration and Certificate approval")
        self.assertContains(response, "Certificate list")
        self.assertContains(
            response,
            f"{reports_url}?event={self.dsti_event.pk}&amp;filter=registration_certificates",
        )
        self.assertContains(
            response,
            f"{reports_url}?event={self.dsti_event.pk}&amp;filter=certificate_list",
        )

    def test_event_administrator_can_authorize_checked_in_certificate(self):
        self.dsti_event.certificate_enabled = True
        self.dsti_event.badge_enabled = True
        self.dsti_event.save(update_fields=[
            "certificate_enabled", "badge_enabled", "updated_at",
        ])
        registration_form = EventForm.objects.create(
            event=self.dsti_event,
            name_sw="Usajili",
            name_en="Registration",
            form_type=EventForm.FormType.EXHIBITOR,
            is_active=True,
        )
        submission = FormSubmission.objects.create(
            event_form=registration_form,
            is_complete=True,
            review_status=FormSubmission.ReviewStatus.APPROVED,
        )
        user = self._staff("certificate-event-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        ParticipantCheckIn.objects.create(
            submission=submission,
            checked_in_by=user,
            method=ParticipantCheckIn.Method.MANUAL,
        )
        self.client.force_login(user)

        response = self.client.post(
            reverse("checkin:reports"),
            {
                "event": self.dsti_event.pk,
                "filter": "certificate_review",
                "submission": [submission.pk],
            },
        )

        self.assertRedirects(
            response,
            f"{reverse('checkin:reports')}?event={self.dsti_event.pk}"
            "&filter=certificate_review",
        )
        record = CertificateRecord.objects.get(submission=submission)
        self.assertEqual(record.status, CertificateRecord.Status.AUTHORIZED)
        self.assertEqual(record.authorized_by, user)

        list_response = self.client.get(
            reverse("checkin:reports"),
            {"event": self.dsti_event.pk, "filter": "all"},
        )
        self.assertContains(list_response, "Registered participants")
        self.assertContains(list_response, "Details")
        self.assertContains(
            list_response,
            reverse(
                "checkin:participant_staff_detail",
                kwargs={"submission_id": submission.pk},
            ),
        )
        self.assertContains(list_response, "Badge / QR")
        self.assertContains(list_response, "Print A4 list")
        self.assertContains(list_response, "Download Excel")
        self.assertContains(
            list_response,
            "Certificate list, filters and bulk PDF",
        )
        self.assertContains(
            list_response,
            f"?event={self.dsti_event.pk}&amp;filter=certificate_list",
        )

        certificate_list_response = self.client.get(
            reverse("checkin:reports"),
            {"event": self.dsti_event.pk, "filter": "certificate_list"},
        )
        self.assertContains(certificate_list_response, "Search")
        self.assertContains(certificate_list_response, "All institutions")
        self.assertContains(certificate_list_response, "Select all")
        self.assertContains(
            certificate_list_response,
            "Download selected certificates PDF",
        )

        bulk_pdf_response = self.client.post(
            reverse("checkin:certificate_bulk_pdf"),
            {"event": self.dsti_event.pk, "submission": [submission.pk]},
        )
        self.assertEqual(bulk_pdf_response.status_code, 200)
        self.assertEqual(bulk_pdf_response["Content-Type"], "application/pdf")
        self.assertIn(
            f'{self.dsti_event.code}-selected-certificates.pdf',
            bulk_pdf_response["Content-Disposition"],
        )
        self.assertEqual(
            len(PdfReader(BytesIO(bulk_pdf_response.content)).pages),
            1,
        )

        filtered_certificate_response = self.client.get(
            reverse("checkin:reports"),
            {
                "event": self.dsti_event.pk,
                "filter": "certificate_list",
                "q": submission.reference_number,
            },
        )
        self.assertContains(
            filtered_certificate_response,
            submission.reference_number,
        )

        print_response = self.client.get(
            reverse("checkin:participant_list_print"),
            {"event": self.dsti_event.pk},
        )
        self.assertEqual(print_response.status_code, 200)
        self.assertContains(print_response, submission.reference_number)
        self.assertContains(print_response, "Certificate number")
        self.assertContains(print_response, record.certificate_number)
        self.assertNotContains(print_response, ">Attendance</th>")

        excel_response = self.client.get(
            reverse("checkin:participant_list_excel"),
            {"event": self.dsti_event.pk},
        )
        self.assertEqual(excel_response.status_code, 200)
        self.assertEqual(
            excel_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        detail_url = reverse(
            "checkin:participant_staff_detail",
            kwargs={"submission_id": submission.pk},
        )
        detail_response = self.client.get(detail_url)
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, "Participant details")
        self.assertContains(detail_response, "View certificate")
        self.assertContains(detail_response, "Revoke certificate")
        certificate_pdf_url = reverse(
            "forms_builder:participant_certificate_pdf",
            kwargs={"participant_token": submission.participant_token},
        )
        self.assertContains(
            detail_response,
            f'{certificate_pdf_url}?view=1',
        )
        inline_pdf_response = self.client.get(
            certificate_pdf_url,
            {"view": "1"},
        )
        self.assertEqual(inline_pdf_response.status_code, 200)
        self.assertEqual(
            inline_pdf_response["Content-Type"],
            "application/pdf",
        )
        self.assertTrue(
            inline_pdf_response["Content-Disposition"].startswith("inline;")
        )

        revoke_response = self.client.post(detail_url, {
            "action": "revoke",
            "reason": "Certificate issued in error",
        })
        self.assertRedirects(revoke_response, detail_url)
        record.refresh_from_db()
        self.assertEqual(record.status, CertificateRecord.Status.REVOKED)
        self.assertEqual(record.revocation_reason, "Certificate issued in error")

    def test_certificate_authorization_waits_for_check_in(self):
        self.dsti_event.certificate_enabled = True
        self.dsti_event.save(update_fields=["certificate_enabled", "updated_at"])
        registration_form = EventForm.objects.create(
            event=self.dsti_event,
            name_sw="Usajili",
            name_en="Registration",
            form_type=EventForm.FormType.EXHIBITOR,
            is_active=True,
        )
        submission = FormSubmission.objects.create(
            event_form=registration_form,
            is_complete=True,
            review_status=FormSubmission.ReviewStatus.APPROVED,
        )
        user = self._staff("pre-checkin-certificate-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)
        detail_url = reverse(
            "checkin:participant_staff_detail",
            kwargs={"submission_id": submission.pk},
        )

        detail_response = self.client.get(detail_url)
        self.assertContains(
            detail_response,
            "Certificate approval will become available after this participant checks in.",
        )
        authorize_response = self.client.post(detail_url, {"action": "authorize"})
        self.assertRedirects(authorize_response, detail_url)
        self.assertFalse(
            CertificateRecord.objects.filter(submission=submission).exists()
        )

    def test_registration_officer_cannot_open_certificate_preview(self):
        self.dsti_event.code = "WEUUTz-2026"
        self.dsti_event.save()
        user = self._staff("elimu-preview-registration", self.dsti)
        user.profile.role = "REGISTRATION_OFFICER"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)

        response = self.client.get(reverse(
            "events:department_certificate_preview",
            kwargs={"event_slug": self.dsti_event.slug},
        ))

        self.assertEqual(response.status_code, 403)

    def test_event_administrator_special_events_are_department_scoped(self):
        special_category = EventCategory.objects.create(
            code="SPECIAL_EVENT",
            name_sw="Tukio maalum",
            name_en="Special Event",
            slug="special-event",
        )
        dsti_special = Event.objects.create(
            owning_department=self.dsti,
            category=special_category,
            code="TUZO-TEST",
            title_sw="Tuzo",
            title_en="Awards",
            starts_at=timezone.now(),
            ends_at=timezone.now() + timedelta(days=1),
        )
        user = self._staff("dsti-event-admin", self.dsti)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)

        response = self.client.get(
            reverse("events:special_event_participant_list")
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, dsti_special.code)
        self.assertNotContains(response, self.dhe_event.code)

        dsti_detail = self.client.get(reverse(
            "events:department_event_detail", args=[dsti_special.slug]
        ))
        self.assertContains(dsti_detail, "DSTI researcher QR records")
        self.assertContains(dsti_detail, "Individual QR records")

        dhe_special = Event.objects.create(
            owning_department=self.dhe,
            category=special_category,
            code="DHE-SPECIAL",
            title_sw="Tukio la DHE",
            title_en="DHE Special Event",
            starts_at=timezone.now(),
            ends_at=timezone.now() + timedelta(days=1),
        )
        dhe_user = self._staff("dhe-event-admin", self.dhe)
        dhe_user.profile.role = "EVENT_ADMIN"
        dhe_user.profile.save(update_fields=["role"])
        self.client.force_login(dhe_user)
        dhe_detail = self.client.get(reverse(
            "events:department_event_detail", args=[dhe_special.slug]
        ))
        self.assertContains(dhe_detail, "Individual QR records")
        self.assertNotContains(dhe_detail, "DSTI researcher QR records")
        self.assertNotContains(dhe_detail, "Manage researchers, publications")

    def test_custom_individual_qr_records_require_explicit_form_opt_in(self):
        user = self._staff("custom-qr-admin", self.dhe)
        user.profile.role = "EVENT_ADMIN"
        user.profile.save(update_fields=["role"])
        self.client.force_login(user)
        qr_form = EventForm.objects.create(
            event=self.dhe_event,
            name_en="Custom individual record",
            name_sw="Rekodi binafsi maalum",
            qr_record_enabled=True,
        )
        submission = FormSubmission.objects.create(
            event_form=qr_form,
            badge_name="Custom Person",
            badge_organization="Custom Department",
            is_complete=True,
        )

        list_response = self.client.get(reverse(
            "forms_builder:individual_qr_record_list",
            args=[self.dhe_event.slug],
        ))
        self.assertContains(list_response, "Custom individual record")
        self.assertContains(list_response, "Custom Person")
        record_url = reverse(
            "forms_builder:individual_qr_record",
            args=[submission.participant_token],
        )
        self.client.logout()
        record_response = self.client.get(record_url)
        self.assertContains(record_response, "Custom Person")
        qr_response = self.client.get(reverse(
            "forms_builder:individual_qr_record_qr",
            args=[submission.participant_token],
        ))
        self.assertEqual(qr_response.status_code, 200)
        self.assertEqual(qr_response["Content-Type"], "image/png")

        qr_form.qr_record_enabled = False
        qr_form.save(update_fields=["qr_record_enabled", "updated_at"])
        self.assertEqual(self.client.get(record_url).status_code, 404)
