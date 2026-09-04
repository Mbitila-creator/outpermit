from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from core.workflow import (
    approve_current_finance_step,
    get_finance_review_context,
    initialize_finance_approval_steps,
    reject_current_finance_step,
    return_current_finance_step,
)
from permits.executive_scope import executive_department_codes, is_executive_viewer
from permits.models import Department, ModuleRoleAssignment
from permits.module_roles import module_role

from .forms import (
    BudgetLineForm,
    FinanceRequestForm,
    MinuteSheetForm,
    RetirementForm,
)
from .models import (
    BudgetLine,
    Disbursement,
    FinanceDocument,
    FinanceRequest,
    MinuteSheet,
    Retirement,
)


MONTH_ORDER = [
    "JULY",
    "AUGUST",
    "SEPTEMBER",
    "OCTOBER",
    "NOVEMBER",
    "DECEMBER",
    "JANUARY",
    "FEBRUARY",
    "MARCH",
    "APRIL",
    "MAY",
    "JUNE",
]


def get_user_profile(user):
    if not user or not getattr(user, "is_authenticated", False):
        return None

    profile = getattr(user, "profile", None)

    if profile is None:
        profile = getattr(user, "userprofile", None)

    return profile


def get_user_role(user):
    if user.is_superuser:
        return "ADMIN"

    assigned_role = module_role(
        user,
        ModuleRoleAssignment.Module.FINANCE,
        priority=("DIVISION_BUDGET_OFFICER", "ACCOUNTANT"),
    )
    if assigned_role:
        return assigned_role

    profile = get_user_profile(user)

    if profile:
        approval_role = getattr(
            profile,
            "approval_role",
            None,
        )

        approval_role_code = getattr(
            approval_role,
            "code",
            None,
        )

        if approval_role_code:
            role_code = (
                approval_role_code
                .strip()
                .upper()
            )

            if role_code == "SYSTEM_ADMIN":
                return "ADMIN"

            return role_code

        legacy_role = getattr(
            profile,
            "role",
            None,
        )

        if legacy_role:
            role_code = (
                legacy_role
                .strip()
                .upper()
            )

            if role_code == "SYSTEM_ADMIN":
                return "ADMIN"

            return role_code

    groups = {
        name.strip().upper()
        for name in user.groups.values_list(
            "name",
            flat=True,
        )
        if name
    }

    group_priority = [
        "SYSTEM_ADMIN",
        "ADMIN",
        "DIVISION_BUDGET_OFFICER",
        "ACCOUNTANT",
        "DIRECTOR",
        "ASSISTANT_DIRECTOR",
        "HEAD_OF_UNIT",
        "REQUESTER",
    ]

    for group_name in group_priority:
        if group_name in groups:
            if group_name == "SYSTEM_ADMIN":
                return "ADMIN"

            return group_name

    return "STAFF"


def get_role_display_name(role_code):
    role_labels = {
        "ADMIN": "System Administrator",
        "SYSTEM_ADMIN": "System Administrator",
        "REQUESTER": "Requester",
        "STAFF": "Staff",
        "HEAD_OF_UNIT": "Head of Unit",
        "ASSISTANT_DIRECTOR": "Assistant Director",
        "DIRECTOR": "Director",
        "DIVISION_BUDGET_OFFICER": (
            "Division Budget Officer"
        ),
        "ACCOUNTANT": "Accountant",
    }

    normalized_role = (
        role_code or "STAFF"
    ).strip().upper()

    return role_labels.get(
        normalized_role,
        normalized_role.replace(
            "_",
            " ",
        ).title(),
    )


def get_user_department(user):
    profile = get_user_profile(user)

    if not profile:
        return None

    return getattr(
        profile,
        "department",
        None,
    )


def get_user_department_unit(user):
    profile = get_user_profile(user)

    if not profile:
        return None

    return getattr(
        profile,
        "department_unit",
        None,
    )


def get_user_unit(user):
    department_unit = get_user_department_unit(user)

    if department_unit:
        return department_unit.name

    return ""


def can_see_all_finance(user):
    """
    Users who may access broad Finance management information.

    Department-level filtering should still be applied separately so that
    departmental officers do not see records from other departments.
    """
    return get_user_role(user) in [
        "ADMIN",
        "DIVISION_BUDGET_OFFICER",
        "ACCOUNTANT",
    ]


def is_director(user):
    return get_user_role(user) == "DIRECTOR"


def is_assistant_director(user):
    return (
        get_user_role(user)
        == "ASSISTANT_DIRECTOR"
    )


def is_hou(user):
    return (
        get_user_role(user)
        == "HEAD_OF_UNIT"
    )


def is_final_approver(user):
    """
    The Director is the final approver for Minute Sheets.

    Assistant Directors review and forward but do not give final approval.
    """
    return is_director(user)


def is_dbo(user):
    return (
        get_user_role(user)
        == "DIVISION_BUDGET_OFFICER"
    )


def is_accountant(user):
    return (
        get_user_role(user)
        == "ACCOUNTANT"
    )


def is_retirement_director(user):
    """
    Only a departmental Director or System Administrator may give
    final approval to a Retirement.
    """
    return (
        user.is_superuser
        or get_user_role(user) in [
            "ADMIN",
            "DIRECTOR",
        ]
    )


def can_view_finance_reports(user):
    return is_executive_viewer(user) or get_user_role(user) in [
        "ADMIN",
        "DIVISION_BUDGET_OFFICER",
        "ACCOUNTANT",
        "HEAD_OF_UNIT",
        "ASSISTANT_DIRECTOR",
        "DIRECTOR",
    ]


def apply_executive_finance_scope(queryset, user, department_field="department"):
    """Limit read-only executive visibility to the official reporting chain."""
    if not is_executive_viewer(user):
        return queryset
    department_codes = executive_department_codes(user)
    if department_codes is None:
        return queryset
    return queryset.filter(**{f"{department_field}__code__in": department_codes})


def executive_finance_departments(user):
    if not is_executive_viewer(user):
        return Department.objects.none()
    departments = Department.objects.filter(is_active=True)
    department_codes = executive_department_codes(user)
    if department_codes is not None:
        departments = departments.filter(code__in=department_codes)
    return departments.order_by("code")


def get_current_financial_year(today=None):
    today = today or date.today()

    if today.month >= 7:
        return f"{today.year}/{today.year + 1}"

    return f"{today.year - 1}/{today.year}"


def get_next_financial_year(today=None):
    today = today or date.today()

    if today.month >= 7:
        start_year = today.year + 1
    else:
        start_year = today.year

    return f"{start_year}/{start_year + 1}"


def get_next_budget_month(month):
    month = (month or "").upper()

    if month not in MONTH_ORDER:
        return None

    index = MONTH_ORDER.index(month)

    if index == len(MONTH_ORDER) - 1:
        return None

    return MONTH_ORDER[index + 1]


def get_unit_initials(unit_name):
    if not unit_name:
        return "MS"

    words = unit_name.replace("-", " ").replace("_", " ").split()
    initials = "".join(word[0].upper() for word in words if word)

    return initials or "MS"


def generate_minute_sheet_reference(user):
    department = get_user_department(user)
    department_unit = get_user_department_unit(user)

    if department_unit and getattr(department_unit, "code", None):
        owner_code = department_unit.code.upper()
    elif department and getattr(department, "code", None):
        owner_code = department.code.upper()
    else:
        owner_code = get_unit_initials(get_user_unit(user))

    prefix = f"{owner_code}-MS"

    last_reference = (
        MinuteSheet.objects.filter(reference_no__startswith=f"{prefix}-")
        .order_by("-id")
        .values_list("reference_no", flat=True)
        .first()
    )

    next_number = 1

    if last_reference:
        try:
            next_number = int(last_reference.rsplit("-", 1)[1]) + 1
        except (IndexError, TypeError, ValueError):
            next_number = (
                MinuteSheet.objects.filter(
                    reference_no__startswith=f"{prefix}-"
                ).count()
                + 1
            )

    return f"{prefix}-{next_number:04d}"


def apply_user_ownership(instance, user):
    instance.department = get_user_department(user)
    instance.department_unit = get_user_department_unit(user)


def get_visible_minute_sheets(user):
    base_qs = MinuteSheet.objects.select_related(
        "requested_by",
        "department",
        "department_unit",
        "budget_line",
    )

    # System Administrator may see all departments.
    if user.is_superuser or get_user_role(user) in [
        "ADMIN",
        "SYSTEM_ADMIN",
    ]:
        return base_qs.order_by("-created_at")

    if is_executive_viewer(user):
        return apply_executive_finance_scope(base_qs, user).order_by("-created_at")

    department = get_user_department(user)
    department_unit = get_user_department_unit(user)

    # Department-level viewers see all Minute Sheets in their department.
    if (
        is_director(user)
        or is_assistant_director(user)
        or is_dbo(user)
        or is_accountant(user)
    ):
        if not department:
            return base_qs.none()

        return base_qs.filter(
            department=department
        ).order_by("-created_at")

    # Head of Unit sees their unit's Minute Sheets.
    if is_hou(user):
        if department_unit:
            return base_qs.filter(
                Q(requested_by=user)
                | Q(department_unit=department_unit)
            ).distinct().order_by("-created_at")

        if department:
            return base_qs.filter(
                Q(requested_by=user)
                | Q(
                    department=department,
                    department_unit__isnull=True,
                )
            ).distinct().order_by("-created_at")

    # Ordinary requester sees only their own Minute Sheets.
    return base_qs.filter(
        requested_by=user
    ).order_by("-created_at")


def get_visible_finance_requests(user):
    base_qs = FinanceRequest.objects.select_related(
        "submitted_by",
        "department",
        "department_unit",
        "minute_sheet",
        "minute_sheet__requested_by",
    )

    # System Administrator may see all departments.
    if user.is_superuser or get_user_role(user) in [
        "ADMIN",
        "SYSTEM_ADMIN",
    ]:
        return base_qs.order_by("-created_at")

    if is_executive_viewer(user):
        return apply_executive_finance_scope(base_qs, user).order_by("-created_at")

    department = get_user_department(user)
    department_unit = get_user_department_unit(user)

    # Department-level viewers see requests from their department.
    if (
        is_director(user)
        or is_assistant_director(user)
        or is_dbo(user)
        or is_accountant(user)
    ):
        if not department:
            return base_qs.none()

        return base_qs.filter(
            department=department
        ).order_by("-created_at")

    # Head of Unit sees requests from their unit.
    if is_hou(user):
        if department_unit:
            return base_qs.filter(
                Q(submitted_by=user)
                | Q(department_unit=department_unit)
            ).distinct().order_by("-created_at")

        if department:
            return base_qs.filter(
                Q(submitted_by=user)
                | Q(
                    department=department,
                    department_unit__isnull=True,
                )
            ).distinct().order_by("-created_at")

    # Ordinary requester sees only their own requests.
    return base_qs.filter(
        Q(submitted_by=user)
        | Q(minute_sheet__requested_by=user)
    ).distinct().order_by("-created_at")


def get_visible_retirements(user):
    base_qs = Retirement.objects.select_related(
        "submitted_by",
        "finance_request",
        "finance_request__department",
        "finance_request__department_unit",
        "finance_request__minute_sheet",
    )

    if can_see_all_finance(user):
        return base_qs.order_by("-created_at")

    if is_executive_viewer(user):
        return apply_executive_finance_scope(
            base_qs, user, "finance_request__department"
        ).order_by("-created_at")

    department = get_user_department(user)
    department_unit = get_user_department_unit(user)

    if is_accountant(user):
        if department:
            return base_qs.filter(
                finance_request__department=department
            ).exclude(status="DRAFT").order_by("-created_at")

        return base_qs.exclude(status="DRAFT").order_by("-created_at")

    if is_director(user) or is_assistant_director(user):
        if not department:
            return base_qs.none()

        return base_qs.filter(
            finance_request__department=department
        ).order_by("-created_at")

    if is_hou(user):
        if department_unit:
            return base_qs.filter(
                Q(submitted_by=user)
                | Q(finance_request__department_unit=department_unit)
            ).distinct().order_by("-created_at")

        if department:
            return base_qs.filter(
                Q(submitted_by=user)
                | Q(
                    finance_request__department=department,
                    finance_request__department_unit__isnull=True,
                )
            ).distinct().order_by("-created_at")

    return base_qs.filter(
        submitted_by=user
    ).order_by("-created_at")


def get_visible_budget_lines(user):
    """
    Return Budget Lines according to organizational ownership.

    Access rules:
    - Superuser/System Administrator: all Budget Lines.
    - Every other user, including DBO, Accountant, Director and
      Assistant Director: only Budget Lines from their department.
    - Users without a department: no Budget Lines.
    """
    base_qs = BudgetLine.objects.select_related(
        "department",
        "department_unit",
        "finance_request",
        "finance_request__department",
    )

    if user.is_superuser or get_user_role(user) == "ADMIN":
        return base_qs.order_by(
            "financial_year",
            "month",
            "task_code",
        )

    if is_executive_viewer(user):
        return apply_executive_finance_scope(base_qs, user).order_by(
            "financial_year", "month", "task_code"
        )

    department = get_user_department(user)

    if not department:
        return base_qs.none()

    return base_qs.filter(
        department=department
    ).order_by(
        "financial_year",
        "month",
        "task_code",
    )


def get_budget_available_fund(budget_line):
    if not budget_line:
        return Decimal("0.00")

    approved = budget_line.monthly_approved_amount or Decimal("0.00")
    disbursed = budget_line.amount_disbursed or Decimal("0.00")

    return approved - disbursed


def parse_decimal(value, fallback=Decimal("0.00")):
    if value in (None, ""):
        return fallback

    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return fallback


def get_virtual_budgetline_date(line):
    month_map = {
        "JULY": 7,
        "AUGUST": 8,
        "SEPTEMBER": 9,
        "OCTOBER": 10,
        "NOVEMBER": 11,
        "DECEMBER": 12,
        "JANUARY": 1,
        "FEBRUARY": 2,
        "MARCH": 3,
        "APRIL": 4,
        "MAY": 5,
        "JUNE": 6,
    }

    month_name = (line.month or "").upper()
    month_number = month_map.get(month_name)

    if not month_number:
        return None

    try:
        years = (line.financial_year or "").split("/")
        first_year = int(years[0])
        second_year = (
            int(years[1])
            if len(years) > 1
            else first_year + 1
        )
    except (IndexError, TypeError, ValueError):
        return None

    year = first_year if month_number >= 7 else second_year
    return date(year, month_number, 1)


def get_filtered_budget_lines(request):
    search_query = request.GET.get("search", "").strip().lower()
    month_filter = request.GET.get("month", "all").strip().upper()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    financial_year = request.GET.get("financial_year", "").strip()
    department_filter = request.GET.get("department", "").strip()

    if not financial_year:
        financial_year = get_current_financial_year()

    budget_lines = list(
        get_visible_budget_lines(request.user).filter(
            **({"department_id": department_filter} if department_filter and is_executive_viewer(request.user) else {})
        ).order_by(
            "financial_year",
            "month",
            "task_code",
        )
    )

    if search_query:
        budget_lines = [
            line
            for line in budget_lines
            if search_query in (line.task_code or "").lower()
            or search_query in (line.task_in_mtef or "").lower()
            or search_query in (line.month or "").lower()
            or search_query in (line.financial_year or "").lower()
        ]

    if month_filter and month_filter != "ALL":
        budget_lines = [
            line
            for line in budget_lines
            if (line.month or "").upper() == month_filter
        ]

    if financial_year:
        budget_lines = [
            line
            for line in budget_lines
            if (line.financial_year or "").strip() == financial_year
        ]

    parsed_from = None
    parsed_to = None

    try:
        if date_from:
            parsed_from = date.fromisoformat(date_from)

        if date_to:
            parsed_to = date.fromisoformat(date_to)
    except ValueError:
        parsed_from = None
        parsed_to = None

    if parsed_from or parsed_to:
        filtered_by_date = []

        for line in budget_lines:
            virtual_date = get_virtual_budgetline_date(line)

            if not virtual_date:
                continue

            if parsed_from and virtual_date < parsed_from:
                continue

            if parsed_to and virtual_date > parsed_to:
                continue

            filtered_by_date.append(line)

        budget_lines = filtered_by_date

    return budget_lines


def calculate_unique_annual_budget(budget_lines):
    seen_tasks = set()
    total = Decimal("0.00")

    for line in budget_lines:
        key = (line.financial_year, line.task_code)

        if key not in seen_tasks:
            total += line.budgeted_amount or Decimal("0.00")
            seen_tasks.add(key)

    return total


def build_finance_analytics_data(budget_lines):
    total_proposed = Decimal("0.00")
    total_approved = Decimal("0.00")
    total_disbursed = Decimal("0.00")

    monthly_data = {}
    task_data = {}

    for line in budget_lines:
        proposed = line.monthly_proposed_amount or Decimal("0.00")
        approved = line.monthly_approved_amount or Decimal("0.00")
        disbursed = line.amount_disbursed or Decimal("0.00")

        total_proposed += proposed
        total_approved += approved
        total_disbursed += disbursed

        month = line.month or "Unknown"

        if month not in monthly_data:
            monthly_data[month] = {
                "proposed": Decimal("0.00"),
                "approved": Decimal("0.00"),
                "disbursed": Decimal("0.00"),
            }

        monthly_data[month]["proposed"] += proposed
        monthly_data[month]["approved"] += approved
        monthly_data[month]["disbursed"] += disbursed

        task_key = f"{line.task_code}-{line.task_in_mtef}"

        if task_key not in task_data:
            task_data[task_key] = {
                "task_code": line.task_code,
                "task_in_mtef": line.task_in_mtef,
                "approved": Decimal("0.00"),
                "disbursed": Decimal("0.00"),
                "balance": Decimal("0.00"),
            }

        task_data[task_key]["approved"] += approved
        task_data[task_key]["disbursed"] += disbursed
        task_data[task_key]["balance"] = (
            task_data[task_key]["approved"]
            - task_data[task_key]["disbursed"]
        )

    total_budgeted = calculate_unique_annual_budget(budget_lines)
    balance = total_approved - total_disbursed

    monthly_data = {
        month: monthly_data[month]
        for month in MONTH_ORDER
        if month in monthly_data
    }

    top_consuming_tasks = sorted(
        task_data.values(),
        key=lambda item: item["disbursed"],
        reverse=True,
    )[:10]

    remaining_funds_tasks = sorted(
        task_data.values(),
        key=lambda item: item["balance"],
        reverse=True,
    )[:10]

    return {
        "total_budgeted": total_budgeted,
        "total_proposed": total_proposed,
        "total_approved": total_approved,
        "total_disbursed": total_disbursed,
        "balance": balance,
        "monthly_data": monthly_data,
        "top_consuming_tasks": top_consuming_tasks,
        "remaining_funds_tasks": remaining_funds_tasks,
    }


@login_required
def finance_dashboard(request):
    role = get_user_role(request.user)
    current_financial_year = get_current_financial_year()
    executive_viewer = is_executive_viewer(request.user)
    selected_department = request.GET.get("department", "").strip()

    minute_sheets = get_visible_minute_sheets(request.user)
    finance_requests = get_visible_finance_requests(request.user)
    retirements = get_visible_retirements(request.user)

    visible_budget_lines = get_visible_budget_lines(request.user)

    if selected_department and executive_viewer:
        allowed_departments = executive_finance_departments(request.user)
        if allowed_departments.filter(pk=selected_department).exists():
            minute_sheets = minute_sheets.filter(department_id=selected_department)
            finance_requests = finance_requests.filter(department_id=selected_department)
            retirements = retirements.filter(
                finance_request__department_id=selected_department
            )
            visible_budget_lines = visible_budget_lines.filter(
                department_id=selected_department
            )
        else:
            selected_department = ""

    minute_sheets = minute_sheets[:10]
    finance_requests = finance_requests[:10]
    retirements = retirements[:10]

    budget_lines = visible_budget_lines.only(
        "task_code",
        "task_in_mtef",
        "month",
        "monthly_approved_amount",
        "amount_disbursed",
        "financial_year",
    ).order_by("-id")[:10]

    visible_request_ids = get_visible_finance_requests(
        request.user
    ).values_list("id", flat=True)

    disbursements = Disbursement.objects.filter(
        finance_request_id__in=visible_request_ids
    ).order_by("-created_at")[:10]

    if can_view_finance_reports(request.user):
        dashboard_budget_lines = visible_budget_lines.filter(
            financial_year=current_financial_year
        )

        totals = dashboard_budget_lines.aggregate(
            total_monthly_proposed=Sum("monthly_proposed_amount"),
            total_monthly_approved=Sum("monthly_approved_amount"),
            total_disbursed=Sum("amount_disbursed"),
        )

        total_budgeted = calculate_unique_annual_budget(
            dashboard_budget_lines
        )
        total_monthly_proposed = (
            totals["total_monthly_proposed"]
            or Decimal("0.00")
        )
        total_monthly_approved = (
            totals["total_monthly_approved"]
            or Decimal("0.00")
        )
        total_disbursed = (
            totals["total_disbursed"]
            or Decimal("0.00")
        )
        balance = total_monthly_approved - total_disbursed
    else:
        total_budgeted = Decimal("0.00")
        total_monthly_proposed = Decimal("0.00")
        total_monthly_approved = Decimal("0.00")
        total_disbursed = Decimal("0.00")
        balance = Decimal("0.00")

    context = {
        "role": role,
        "role_display": get_role_display_name(role),
        "can_manage_budget_lines": (
            request.user.is_superuser
            or is_dbo(request.user)
        ),
        "can_view_budget_lines": (
            request.user.is_superuser
            or executive_viewer
            or is_dbo(request.user)
            or is_director(request.user)
            or is_assistant_director(request.user)
        ),
        "can_view_reports": can_view_finance_reports(request.user),
        "can_start_financial_year": (
            request.user.is_superuser
            or is_dbo(request.user)
        ),
        "current_financial_year": current_financial_year,
        "minute_sheets": minute_sheets,
        "finance_requests": finance_requests,
        "budget_lines": budget_lines,
        "disbursements": disbursements,
        "retirements": retirements,
        "total_budgeted": total_budgeted,
        "total_monthly_proposed": total_monthly_proposed,
        "total_monthly_approved": total_monthly_approved,
        "total_disbursed": total_disbursed,
        "balance": balance,
        "is_read_only_executive": executive_viewer,
        "departments": executive_finance_departments(request.user),
        "selected_department": selected_department,
    }

    return render(request, "finance/dashboard.html", context)


@login_required
@transaction.atomic
def minute_sheet_create(request):
    if is_executive_viewer(request.user):
        raise PermissionDenied("Executive access to Finance is read-only.")
    user_department = get_user_department(request.user)

    if not user_department:
        messages.error(
            request,
            "Your user profile has no department. Contact the System Administrator.",
        )
        return redirect("finance:dashboard")

    if request.method == "POST":
        form = MinuteSheetForm(
            request.POST,
            request.FILES,
            user=request.user,
        )

        if form.is_valid():
            minute_sheet = form.save(commit=False)
            minute_sheet.requested_by = request.user
            apply_user_ownership(minute_sheet, request.user)

            minute_sheet.title = minute_sheet.subject
            minute_sheet.description = minute_sheet.activity_series
            minute_sheet.estimated_total = (
                minute_sheet.requested_amount
                or Decimal("0.00")
            )
            minute_sheet.status = "SUBMITTED"

            if not minute_sheet.reference_no:
                minute_sheet.reference_no = (
                    generate_minute_sheet_reference(request.user)
                )

            minute_sheet.save()

            finance_request = FinanceRequest(
                minute_sheet=minute_sheet,
                request_type="SPECIAL_ACTIVITY",
                requested_amount=minute_sheet.requested_amount,
                approved_amount=Decimal("0.00"),
                currency="TZS",
                financial_year=(
                    minute_sheet.budget_line.financial_year
                    if minute_sheet.budget_line
                    else get_current_financial_year()
                ),
                status="SUBMITTED",
                submitted_by=request.user,
            )
            apply_user_ownership(finance_request, request.user)
            finance_request.save()
            initialize_finance_approval_steps(finance_request)

            messages.success(
                request,
                (
                    "Minute Sheet and Finance Request submitted successfully. "
                    "The approval workflow has been initialized."
                ),
            )
            return redirect(
                "finance:minute_sheet_detail",
                pk=minute_sheet.pk,
            )

        messages.error(
            request,
            "Minute sheet was not saved. Correct the errors below.",
        )
    else:
        form = MinuteSheetForm(user=request.user)

    return render(
        request,
        "finance/minute_sheet_form.html",
        {"form": form},
    )


@login_required
def minute_sheet_detail(request, pk):
    minute_sheet = get_object_or_404(
        get_visible_minute_sheets(request.user),
        pk=pk,
    )
    role = get_user_role(request.user)
    if request.method == "POST" and is_executive_viewer(request.user):
        raise PermissionDenied("Executive access to Finance is read-only.")
    available_fund = get_budget_available_fund(
        minute_sheet.budget_line
    )

    finance_request = (
        minute_sheet.finance_requests.select_related(
            "department",
            "department_unit",
            "submitted_by",
        )
        .order_by("created_at")
        .first()
    )

    if (
        finance_request
        and finance_request.status == "SUBMITTED"
        and not finance_request.approval_steps.exists()
    ):
        try:
            initialize_finance_approval_steps(finance_request)
        except ValueError as error:
            messages.error(
                request,
                f"Approval workflow could not be initialized: {error}",
            )

    if finance_request:
        workflow_context = get_finance_review_context(
            request.user,
            finance_request,
        )
    else:
        workflow_context = {
            "workflow_steps": [],
            "current_workflow_step": None,
            "can_review": False,
            "can_approve": False,
            "can_return": False,
            "can_reject": False,
            "workflow_complete": False,
        }

    if request.method == "POST":
        if not finance_request:
            messages.error(
                request,
                "This Minute Sheet has no related Finance Request.",
            )
            return redirect(
                "finance:minute_sheet_detail",
                pk=minute_sheet.pk,
            )

        if not workflow_context["can_review"]:
            messages.error(
                request,
                "You are not the assigned approver for the current step.",
            )
            return redirect(
                "finance:minute_sheet_detail",
                pk=minute_sheet.pk,
            )

        action = request.POST.get("action", "").strip().lower()
        remarks = request.POST.get("remarks", "").strip()
        current_step = workflow_context["current_workflow_step"]

        try:
            if action == "return":
                return_current_finance_step(
                    request.user,
                    finance_request,
                    remarks,
                )

                minute_sheet.status = "DRAFT"
                minute_sheet.is_approved = False
                minute_sheet.director_remarks = remarks
                minute_sheet.save(
                    update_fields=[
                        "status",
                        "is_approved",
                        "director_remarks",
                        "updated_at",
                    ]
                )

                messages.success(
                    request,
                    "Minute Sheet returned to the requester for correction.",
                )
                return redirect(
                    "finance:minute_sheet_detail",
                    pk=minute_sheet.pk,
                )

            if action == "reject":
                reject_current_finance_step(
                    request.user,
                    finance_request,
                    remarks,
                )

                minute_sheet.status = "REJECTED"
                minute_sheet.is_approved = False
                minute_sheet.director_remarks = remarks
                minute_sheet.save(
                    update_fields=[
                        "status",
                        "is_approved",
                        "director_remarks",
                        "updated_at",
                    ]
                )

                messages.success(request, "Minute Sheet rejected.")
                return redirect(
                    "finance:minute_sheet_detail",
                    pk=minute_sheet.pk,
                )

            if action != "approve":
                messages.error(
                    request,
                    "Select Approve, Return, or Reject.",
                )
                return redirect(
                    "finance:minute_sheet_detail",
                    pk=minute_sheet.pk,
                )

            if not minute_sheet.budget_line:
                messages.error(
                    request,
                    "Approval blocked. No related Budget Line.",
                )
                return redirect(
                    "finance:minute_sheet_detail",
                    pk=minute_sheet.pk,
                )

            step_role = current_step.step_name if current_step else ""

            amount = parse_decimal(
                request.POST.get("amount"),
                (
                    minute_sheet.director_approved_amount
                    or minute_sheet.requested_amount
                ),
            )

            if step_role == "DIRECTOR":
                if amount <= 0:
                    messages.error(
                        request,
                        "The amount must be greater than zero.",
                    )
                    return redirect(
                        "finance:minute_sheet_detail",
                        pk=minute_sheet.pk,
                    )

                if amount > minute_sheet.requested_amount:
                    messages.error(
                        request,
                        "The amount cannot exceed the requested amount.",
                    )
                    return redirect(
                        "finance:minute_sheet_detail",
                        pk=minute_sheet.pk,
                    )

            if step_role == "HEAD_OF_UNIT":
                finance_request.status = "HOU_REVIEWED"

            elif step_role == "DIRECTOR":
                available_fund = get_budget_available_fund(
                    minute_sheet.budget_line
                )

                if amount > available_fund:
                    messages.error(
                        request,
                        (
                            "Approval blocked. Available fund is only "
                            f"{available_fund:,.2f} TZS."
                        ),
                    )
                    return redirect(
                        "finance:minute_sheet_detail",
                        pk=minute_sheet.pk,
                    )

                minute_sheet.director_approved_amount = amount
                finance_request.approved_amount = amount
                finance_request.status = "DIRECTOR_APPROVED"

            elif step_role == "DIVISION_BUDGET_OFFICER":
                approved_amount = (
                    finance_request.approved_amount
                    or minute_sheet.director_approved_amount
                    or minute_sheet.requested_amount
                )

                available_fund = get_budget_available_fund(
                    minute_sheet.budget_line
                )

                if approved_amount > available_fund:
                    messages.error(
                        request,
                        (
                            "Budget verification blocked. Available fund is "
                            f"only {available_fund:,.2f} TZS."
                        ),
                    )
                    return redirect(
                        "finance:minute_sheet_detail",
                        pk=minute_sheet.pk,
                    )

            approved_step, next_step, workflow_complete = (
                approve_current_finance_step(
                    request.user,
                    finance_request,
                    remarks,
                )
            )

            minute_sheet.director_remarks = remarks
            minute_sheet.save()

            finance_request.remarks = remarks
            finance_request.save()

            if workflow_complete:
                approved_amount = (
                    finance_request.approved_amount
                    or minute_sheet.director_approved_amount
                    or minute_sheet.hou_recommended_amount
                    or minute_sheet.requested_amount
                )

                with transaction.atomic():
                    locked_request = (
                        FinanceRequest.objects.select_for_update().get(
                            pk=finance_request.pk
                        )
                    )
                    budget_line = (
                        BudgetLine.objects.select_for_update().get(
                            pk=minute_sheet.budget_line_id
                        )
                    )

                    disbursement_no = (
                        f"AUTO-{minute_sheet.reference_no}"
                    )

                    disbursement, created = (
                        Disbursement.objects.get_or_create(
                            finance_request=locked_request,
                            disbursement_no=disbursement_no,
                            defaults={
                                "amount": approved_amount,
                                "payment_method": "BANK",
                                "recipient_name": (
                                    minute_sheet.requested_by.get_full_name()
                                    or minute_sheet.requested_by.username
                                    if minute_sheet.requested_by
                                    else "System"
                                ),
                                "status": "PAID",
                                "disbursed_by": request.user,
                                "disbursed_at": timezone.now(),
                                "notes": (
                                    "Automatically recorded after completion "
                                    "of the configured Finance workflow."
                                ),
                            },
                        )
                    )

                    if created:
                        budget_line.amount_disbursed = (
                            budget_line.amount_disbursed
                            or Decimal("0.00")
                        ) + approved_amount
                        budget_line.save()

                    locked_request.approved_amount = approved_amount
                    locked_request.status = "DISBURSED"
                    locked_request.remarks = remarks
                    locked_request.save()

                    minute_sheet.director_approved_amount = approved_amount
                    minute_sheet.status = "APPROVED"
                    minute_sheet.is_approved = True
                    minute_sheet.director_remarks = remarks
                    minute_sheet.save()

                messages.success(
                    request,
                    "All approval steps are complete. Payment was recorded.",
                )
            else:
                next_role = (
                    next_step.get_step_name_display()
                    if next_step
                    else "the next approver"
                )
                messages.success(
                    request,
                    f"Approved and forwarded to {next_role}.",
                )

            return redirect(
                "finance:minute_sheet_detail",
                pk=minute_sheet.pk,
            )

        except (PermissionDenied, ValueError) as error:
            messages.error(request, str(error))
            return redirect(
                "finance:minute_sheet_detail",
                pk=minute_sheet.pk,
            )

    if finance_request:
        workflow_context = get_finance_review_context(
            request.user,
            finance_request,
        )

    return render(
        request,
        "finance/minute_sheet_detail.html",
        {
            "minute_sheet": minute_sheet,
            "finance_request": finance_request,
            "role": role,
            "available_fund": available_fund,
            **workflow_context,
        },
    )


@login_required
@transaction.atomic
def finance_request_create(request):
    if is_executive_viewer(request.user):
        raise PermissionDenied("Executive access to Finance is read-only.")
    user_department = get_user_department(request.user)

    if not user_department:
        messages.error(
            request,
            "Your user profile has no department. Contact the System Administrator.",
        )
        return redirect("finance:dashboard")

    if request.method == "POST":
        form = FinanceRequestForm(
            request.POST,
            user=request.user,
        )

        if form.is_valid():
            finance_request = form.save(commit=False)
            finance_request.submitted_by = request.user
            apply_user_ownership(finance_request, request.user)
            finance_request.approved_amount = Decimal("0.00")
            finance_request.status = "SUBMITTED"
            finance_request.save()
            initialize_finance_approval_steps(finance_request)

            messages.success(
                request,
                (
                    "Finance Request submitted successfully. "
                    "The approval workflow has been initialized."
                ),
            )
            return redirect(
                "finance:finance_request_detail",
                pk=finance_request.pk,
            )
    else:
        form = FinanceRequestForm(user=request.user)

    return render(
        request,
        "finance/finance_request_form.html",
        {"form": form},
    )


@login_required
def finance_request_detail(request, pk):
    finance_request = get_object_or_404(
        get_visible_finance_requests(request.user),
        pk=pk,
    )
    budget_lines = finance_request.budget_lines.all()
    role = get_user_role(request.user)

    if request.method == "POST" and is_executive_viewer(request.user):
        raise PermissionDenied("Executive access to Finance is read-only.")

    if request.method == "POST":
        action = request.POST.get("action")
        remarks = request.POST.get("remarks", "").strip()

        if action == "hou_review" and is_hou(request.user):
            finance_request.status = "HOU_REVIEWED"
            finance_request.remarks = remarks
            finance_request.save()

            messages.success(
                request,
                "Finance request reviewed by Head of Unit.",
            )
            return redirect(
                "finance:finance_request_detail",
                pk=finance_request.pk,
            )

        if action == "director_approve" and is_final_approver(request.user):
            approved_amount = parse_decimal(
                request.POST.get("approved_amount"),
                finance_request.requested_amount,
            )

            if approved_amount <= 0:
                messages.error(
                    request,
                    "Approved amount must be greater than zero.",
                )
                return redirect(
                    "finance:finance_request_detail",
                    pk=finance_request.pk,
                )

            if approved_amount > finance_request.requested_amount:
                messages.error(
                    request,
                    "Approved amount cannot exceed requested amount.",
                )
                return redirect(
                    "finance:finance_request_detail",
                    pk=finance_request.pk,
                )

            finance_request.approved_amount = approved_amount
            finance_request.status = "DIRECTOR_APPROVED"
            finance_request.remarks = remarks
            finance_request.save()

            messages.success(
                request,
                "Finance request approved.",
            )
            return redirect(
                "finance:finance_request_detail",
                pk=finance_request.pk,
            )

        if action == "reject" and (
            is_hou(request.user)
            or is_final_approver(request.user)
        ):
            finance_request.status = "REJECTED"
            finance_request.remarks = remarks
            finance_request.save()

            messages.success(request, "Finance request rejected.")
            return redirect(
                "finance:finance_request_detail",
                pk=finance_request.pk,
            )

        messages.error(
            request,
            "You are not allowed to perform that action.",
        )

    return render(
        request,
        "finance/finance_request_detail.html",
        {
            "finance_request": finance_request,
            "budget_lines": budget_lines,
            "role": role,
        },
    )


@login_required
def budget_line_list(request):
    role = get_user_role(request.user)

    if not (
        request.user.is_superuser
        or is_executive_viewer(request.user)
        or is_dbo(request.user)
        or is_director(request.user)
        or is_assistant_director(request.user)
    ):
        messages.error(
            request,
            "You are not allowed to view budget lines.",
        )
        return redirect("finance:dashboard")

    selected_financial_year = request.GET.get(
        "financial_year",
        "",
    ).strip()

    selected_month = request.GET.get(
        "month",
        "",
    ).strip().upper()

    visible_lines = get_visible_budget_lines(
        request.user
    )

    financial_years = (
        visible_lines
        .exclude(financial_year__isnull=True)
        .exclude(financial_year="")
        .values_list(
            "financial_year",
            flat=True,
        )
        .distinct()
        .order_by("financial_year")
    )

    if not selected_financial_year:
        selected_financial_year = (
            get_current_financial_year()
        )

    budget_lines = visible_lines.filter(
        financial_year=selected_financial_year
    )

    if selected_month:
        budget_lines = budget_lines.filter(
            month=selected_month
        )

    budget_lines = budget_lines.order_by("-id")

    return render(
        request,
        "finance/budget_line_list.html",
        {
            "budget_lines": budget_lines,
            "role": role,
            "role_display": get_role_display_name(role),
            "selected_financial_year": (
                selected_financial_year
            ),
            "selected_month": selected_month,
            "financial_years": financial_years,
            "months": MONTH_ORDER,
        },
    )


@login_required
def budget_line_create(request):
    role = get_user_role(request.user)

    if not is_dbo(request.user) and role != "ADMIN":
        messages.error(
            request,
            "Only the Division Budget Officer or System Administrator "
            "can add budget lines.",
        )
        return redirect("finance:dashboard")

    if request.method == "POST":
        form = BudgetLineForm(request.POST)

        if form.is_valid():
            budget_line = form.save()
            messages.success(
                request,
                "Budget line added successfully.",
            )
            return redirect("finance:budget_line_list")

        messages.error(
            request,
            "Budget line was not saved. Correct the errors below.",
        )
    else:
        form = BudgetLineForm()

    return render(
        request,
        "finance/budget_line_form.html",
        {
            "form": form,
            "role": role,
        },
    )


@login_required
def budget_line_edit(request, pk):
    role = get_user_role(request.user)

    if not is_dbo(request.user) and role != "ADMIN":
        messages.error(
            request,
            "Only the Division Budget Officer or System Administrator "
            "can edit budget lines.",
        )
        return redirect("finance:dashboard")

    budget_line = get_object_or_404(
        get_visible_budget_lines(request.user),
        pk=pk,
    )

    if request.method == "POST":
        form = BudgetLineForm(
            request.POST,
            instance=budget_line,
        )

        if form.is_valid():
            form.save()
            messages.success(
                request,
                "Budget line updated successfully.",
            )
            return redirect("finance:budget_line_list")

        messages.error(
            request,
            "Budget line was not updated. Correct the errors below.",
        )
    else:
        form = BudgetLineForm(instance=budget_line)

    return render(
        request,
        "finance/budget_line_form.html",
        {
            "form": form,
            "role": role,
            "budget_line": budget_line,
            "is_edit": True,
        },
    )


@login_required
def clone_month_to_next(request):
    role = get_user_role(request.user)

    if not request.user.is_superuser and not is_dbo(request.user):
        messages.error(
            request,
            "You are not allowed to perform this action.",
        )
        return redirect("finance:budget_line_list")

    current_month = request.GET.get("month", "").upper()
    financial_year = request.GET.get("financial_year", "")

    if not current_month:
        messages.error(request, "No month selected.")
        return redirect("finance:budget_line_list")

    next_month = get_next_budget_month(current_month)

    if not next_month:
        messages.error(
            request,
            "Cannot clone the last month of the financial year.",
        )
        return redirect("finance:budget_line_list")

    current_lines = get_visible_budget_lines(request.user).filter(
        financial_year=financial_year,
        month=current_month,
    )

    if not current_lines.exists():
        messages.warning(
            request,
            f"No data found for {current_month}.",
        )
        return redirect("finance:budget_line_list")

    created_count = 0
    skipped_count = 0

    for line in current_lines:
        exists = BudgetLine.objects.filter(
            financial_year=financial_year,
            month=next_month,
            task_code=line.task_code,
        ).exists()

        if exists:
            skipped_count += 1
            continue

        BudgetLine.objects.create(
            financial_year=line.financial_year,
            month=next_month,
            task_code=line.task_code,
            task_in_mtef=line.task_in_mtef,
            monthly_activity_description=(
                line.monthly_activity_description
            ),
            budgeted_amount=line.budgeted_amount,
            monthly_proposed_amount=Decimal("0.00"),
            monthly_approved_amount=Decimal("0.00"),
            amount_disbursed=Decimal("0.00"),
            remarks=f"Cloned from {current_month}",
        )
        created_count += 1

    messages.success(
        request,
        (
            f"{created_count} lines cloned to {next_month}. "
            f"{skipped_count} skipped because they already existed."
        ),
    )
    return redirect("finance:budget_line_list")


@login_required
def disbursement_create(request):
    messages.error(
        request,
        "Manual disbursement is disabled. The system automatically "
        "records an approved Minute Sheet amount.",
    )
    return redirect("finance:dashboard")


@login_required
def finance_reports(request):
    role = get_user_role(request.user)

    if not can_view_finance_reports(request.user):
        messages.error(
            request,
            "You are not allowed to view finance reports.",
        )
        return redirect("finance:dashboard")

    budget_lines = get_filtered_budget_lines(request)

    if request.GET.get("export") == "excel":
        return export_finance_reports_excel(request, budget_lines)

    totals = {
        "total_budgeted": calculate_unique_annual_budget(budget_lines),
        "total_proposed": sum(
            (
                line.monthly_proposed_amount
                or Decimal("0.00")
                for line in budget_lines
            ),
            Decimal("0.00"),
        ),
        "total_approved": sum(
            (
                line.monthly_approved_amount
                or Decimal("0.00")
                for line in budget_lines
            ),
            Decimal("0.00"),
        ),
        "total_disbursed": sum(
            (
                line.amount_disbursed
                or Decimal("0.00")
                for line in budget_lines
            ),
            Decimal("0.00"),
        ),
    }

    financial_years = (
        get_visible_budget_lines(request.user)
        .exclude(financial_year__isnull=True)
        .exclude(financial_year="")
        .values_list("financial_year", flat=True)
        .distinct()
        .order_by("financial_year")
    )

    context = {
        "role": role,
        "budget_lines": budget_lines,
        "total_budgeted": totals["total_budgeted"],
        "total_proposed": totals["total_proposed"],
        "total_approved": totals["total_approved"],
        "total_disbursed": totals["total_disbursed"],
        "search_query": request.GET.get("search", ""),
        "selected_month": request.GET.get("month", "all"),
        "selected_financial_year": request.GET.get(
            "financial_year",
            "",
        ),
        "date_from": request.GET.get("date_from", ""),
        "date_to": request.GET.get("date_to", ""),
        "financial_years": financial_years,
        "months": MONTH_ORDER,
        "departments": executive_finance_departments(request.user),
        "selected_department": request.GET.get("department", ""),
        "is_read_only_executive": is_executive_viewer(request.user),
    }

    return render(request, "finance/reports.html", context)


@login_required
def export_finance_reports_excel(request, budget_lines=None):
    if not can_view_finance_reports(request.user):
        messages.error(
            request,
            "You are not allowed to export finance reports.",
        )
        return redirect("finance:dashboard")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    budget_lines = (
        budget_lines
        if budget_lines is not None
        else get_filtered_budget_lines(request)
    )

    total_budgeted = calculate_unique_annual_budget(
        budget_lines
    )

    total_proposed = sum(
        (
            line.monthly_proposed_amount
            or Decimal("0.00")
            for line in budget_lines
        ),
        Decimal("0.00"),
    )

    total_approved = sum(
        (
            line.monthly_approved_amount
            or Decimal("0.00")
            for line in budget_lines
        ),
        Decimal("0.00"),
    )

    total_disbursed = sum(
        (
            line.amount_disbursed
            or Decimal("0.00")
            for line in budget_lines
        ),
        Decimal("0.00"),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Finance Reports"

    ws.append(["Finance Reports"])

    ws.append([
        "Generated At",
        timezone.now().strftime("%d/%m/%Y %H:%M"),
    ])

    ws.append([
        "Financial Year",
        request.GET.get("financial_year", "") or "All",
    ])

    ws.append([
        "Month",
        request.GET.get("month", "all") or "All",
    ])

    ws.append([
        "Search",
        request.GET.get("search", "") or "-",
    ])

    ws.append([
        "Date From",
        request.GET.get("date_from", "") or "-",
    ])

    ws.append([
        "Date To",
        request.GET.get("date_to", "") or "-",
    ])

    ws.append([])

    ws.append(["Summary"])
    ws.append(["Total Budgeted", float(total_budgeted)])
    ws.append(["Total Proposed", float(total_proposed)])
    ws.append(["Total Approved", float(total_approved)])
    ws.append(["Total Disbursed", float(total_disbursed)])

    ws.append([])

    headers = [
        "#",
        "Financial Year",
        "Month",
        "Task Code",
        "Task in MTEF",
        "Activity",
        "Budgeted",
        "Proposed",
        "Approved",
        "Disbursed",
        "Available",
        "Variance",
        "Remarks",
    ]

    ws.append(headers)
    header_row = ws.max_row

    for index, line in enumerate(
        budget_lines,
        start=1,
    ):
        proposed = (
            line.monthly_proposed_amount
            or Decimal("0.00")
        )

        approved = (
            line.monthly_approved_amount
            or Decimal("0.00")
        )

        disbursed = (
            line.amount_disbursed
            or Decimal("0.00")
        )

        available = approved - disbursed
        variance = proposed - approved

        ws.append([
            index,
            line.financial_year,
            line.month,
            line.task_code,
            line.task_in_mtef,
            line.monthly_activity_description,
            float(
                line.budgeted_amount
                or Decimal("0.00")
            ),
            float(proposed),
            float(approved),
            float(disbursed),
            float(available),
            float(variance),
            line.remarks,
        ])

    title_font = Font(
        bold=True,
        size=14,
    )

    bold_font = Font(
        bold=True,
    )

    header_fill = PatternFill(
        "solid",
        fgColor="E9F1FF",
    )

    ws["A1"].font = title_font

    for cell in ws[9]:
        cell.font = bold_font

    for cell in ws[header_row]:
        cell.font = bold_font
        cell.fill = header_fill

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for column in ws.columns:
        max_length = max(
            (
                len(str(cell.value))
                if cell.value is not None
                else 0
            )
            for cell in column
        )

        ws.column_dimensions[
            get_column_letter(column[0].column)
        ].width = min(
            max_length + 2,
            45,
        )

    for row in ws.iter_rows(
        min_row=header_row + 1,
        min_col=7,
        max_col=12,
    ):
        for cell in row:
            cell.number_format = "#,##0.00"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

    response["Content-Disposition"] = (
        'attachment; filename="finance_reports.xlsx"'
    )

    return response


@login_required
def search_mtef_task(request):
    task_code = request.GET.get("task_code", "").strip()
    financial_year = request.GET.get(
        "financial_year",
        "",
    ).strip()

    if not financial_year:
        financial_year = get_current_financial_year()

    if not task_code:
        return JsonResponse({
            "found": False,
            "message": "Task code is required.",
        })

    budget_query = get_visible_budget_lines(request.user).filter(
        task_code__iexact=task_code,
        financial_year=financial_year,
    )

    latest_line = budget_query.order_by("-id").first()

    if not latest_line:
        return JsonResponse({
            "found": False,
            "message": "No MTEF task found for this Task Code.",
        })

    total_budgeted = calculate_unique_annual_budget(
        list(budget_query)
    )

    totals = budget_query.aggregate(
        total_approved=Sum("monthly_approved_amount"),
        total_disbursed=Sum("amount_disbursed"),
    )

    total_approved = (
        totals["total_approved"]
        or Decimal("0.00")
    )
    total_disbursed = (
        totals["total_disbursed"]
        or Decimal("0.00")
    )
    previous_balance = total_approved - total_disbursed

    return JsonResponse({
        "found": True,
        "task_code": latest_line.task_code,
        "task_in_mtef": latest_line.task_in_mtef,
        "financial_year": latest_line.financial_year,
        "budgeted_amount": str(total_budgeted),
        "previous_approved": str(total_approved),
        "previous_disbursed": str(total_disbursed),
        "previous_balance": str(previous_balance),
    })


@login_required
def finance_analytics(request):
    role = get_user_role(request.user)

    if not can_view_finance_reports(request.user):
        messages.error(
            request,
            "You are not allowed to view finance analytics.",
        )
        return redirect("finance:dashboard")

    budget_lines = get_filtered_budget_lines(request)
    analytics_data = build_finance_analytics_data(budget_lines)

    context = {
        "role": role,
        "selected_table": request.GET.get("table", "all"),
        "search_query": request.GET.get("search", ""),
        "selected_month": request.GET.get("month", "all"),
        "financial_year": request.GET.get("financial_year", ""),
        "departments": executive_finance_departments(request.user),
        "selected_department": request.GET.get("department", ""),
        "is_read_only_executive": is_executive_viewer(request.user),
        **analytics_data,
    }

    return render(
        request,
        "finance/analytics.html",
        context,
    )


@login_required
def start_new_financial_year(request):
    role = get_user_role(request.user)

    if not request.user.is_superuser and not is_dbo(request.user):
        messages.error(
            request,
            "You are not allowed to start a new financial year.",
        )
        return redirect("finance:dashboard")

    new_financial_year = get_current_financial_year()

    if request.method == "POST":
        messages.success(
            request,
            (
                f"Financial year {new_financial_year} is ready. "
                "You can now enter the new departmental Budget Lines."
            ),
        )
        return redirect("finance:budget_line_list")

    return render(
        request,
        "finance/start_new_financial_year.html",
        {
            "role": role,
            "role_display": get_role_display_name(role),
            "new_financial_year": new_financial_year,
        },
    )

@login_required
def export_finance_analytics_excel(request):
    if not can_view_finance_reports(request.user):
        messages.error(
            request,
            "You are not allowed to export finance analytics.",
        )
        return redirect("finance:dashboard")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    budget_lines = get_filtered_budget_lines(request)
    analytics_data = build_finance_analytics_data(budget_lines)
    selected_table = request.GET.get("table", "all")

    wb = Workbook()
    ws = wb.active
    ws.title = "Finance Analytics"

    ws.append(["Finance Analytics Report"])
    ws.append([
        "Generated At",
        timezone.now().strftime("%d/%m/%Y %H:%M"),
    ])
    ws.append([])

    def bold_last_row():
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    if selected_table in ["all", "monthly"]:
        ws.append(["Monthly Spending Trend"])
        bold_last_row()
        ws.append([
            "Month",
            "Proposed",
            "Approved",
            "Disbursed",
        ])
        bold_last_row()

        for month, data in analytics_data["monthly_data"].items():
            ws.append([
                month,
                float(data["proposed"]),
                float(data["approved"]),
                float(data["disbursed"]),
            ])

        ws.append([])

    if selected_table in ["all", "top"]:
        ws.append(["Top Consuming Tasks"])
        bold_last_row()
        ws.append([
            "#",
            "Task Code",
            "Task in MTEF",
            "Approved",
            "Disbursed",
            "Balance",
        ])
        bold_last_row()

        for index, task in enumerate(
            analytics_data["top_consuming_tasks"],
            start=1,
        ):
            ws.append([
                index,
                task["task_code"],
                str(task["task_in_mtef"]),
                float(task["approved"]),
                float(task["disbursed"]),
                float(task["balance"]),
            ])

        ws.append([])

    if selected_table in ["all", "remaining"]:
        ws.append(["Remaining Funds Per Task"])
        bold_last_row()
        ws.append([
            "#",
            "Task Code",
            "Task in MTEF",
            "Approved",
            "Disbursed",
            "Remaining",
        ])
        bold_last_row()

        for index, task in enumerate(
            analytics_data["remaining_funds_tasks"],
            start=1,
        ):
            ws.append([
                index,
                task["task_code"],
                str(task["task_in_mtef"]),
                float(task["approved"]),
                float(task["disbursed"]),
                float(task["balance"]),
            ])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for column in ws.columns:
        max_length = max(
            len(str(cell.value))
            if cell.value is not None
            else 0
            for cell in column
        )
        ws.column_dimensions[
            get_column_letter(column[0].column)
        ].width = min(max_length + 2, 45)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        'attachment; filename="finance_analytics.xlsx"'
    )
    return response


@login_required
def export_finance_analytics_pdf(request):
    if not can_view_finance_reports(request.user):
        messages.error(
            request,
            "You are not allowed to export finance analytics.",
        )
        return redirect("finance:dashboard")

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import (
        ParagraphStyle,
        getSampleStyleSheet,
    )
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=20,
        bottomMargin=20,
    )

    styles = getSampleStyleSheet()
    elements = []

    budget_lines = get_filtered_budget_lines(request)
    analytics_data = build_finance_analytics_data(budget_lines)
    selected_table = request.GET.get("table", "all")

    title_style = styles["Title"]
    normal_style = styles["Normal"]

    wrap_style = ParagraphStyle(
        "WrapStyle",
        parent=styles["Normal"],
        fontSize=7,
        leading=8,
    )

    elements.append(
        Paragraph("Finance Analytics Report", title_style)
    )
    elements.append(
        Paragraph(
            (
                "Generated At: "
                f"{timezone.now().strftime('%d/%m/%Y %H:%M')}"
            ),
            normal_style,
        )
    )

    financial_year = request.GET.get(
        "financial_year",
        "",
    ).strip()
    month_filter = request.GET.get("month", "all")
    table_filter = request.GET.get("table", "all")

    elements.append(
        Paragraph(
            f"Financial Year: {financial_year or 'All'}",
            normal_style,
        )
    )
    elements.append(
        Paragraph(
            f"Month: {month_filter}",
            normal_style,
        )
    )
    elements.append(
        Paragraph(
            f"Report Table: {table_filter}",
            normal_style,
        )
    )
    elements.append(Spacer(1, 12))

    def add_table(title, data, column_widths=None):
        elements.append(Paragraph(title, styles["Heading2"]))

        if column_widths is None:
            column_count = len(data[0])
            column_widths = [
                doc.width / column_count
            ] * column_count

        table = Table(
            data,
            colWidths=column_widths,
            repeatRows=1,
        )

        table.setStyle(TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#e9f1ff"),
            ),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold",
            ),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("LEADING", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("ALIGN", (-3, 1), (-1, -1), "RIGHT"),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 10))

    if selected_table in ["all", "monthly"]:
        rows = [[
            "Month",
            "Proposed",
            "Approved",
            "Disbursed",
        ]]

        for month_name, data in analytics_data[
            "monthly_data"
        ].items():
            rows.append([
                month_name,
                f"{data['proposed']:,.2f}",
                f"{data['approved']:,.2f}",
                f"{data['disbursed']:,.2f}",
            ])

        add_table(
            "Monthly Spending Trend",
            rows,
            [120, 120, 120, 120],
        )

    if selected_table in ["all", "top"]:
        rows = [[
            "#",
            "Task Code",
            "Task in MTEF",
            "Approved",
            "Disbursed",
            "Balance",
        ]]

        for index, task in enumerate(
            analytics_data["top_consuming_tasks"],
            start=1,
        ):
            rows.append([
                index,
                task["task_code"],
                Paragraph(
                    str(task["task_in_mtef"]),
                    wrap_style,
                ),
                f"{task['approved']:,.2f}",
                f"{task['disbursed']:,.2f}",
                f"{task['balance']:,.2f}",
            ])

        add_table(
            "Top Consuming Tasks",
            rows,
            [30, 80, 260, 90, 90, 90],
        )

    if selected_table in ["all", "remaining"]:
        rows = [[
            "#",
            "Task Code",
            "Task in MTEF",
            "Approved",
            "Disbursed",
            "Remaining",
        ]]

        for index, task in enumerate(
            analytics_data["remaining_funds_tasks"],
            start=1,
        ):
            rows.append([
                index,
                task["task_code"],
                Paragraph(
                    str(task["task_in_mtef"]),
                    wrap_style,
                ),
                f"{task['approved']:,.2f}",
                f"{task['disbursed']:,.2f}",
                f"{task['balance']:,.2f}",
            ])

        add_table(
            "Remaining Funds Per Task",
            rows,
            [30, 80, 260, 90, 90, 90],
        )

    def add_page_number(canvas, document):
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(
            document.pagesize[0] - 20,
            15,
            f"Page {canvas.getPageNumber()}",
        )

    doc.build(
        elements,
        onFirstPage=add_page_number,
        onLaterPages=add_page_number,
    )

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(
        pdf,
        content_type="application/pdf",
    )
    response["Content-Disposition"] = (
        'attachment; filename="finance_analytics.pdf"'
    )
    return response


@login_required
def minute_sheet_list(request):
    return render(
        request,
        "finance/minute_sheet_list.html",
        {
            "minute_sheets": get_visible_minute_sheets(
                request.user
            ),
            "role": get_user_role(request.user),
        },
    )


@login_required
def retirement_list(request):
    return render(
        request,
        "finance/retirement_list.html",
        {
            "retirements": get_visible_retirements(
                request.user
            ),
        },
    )


def is_retirement_director(user):
    return (
        user.is_superuser
        or get_user_role(user) in [
            "ADMIN",
            "SYSTEM_ADMIN",
            "DIRECTOR",
        ]
    )


@login_required
def retirement_detail(request, pk):
    retirement = get_object_or_404(
        get_visible_retirements(request.user),
        pk=pk,
    )

    role = get_user_role(request.user)
    if request.method == "POST" and is_executive_viewer(request.user):
        raise PermissionDenied("Executive access to Finance is read-only.")

    can_accountant_review = (
        is_accountant(request.user)
        and retirement.status == "SUBMITTED"
    )

    can_director_review = (
        is_retirement_director(request.user)
        and retirement.status == "VERIFIED"
    )

    if request.method == "POST":
        action = request.POST.get("action", "").strip().lower()
        remarks = request.POST.get("remarks", "").strip()

        if retirement.status == "SUBMITTED":
            if not is_accountant(request.user):
                messages.error(
                    request,
                    "Only the Accountant can review this Retirement.",
                )
                return redirect("finance:retirement_detail", pk=retirement.pk)

            if action == "verify":
                retirement.status = "VERIFIED"
                retirement.verified_by = request.user
                retirement.verified_at = timezone.now()
                retirement.save()
                messages.success(
                    request,
                    "Retirement verified and forwarded to the Director.",
                )

            elif action == "return":
                if not remarks:
                    messages.error(request, "A return reason is required.")
                    return redirect("finance:retirement_detail", pk=retirement.pk)

                retirement.status = "RETURNED"
                retirement.save()
                messages.success(
                    request,
                    "Retirement returned to the requester for correction.",
                )

            elif action == "reject":
                if not remarks:
                    messages.error(request, "A rejection reason is required.")
                    return redirect("finance:retirement_detail", pk=retirement.pk)

                retirement.status = "REJECTED"
                retirement.save()
                messages.success(
                    request,
                    "Retirement rejected by the Accountant.",
                )

            else:
                messages.error(request, "Select Verify, Return, or Reject.")

            return redirect("finance:retirement_detail", pk=retirement.pk)

        if retirement.status == "VERIFIED":
            if not is_retirement_director(request.user):
                messages.error(
                    request,
                    "Only the Director can give final Retirement approval.",
                )
                return redirect("finance:retirement_detail", pk=retirement.pk)

            if action == "approve":
                retirement.status = "ACCEPTED"
                retirement.approved_by = request.user
                retirement.approved_at = timezone.now()
                retirement.save()

                retirement.finance_request.status = "RETIRED"
                retirement.finance_request.save(
                    update_fields=["status", "updated_at"]
                )

                messages.success(
                    request,
                    "Retirement approved successfully.",
                )

            elif action == "return":
                if not remarks:
                    messages.error(request, "A return reason is required.")
                    return redirect("finance:retirement_detail", pk=retirement.pk)

                retirement.status = "RETURNED"
                retirement.save()
                messages.success(
                    request,
                    "Retirement returned to the requester for correction.",
                )

            elif action == "reject":
                if not remarks:
                    messages.error(request, "A rejection reason is required.")
                    return redirect("finance:retirement_detail", pk=retirement.pk)

                retirement.status = "REJECTED"
                retirement.save()
                messages.success(
                    request,
                    "Retirement rejected by the Director.",
                )

            else:
                messages.error(request, "Select Approve, Return, or Reject.")

            return redirect("finance:retirement_detail", pk=retirement.pk)

        messages.error(
            request,
            "This Retirement is not currently awaiting approval.",
        )

    return render(
        request,
        "finance/retirement_detail.html",
        {
            "retirement": retirement,
            "role": role,
            "role_display": get_role_display_name(role),
            "can_accountant_review": can_accountant_review,
            "can_director_review": can_director_review,
        },
    )


@login_required
def retirement_create(request):
    if is_executive_viewer(request.user):
        raise PermissionDenied("Executive access to Finance is read-only.")
    if request.method == "POST":
        form = RetirementForm(
            request.POST,
            request.FILES,
            user=request.user,
        )

        if form.is_valid():
            retirement = form.save(commit=False)
            retirement.submitted_by = request.user

            approved_amount = (
                retirement.finance_request.approved_amount
                or Decimal("0.00")
            )
            accounted = (
                retirement.amount_accounted
                or Decimal("0.00")
            )

            if accounted > approved_amount:
                form.add_error(
                    "amount_accounted",
                    "Accounted amount cannot exceed approved amount.",
                )
            else:
                unaccounted = approved_amount - accounted
                retirement.amount_unaccounted = unaccounted
                retirement.refund_amount = unaccounted

                action = request.POST.get("action")

                if action == "submit":
                    retirement.status = "SUBMITTED"
                    retirement.submitted_at = timezone.now()
                else:
                    retirement.status = "DRAFT"
                    retirement.submitted_at = None

                retirement.save()

                for uploaded_file in request.FILES.getlist(
                    "supporting_files"
                ):
                    FinanceDocument.objects.create(
                        retirement=retirement,
                        document_type="RECEIPT",
                        file=uploaded_file,
                        uploaded_by=request.user,
                    )

                messages.success(
                    request,
                    (
                        "Retirement submitted successfully."
                        if action == "submit"
                        else "Retirement draft saved successfully."
                    ),
                )
                return redirect("finance:dashboard")
    else:
        form = RetirementForm(user=request.user)

    return render(
        request,
        "finance/retirement_form.html",
        {
            "form": form,
            "finance_requests": (
                form.fields["finance_request"].queryset
            ),
        },
    )


@login_required
def get_finance_request_amount(request):
    finance_request_id = request.GET.get(
        "finance_request_id"
    )

    finance_request = get_object_or_404(
        get_visible_finance_requests(request.user),
        pk=finance_request_id,
    )

    return JsonResponse({
        "approved_amount": str(
            finance_request.approved_amount
            or Decimal("0.00")
        ),
    })


@login_required
def edit_retirement(request, pk):
    if is_executive_viewer(request.user):
        raise PermissionDenied("Executive access to Finance is read-only.")
    retirement = get_object_or_404(
        Retirement.objects.filter(submitted_by=request.user),
        pk=pk,
    )

    if retirement.status not in ["DRAFT", "RETURNED"]:
        messages.error(
            request,
            "Only draft or returned retirements can be edited.",
        )
        return redirect(
            "finance:retirement_detail",
            pk=pk,
        )

    success = False

    if request.method == "POST":
        post_data = request.POST.copy()
        post_data["finance_request"] = (
            retirement.finance_request_id
        )

        form = RetirementForm(
            post_data,
            request.FILES,
            instance=retirement,
            user=request.user,
        )

        form.fields["finance_request"].queryset = (
            FinanceRequest.objects.filter(
                pk=retirement.finance_request_id
            )
        )

        if form.is_valid():
            updated_retirement = form.save(commit=False)

            approved_amount = (
                updated_retirement.finance_request.approved_amount
                or Decimal("0.00")
            )
            accounted = (
                updated_retirement.amount_accounted
                or Decimal("0.00")
            )

            if accounted > approved_amount:
                form.add_error(
                    "amount_accounted",
                    "Accounted amount cannot exceed approved amount.",
                )
            else:
                unaccounted = approved_amount - accounted
                updated_retirement.amount_unaccounted = unaccounted
                updated_retirement.refund_amount = unaccounted

                action = request.POST.get("action")

                if action == "submit":
                    updated_retirement.status = "SUBMITTED"
                    updated_retirement.submitted_at = timezone.now()
                    updated_retirement.verified_by = None
                    updated_retirement.verified_at = None
                    updated_retirement.approved_by = None
                    updated_retirement.approved_at = None
                    updated_retirement.save()

                    messages.success(
                        request,
                        "Retirement resubmitted successfully.",
                    )
                    return redirect(
                        "finance:retirement_detail",
                        pk=updated_retirement.pk,
                    )

                updated_retirement.status = "DRAFT"
                updated_retirement.submitted_at = None
                updated_retirement.save()
                success = True

                messages.success(
                    request,
                    "Retirement draft saved successfully.",
                )
    else:
        form = RetirementForm(
            instance=retirement,
            user=request.user,
        )
        form.fields["finance_request"].queryset = (
            FinanceRequest.objects.filter(
                pk=retirement.finance_request_id
            )
        )

    return render(
        request,
        "finance/edit_retirement.html",
        {
            "form": form,
            "retirement": retirement,
            "success": success,
        },
    )
