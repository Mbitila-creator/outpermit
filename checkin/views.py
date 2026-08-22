import csv
from io import BytesIO

from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.urls import reverse
from django.utils import timezone
from django.utils.translation import gettext as _
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from events.auth import User, has_event_role
from events.models import Event
from events.access import events_visible_to
from forms_builder.models import CertificateRecord, EventForm, FormSubmission
from forms_builder.models import NotificationLog
from forms_builder.notifications import send_submission_notification
from forms_builder.services import (
    answer_export_value,
    certificate_number,
    participant_certificate_path,
    public_form_path,
    safe_spreadsheet_value,
)

from .models import ParticipantCheckIn


CHECK_IN_ROLES = {
    User.Role.SYSTEM_ADMIN,
    User.Role.EVENT_ADMIN,
    User.Role.REGISTRATION_OFFICER,
    User.Role.ATTENDANCE_OFFICER,
}


def can_check_in(user):
    return bool(
        user.is_authenticated
        and (
            user.is_superuser
            or has_event_role(user, CHECK_IN_ROLES)
        )
    )


check_in_required = user_passes_test(
    can_check_in,
    login_url="login",
)


REPORT_ROLES = {
    User.Role.SYSTEM_ADMIN,
    User.Role.EVENT_ADMIN,
    User.Role.REPORT_OFFICER,
    User.Role.DIRECTOR,
    User.Role.ASSISTANT_DIRECTOR,
}


def can_view_reports(user):
    return bool(
        user.is_authenticated
        and (
            user.is_superuser
            or has_event_role(user, REPORT_ROLES)
        )
    )


report_required = user_passes_test(
    can_view_reports,
    login_url="login",
)


def can_authorize_certificates(user):
    return bool(
        user.is_authenticated
        and (
            user.is_superuser
            or has_event_role(user, {
                User.Role.SYSTEM_ADMIN,
                User.Role.EVENT_ADMIN,
            })
        )
    )


def approved_submission_queryset():
    return FormSubmission.objects.select_related(
        "event_form",
        "event_form__event",
        "event_form__event__venue",
    ).filter(
        review_status=FormSubmission.ReviewStatus.APPROVED,
        is_complete=True,
        is_active=True,
        event_form__event__qr_checkin_enabled=True,
        event_form__form_type__in=[
            EventForm.FormType.REGISTRATION,
            EventForm.FormType.EXHIBITOR,
            EventForm.FormType.SPEAKER,
        ],
    )


def report_submissions(event):
    return FormSubmission.objects.filter(
        event_form__event=event,
        is_active=True,
        is_complete=True,
        event_form__form_type__in=[
            EventForm.FormType.REGISTRATION,
            EventForm.FormType.EXHIBITOR,
            EventForm.FormType.SPEAKER,
        ],
    ).select_related(
        "event_form__event",
        "check_in",
        "certificate_record",
    )


def participant_report_rows(event):
    return report_submissions(event).order_by("badge_name", "reference_number")


@report_required
@require_http_methods(["GET", "POST"])
def attendance_reports(request):
    events = events_visible_to(request.user).annotate(
        registration_total=Count(
            "forms__submissions",
            filter=Q(
                forms__submissions__is_active=True,
                forms__submissions__is_complete=True,
            ),
            distinct=True,
        )
    ).order_by("-starts_at")
    selected_event = None
    request_data = request.POST if request.method == "POST" else request.GET
    selected_event_id = request_data.get("event", "").strip()
    selected_filter = request_data.get("filter", "all").strip()
    available_filters = {
        "all",
        "approved",
        "checked_in",
        "not_checked_in",
        "attendance_rate",
        "certificate_eligible",
        "certificate_review",
        "pending",
        "rejected",
    }
    if selected_filter not in available_filters:
        selected_filter = "all"

    if selected_event_id:
        selected_event = get_object_or_404(events, pk=selected_event_id)
    else:
        selected_event = events.first()

    if request.method == "POST":
        if not selected_event or not can_authorize_certificates(request.user):
            raise PermissionDenied
        if not selected_event.certificate_enabled:
            messages.error(request, _("Certificates are not enabled for this event."))
        else:
            selected_ids = request.POST.getlist("submission")
            eligible = report_submissions(selected_event).filter(
                pk__in=selected_ids,
                review_status=FormSubmission.ReviewStatus.APPROVED,
                check_in__isnull=False,
            ).exclude(
                certificate_record__status=CertificateRecord.Status.AUTHORIZED,
            )
            authorized = 0
            with transaction.atomic():
                for submission in eligible:
                    _record, _record_created = CertificateRecord.objects.update_or_create(
                        submission=submission,
                        defaults={
                            "certificate_number": certificate_number(submission),
                            "status": CertificateRecord.Status.AUTHORIZED,
                            "authorized_by": request.user,
                            "authorized_at": timezone.now(),
                            "denied_by": None,
                            "denied_at": None,
                            "denial_reason": "",
                            "revoked_by": None,
                            "revoked_at": None,
                            "revocation_reason": "",
                        },
                    )
                    send_submission_notification(
                        submission,
                        NotificationLog.NotificationType.CERTIFICATE_AUTHORIZED,
                        request=request,
                    )
                    authorized += 1
            if authorized:
                messages.success(
                    request,
                    _("%(count)s certificate(s) authorized successfully.")
                    % {"count": authorized},
                )
            else:
                messages.info(request, _("No eligible participants were selected."))
        return redirect(
            f"{reverse('checkin:reports')}?event={selected_event.pk}"
            "&filter=certificate_review"
        )

    summary = None
    rows = FormSubmission.objects.none()
    if selected_event:
        rows = report_submissions(selected_event)
        registered = rows.count()
        approved = rows.filter(
            review_status=FormSubmission.ReviewStatus.APPROVED
        ).count()
        checked_in = rows.filter(check_in__isnull=False).count()
        approved_checked_in = rows.filter(
            review_status=FormSubmission.ReviewStatus.APPROVED,
            check_in__isnull=False,
        ).count()
        certificate_authorized = rows.filter(
            review_status=FormSubmission.ReviewStatus.APPROVED,
            check_in__isnull=False,
            certificate_record__status=CertificateRecord.Status.AUTHORIZED,
        ).count()
        certificate_review = rows.filter(
            review_status=FormSubmission.ReviewStatus.APPROVED,
            check_in__isnull=False,
        ).exclude(
            certificate_record__status=CertificateRecord.Status.AUTHORIZED,
        ).count()
        approved_not_checked_in = rows.filter(
            review_status=FormSubmission.ReviewStatus.APPROVED,
            check_in__isnull=True,
        ).count()
        summary = {
            "registered": registered,
            "approved": approved,
            "pending": rows.filter(
                review_status=FormSubmission.ReviewStatus.PENDING
            ).count(),
            "rejected": rows.filter(
                review_status=FormSubmission.ReviewStatus.REJECTED
            ).count(),
            "checked_in": checked_in,
            "not_checked_in": approved_not_checked_in,
            "certificate_eligible": (
                certificate_authorized
                if selected_event.certificate_enabled
                else 0
            ),
            "certificate_review": (
                certificate_review
                if selected_event.certificate_enabled
                else 0
            ),
            "attendance_rate": (
                round((approved_checked_in / approved) * 100, 1)
                if approved
                else 0
            ),
        }
        row_filters = {
            "approved": Q(
                review_status=FormSubmission.ReviewStatus.APPROVED
            ),
            "checked_in": Q(check_in__isnull=False),
            "attendance_rate": Q(
                review_status=FormSubmission.ReviewStatus.APPROVED,
                check_in__isnull=False,
            ),
            "not_checked_in": Q(
                review_status=FormSubmission.ReviewStatus.APPROVED,
                check_in__isnull=True,
            ),
            "pending": Q(
                review_status=FormSubmission.ReviewStatus.PENDING
            ),
            "rejected": Q(
                review_status=FormSubmission.ReviewStatus.REJECTED
            ),
        }
        if selected_filter == "certificate_eligible":
            rows = (
                rows.filter(
                    review_status=FormSubmission.ReviewStatus.APPROVED,
                    check_in__isnull=False,
                    certificate_record__status=CertificateRecord.Status.AUTHORIZED,
                )
                if selected_event.certificate_enabled
                else rows.none()
            )
        elif selected_filter == "certificate_review":
            rows = (
                rows.filter(
                    review_status=FormSubmission.ReviewStatus.APPROVED,
                    check_in__isnull=False,
                ).exclude(
                    certificate_record__status=CertificateRecord.Status.AUTHORIZED,
                )
                if selected_event.certificate_enabled
                else rows.none()
            )
        elif selected_filter in row_filters:
            rows = rows.filter(row_filters[selected_filter])

        rows = rows.order_by("badge_name", "reference_number")

    return render(
        request,
        "checkin/reports.html",
        {
            "events": events,
            "selected_event": selected_event,
            "summary": summary,
            "rows": rows,
            "selected_filter": selected_filter,
            "filtered_total": rows.count() if selected_event else 0,
            "can_authorize_certificates": can_authorize_certificates(request.user),
        },
    )


@report_required
@require_http_methods(["GET"])
def attendance_report_csv(request):
    event = get_object_or_404(
        events_visible_to(request.user), pk=request.GET.get("event")
    )
    report_type = request.GET.get("report", "attendance")
    rows = report_submissions(event).order_by(
        "badge_name",
        "reference_number",
    )

    if report_type == "certificates":
        rows = rows.filter(
            review_status=FormSubmission.ReviewStatus.APPROVED,
            check_in__isnull=False,
            certificate_record__status=CertificateRecord.Status.AUTHORIZED,
        )

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response.write("\ufeff")
    response["Content-Disposition"] = (
        f'attachment; filename="{event.code}-{report_type}-report.csv"'
    )
    writer = csv.writer(response)
    headers = [
        _("Reference number"),
        _("Representative"),
        _("Organization"),
        _("Review status"),
        _("Checked in"),
        _("Checked in at"),
        _("Checked in by"),
    ]
    if report_type == "certificates":
        headers.append(_("Certificate number"))
    writer.writerow(headers)

    for submission in rows:
        check_in = getattr(submission, "check_in", None)
        values = [
            submission.reference_number,
            submission.badge_display_name,
            submission.badge_organization,
            submission.get_review_status_display(),
            _("Yes") if check_in else _("No"),
            check_in.checked_in_at.isoformat() if check_in else "",
            str(check_in.checked_in_by) if check_in else "",
        ]
        if report_type == "certificates":
            values.append(certificate_number(submission))
        writer.writerow([safe_spreadsheet_value(value) for value in values])

    return response


@report_required
@require_http_methods(["GET"])
def participant_list_print(request):
    event = get_object_or_404(
        events_visible_to(request.user).select_related("venue"),
        pk=request.GET.get("event"),
    )
    rows = participant_report_rows(event)
    return render(request, "checkin/participant_list_print.html", {
        "event": event,
        "rows": rows,
        "generated_at": timezone.localtime(),
    })


@report_required
@require_http_methods(["GET"])
def participant_list_excel(request):
    event = get_object_or_404(
        events_visible_to(request.user).select_related("venue"),
        pk=request.GET.get("event"),
    )
    rows = participant_report_rows(event)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Registered participants"
    navy = "173B67"
    teal = "087F73"
    white = "FFFFFF"

    sheet.merge_cells("A1:H1")
    sheet["A1"] = event.title_en
    sheet["A1"].font = Font(size=16, bold=True, color=white)
    sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:H2")
    sheet["A2"] = f"{event.code} — REGISTERED PARTICIPANTS"
    sheet["A2"].font = Font(size=12, bold=True, color=teal)
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.append([])
    headers = [
        "#", "Reference", "Representative", "Institution", "Email", "Phone",
        "Attendance", "Certificate",
    ]
    sheet.append(headers)
    for cell in sheet[4]:
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=teal)
    for number, submission in enumerate(rows, start=1):
        check_in = getattr(submission, "check_in", None)
        certificate = getattr(submission, "certificate_record", None)
        sheet.append([
            number,
            safe_spreadsheet_value(submission.reference_number),
            safe_spreadsheet_value(submission.badge_display_name),
            safe_spreadsheet_value(submission.badge_organization),
            safe_spreadsheet_value(submission.submitter_email),
            safe_spreadsheet_value(submission.submitter_phone),
            "Checked in" if check_in else "Not checked in",
            (
                certificate.get_status_display()
                if certificate
                else "Awaiting check-in/authorization"
            ),
        ])
    widths = [7, 25, 28, 32, 32, 20, 18, 28]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{event.code}-registered-participants.xlsx"'
    )
    return response


@report_required
@require_http_methods(["GET", "POST"])
def participant_staff_detail(request, submission_id):
    submission = get_object_or_404(
        FormSubmission.objects.select_related(
            "event_form__event",
            "event_form__event__venue",
            "check_in",
            "certificate_record",
        ).prefetch_related(
            "answers__question__section",
            "answers__selected_options",
        ),
        pk=submission_id,
        event_form__event__in=events_visible_to(request.user),
        is_active=True,
        is_complete=True,
        event_form__form_type__in=[
            EventForm.FormType.REGISTRATION,
            EventForm.FormType.EXHIBITOR,
            EventForm.FormType.SPEAKER,
        ],
    )
    event = submission.event_form.event
    certificate_record = getattr(submission, "certificate_record", None)

    if request.method == "POST":
        if not can_authorize_certificates(request.user):
            raise PermissionDenied
        action = request.POST.get("action", "").strip()
        if action == "authorize":
            if not event.certificate_enabled:
                messages.error(request, _("Certificates are not enabled for this event."))
            elif not getattr(submission, "check_in", None):
                messages.error(
                    request,
                    _("The participant must check in before certificate authorization."),
                )
            else:
                _record, _created = CertificateRecord.objects.update_or_create(
                    submission=submission,
                    defaults={
                        "certificate_number": certificate_number(submission),
                        "status": CertificateRecord.Status.AUTHORIZED,
                        "authorized_by": request.user,
                        "authorized_at": timezone.now(),
                        "denied_by": None,
                        "denied_at": None,
                        "denial_reason": "",
                        "revoked_by": None,
                        "revoked_at": None,
                        "revocation_reason": "",
                    },
                )
                send_submission_notification(
                    submission,
                    NotificationLog.NotificationType.CERTIFICATE_AUTHORIZED,
                    request=request,
                )
                messages.success(request, _("Certificate authorized successfully."))
        elif action == "revoke":
            reason = request.POST.get("reason", "").strip()
            if not certificate_record or certificate_record.status != CertificateRecord.Status.AUTHORIZED:
                messages.error(request, _("Only an authorized certificate can be revoked."))
            elif not reason:
                messages.error(request, _("Enter a reason before revoking the certificate."))
            else:
                certificate_record.status = CertificateRecord.Status.REVOKED
                certificate_record.revoked_by = request.user
                certificate_record.revoked_at = timezone.now()
                certificate_record.revocation_reason = reason
                certificate_record.denied_by = None
                certificate_record.denied_at = None
                certificate_record.denial_reason = ""
                certificate_record.save()
                messages.success(request, _("Certificate revoked successfully."))
        else:
            messages.error(request, _("Select a valid certificate action."))
        return redirect(
            "checkin:participant_staff_detail",
            submission_id=submission.pk,
        )

    answers = []
    for answer in submission.answers.all():
        value = answer_export_value(answer)
        if value not in (None, ""):
            answers.append({
                "section": (
                    answer.question.section.title_en
                    if submission.language == "en"
                    else answer.question.section.title_sw
                ),
                "question": (
                    answer.question.label_en
                    if submission.language == "en"
                    else answer.question.label_sw
                ),
                "value": value,
            })
    return render(request, "checkin/participant_staff_detail.html", {
        "submission": submission,
        "event": event,
        "check_in": getattr(submission, "check_in", None),
        "certificate_record": certificate_record,
        "answers": answers,
        "can_authorize_certificates": can_authorize_certificates(request.user),
    })


@check_in_required
@require_http_methods(["GET", "POST"])
def check_in_lookup(request):
    lookup_error = ""
    selected_event_id = request.GET.get("event", "").strip()
    selected_event = None
    if selected_event_id:
        selected_event = get_object_or_404(
            events_visible_to(request.user), pk=selected_event_id
        )

    if request.method == "POST":
        selected_event_id = request.POST.get("event", "").strip()
        if selected_event_id:
            selected_event = get_object_or_404(
                events_visible_to(request.user), pk=selected_event_id
            )
        identifier = request.POST.get("identifier", "").strip()
        submissions = approved_submission_queryset().filter(
            event_form__event__in=events_visible_to(request.user),
        )
        if selected_event:
            submissions = submissions.filter(event_form__event=selected_event)
        submission = submissions.filter(
            reference_number__iexact=identifier
        ).first()

        if submission is None:
            try:
                submission = submissions.filter(
                    participant_token=identifier
                ).first()
            except (TypeError, ValueError):
                submission = None

        if submission:
            return redirect(
                "checkin:participant",
                participant_token=submission.participant_token,
            )

        lookup_error = "Participant not found or not eligible for check-in."

    return render(
        request,
        "checkin/lookup.html",
        {"lookup_error": lookup_error, "selected_event": selected_event},
    )


@check_in_required
@require_http_methods(["GET", "POST"])
def participant_check_in(request, participant_token):
    submission = get_object_or_404(
        FormSubmission.objects.select_related(
            "event_form",
            "event_form__event",
            "event_form__event__venue",
        ),
        participant_token=participant_token,
        event_form__event__in=events_visible_to(request.user),
    )
    is_eligible = bool(
        submission.review_status
        == FormSubmission.ReviewStatus.APPROVED
        and submission.event_form.form_type
        in {
            EventForm.FormType.REGISTRATION,
            EventForm.FormType.EXHIBITOR,
            EventForm.FormType.SPEAKER,
        }
        and submission.is_complete
        and submission.is_active
        and submission.event_form.event.qr_checkin_enabled
    )
    check_in = ParticipantCheckIn.objects.filter(
        submission=submission
    ).select_related("checked_in_by").first()
    just_checked_in = False
    automatic_check_in = (
        request.method == "GET" and request.GET.get("auto") == "1"
    )

    if (
        (request.method == "POST" or automatic_check_in)
        and check_in is None
        and is_eligible
    ):
        with transaction.atomic():
            locked_submission = (
                approved_submission_queryset()
                .select_for_update()
                .get(pk=submission.pk)
            )
            check_in, just_checked_in = (
                ParticipantCheckIn.objects.get_or_create(
                    submission=locked_submission,
                    defaults={
                        "checked_in_by": request.user,
                        "method": ParticipantCheckIn.Method.QR,
                        "created_by": request.user,
                        "updated_by": request.user,
                    },
                )
            )

    if just_checked_in:
        send_submission_notification(
            submission,
            NotificationLog.NotificationType.CHECK_IN_CONFIRMED,
            request=request,
        )

    evaluation_form = None
    if submission.event_form.event.evaluation_enabled:
        evaluation_form = (
            submission.event_form.event.forms.filter(
                form_type=EventForm.FormType.EVALUATION,
                is_active=True,
                is_published=True,
            )
            .order_by("id")
            .first()
        )

    return render(
        request,
        "checkin/participant_check_in.html",
        {
            "submission": submission,
            "event": submission.event_form.event,
            "check_in": check_in,
            "just_checked_in": just_checked_in,
            "is_eligible": is_eligible,
            "certificate_path": (
                participant_certificate_path(
                    submission,
                    language=request.LANGUAGE_CODE,
                )
                if (
                    check_in
                    and hasattr(submission, "certificate_record")
                    and submission.certificate_record.status
                    == CertificateRecord.Status.AUTHORIZED
                    and submission.event_form.event.certificate_enabled
                )
                else ""
            ),
            "evaluation_path": (
                public_form_path(
                    evaluation_form,
                    language=request.LANGUAGE_CODE,
                )
                if check_in and evaluation_form
                else ""
            ),
        },
    )
