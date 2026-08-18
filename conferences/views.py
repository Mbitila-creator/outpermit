import csv
import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Avg, Count, Prefetch, Q
from django.core.mail import send_mail
from django.http import FileResponse, Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.utils.translation import gettext as _
from django.views.decorators.http import require_GET, require_http_methods, require_POST

from events.auth import User
from events.models import Event
from events.access import events_visible_to
from forms_builder.models import EventForm, FormSubmission, NotificationLog
from forms_builder.notifications import send_submission_notification
from forms_builder.services import (
    generate_qr_png,
    public_form_path,
    public_form_url,
    safe_spreadsheet_value,
)

from .forms import (
    ConferencePaperSubmissionForm,
    ConferencePeerReviewForm,
    ConferencePresentationConfirmationForm,
    ConferencePresentationScheduleForm,
    ConferencePaperCommunicationForm,
    ConferenceFeedbackForm,
)
from .services import generate_programme_pdf

from .models import (
    ConferenceCallForPapers,
    ConferencePaper,
    ConferencePaperReview,
    ConferencePaperReviewAssignment,
    ConferenceProgrammeContributor,
    ConferenceProgrammeItem,
    ConferencePresentation,
    ConferencePaperCommunication,
    ConferenceCertificate,
    ConferenceFeedback,
    ConferenceReviewer,
    ConferenceSession,
    ConferenceSessionAttendance,
    ConferenceGuidingResponse,
    ConferenceGuidingSubmission,
)


CONFERENCE_VIEW_ROLES = {
    User.Role.SYSTEM_ADMIN,
    User.Role.EVENT_ADMIN,
    User.Role.REGISTRATION_OFFICER,
    User.Role.ATTENDANCE_OFFICER,
    User.Role.REPORT_OFFICER,
    User.Role.DIRECTOR,
    User.Role.ASSISTANT_DIRECTOR,
}
CONFERENCE_MANAGER_ROLES = {
    User.Role.SYSTEM_ADMIN,
    User.Role.EVENT_ADMIN,
    User.Role.REGISTRATION_OFFICER,
}
CONFERENCE_CHECKIN_ROLES = CONFERENCE_MANAGER_ROLES | {
    User.Role.ATTENDANCE_OFFICER,
}
UUID_PATTERN = re.compile(
    r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[1-5][0-9a-fA-F]{3}-"
    r"[89abAB][0-9a-fA-F]{3}-[0-9a-fA-F]{12}"
)


def _require_access(user):
    if not (
        user.is_authenticated
        and user.is_active
        and (user.is_superuser or user.role in CONFERENCE_VIEW_ROLES)
    ):
        raise PermissionDenied


def _require_manager(user):
    if not (
        user.is_authenticated
        and user.is_active
        and (user.is_superuser or user.role in CONFERENCE_MANAGER_ROLES)
    ):
        raise PermissionDenied


def _require_checkin(user):
    if not (
        user.is_authenticated
        and user.is_active
        and (user.is_superuser or user.role in CONFERENCE_CHECKIN_ROLES)
    ):
        raise PermissionDenied


def _certificate_number(event, recipient_type, source_pk):
    prefixes = {
        ConferenceCertificate.RecipientType.PARTICIPANT: "P",
        ConferenceCertificate.RecipientType.PRESENTER: "PR",
        ConferenceCertificate.RecipientType.REVIEWER: "R",
    }
    return f"{event.code}-CERT-{prefixes[recipient_type]}-{source_pk:05d}"


def _certificate_defaults(recipient_type, source):
    if recipient_type == ConferenceCertificate.RecipientType.PARTICIPANT:
        return source.badge_display_name, source.badge_organization or ""
    if recipient_type == ConferenceCertificate.RecipientType.PRESENTER:
        return source.presentation.presenter_name, source.institution
    user = source.user
    return user.get_full_name() or user.username, source.institution


def _eligible_certificate_sources(event_form, recipient_type):
    event = event_form.event
    if recipient_type == ConferenceCertificate.RecipientType.PARTICIPANT:
        return FormSubmission.objects.filter(
            event_form=event_form,
            is_active=True,
            is_complete=True,
            review_status=FormSubmission.ReviewStatus.APPROVED,
            conference_session_attendance__session__event=event,
            conference_session_attendance__is_active=True,
        ).distinct()
    if recipient_type == ConferenceCertificate.RecipientType.PRESENTER:
        return ConferencePaper.objects.filter(
            call__event=event,
            is_active=True,
            presentation__is_active=True,
            presentation__status=ConferencePresentation.Status.DELIVERED,
        ).select_related("presentation")
    return ConferenceReviewer.objects.filter(
        event=event,
        is_active=True,
        assignments__is_active=True,
        assignments__status=ConferencePaperReviewAssignment.Status.COMPLETED,
    ).select_related("user").distinct()


def _conference_registration_forms(user):
    visible_events = events_visible_to(user)
    return (
        EventForm.objects.filter(
            Q(event__category__code__iexact="CONFERENCE")
            | Q(event__category__name_en__iexact="Conference")
            | Q(event__category__slug__in=("conference", "kongamano")),
            event__in=visible_events,
            form_type=EventForm.FormType.REGISTRATION,
            is_active=True,
        )
        .select_related("event", "event__venue")
        .order_by("-event__starts_at", "event__code")
    )


def _public_conference(event_slug):
    event = get_object_or_404(
        Event.objects.select_related("category", "venue"),
        slug=event_slug,
        is_active=True,
        is_public=True,
    )
    if not event.category.is_conference:
        raise PermissionDenied
    return event


def _selected_submissions(event_form, session, approved_only=True):
    queryset = FormSubmission.objects.filter(
        event_form=event_form,
        is_active=True,
        is_complete=True,
        answers__question__section__event_form=event_form,
        answers__selected_options__value=session.registration_option_value,
        answers__selected_options__is_active=True,
    )
    if approved_only:
        queryset = queryset.filter(
            review_status=FormSubmission.ReviewStatus.APPROVED,
        )
    return queryset.distinct()


def _submission_session_values(submission):
    return set(
        submission.answers.filter(
            selected_options__isnull=False,
        ).values_list("selected_options__value", flat=True)
    )


@require_GET
def public_programme(request, event_slug):
    event = _public_conference(event_slug)

    contributors = ConferenceProgrammeContributor.objects.filter(
        is_active=True,
        speaker__is_active=True,
    ).select_related("speaker")
    programme_items = ConferenceProgrammeItem.objects.filter(
        is_active=True,
        is_published=True,
    ).prefetch_related(
        Prefetch("contributors", queryset=contributors),
    )
    sessions = (
        ConferenceSession.objects.filter(event=event, is_active=True)
        .prefetch_related(
            Prefetch("programme_items", queryset=programme_items),
            Prefetch(
                "paper_presentations",
                queryset=ConferencePresentation.objects.exclude(
                    status=ConferencePresentation.Status.CANCELLED,
                ).select_related("paper"),
                to_attr="public_presentations",
            ),
        )
        .order_by("starts_at", "display_order")
    )
    return render(
        request,
        "conferences/public_programme.html",
        {"event": event, "sessions": sessions},
    )


@require_GET
def participant_guiding_questions(request, participant_token):
    submission = get_object_or_404(
        FormSubmission.objects.select_related("event_form__event"),
        participant_token=participant_token,
        is_active=True,
        is_complete=True,
    )
    selected_values = set(
        submission.answers.filter(selected_options__is_active=True).values_list(
            "selected_options__value", flat=True
        )
    )
    sessions = list(
        ConferenceSession.objects.filter(
            event=submission.event_form.event,
            registration_option_value__in=selected_values,
            is_active=True,
        ).order_by("starts_at", "display_order", "id")
    )
    progress = {
        item.session_id: item
        for item in submission.conference_guiding_submissions.filter(
            session_id__in=[session.pk for session in sessions]
        )
    }
    session_rows = [
        {"session": session, "progress": progress.get(session.pk)}
        for session in sessions
    ]
    session_query = "&".join(f"session={session.pk}" for session in sessions)
    programme_url = reverse(
        "conferences:public_programme",
        kwargs={"event_slug": submission.event_form.event.slug},
    )
    return render(request, "conferences/participant_guiding_questions.html", {
        "submission": submission,
        "event": submission.event_form.event,
        "session_rows": session_rows,
        "programme_url": f"{programme_url}?{session_query}",
    })


@require_http_methods(["GET", "POST"])
def participant_session_guiding_questions(request, participant_token, session_id):
    submission = get_object_or_404(
        FormSubmission.objects.select_related("event_form__event"),
        participant_token=participant_token,
        is_active=True,
        is_complete=True,
    )
    selected_values = submission.answers.filter(
        selected_options__is_active=True
    ).values_list("selected_options__value", flat=True)
    session = get_object_or_404(
        ConferenceSession.objects.prefetch_related("guiding_topics__questions"),
        pk=session_id,
        event=submission.event_form.event,
        registration_option_value__in=selected_values,
        is_active=True,
    )
    progress, progress_created = ConferenceGuidingSubmission.objects.get_or_create(
        submission=submission,
        session=session,
    )
    questions = [
        question
        for topic in session.guiding_topics.all()
        if topic.is_active
        for question in topic.questions.all()
        if question.is_active
    ]
    existing = {
        item.question_id: item
        for item in submission.conference_guiding_responses.filter(
            question_id__in=[question.pk for question in questions]
        )
    }

    if request.method == "POST" and progress.status != ConferenceGuidingSubmission.Status.SUBMITTED:
        with transaction.atomic():
            for question in questions:
                value = request.POST.get(f"question_{question.pk}", "").strip()
                if value:
                    ConferenceGuidingResponse.objects.update_or_create(
                        submission=submission,
                        question=question,
                        defaults={"response": value, "is_active": True},
                    )
                elif question.pk in existing:
                    existing[question.pk].delete()
            if request.POST.get("action") == "submit":
                progress.status = ConferenceGuidingSubmission.Status.SUBMITTED
                progress.submitted_at = timezone.now()
                message = _("Your responses for this session have been submitted.")
            else:
                progress.status = ConferenceGuidingSubmission.Status.DRAFT
                progress.submitted_at = None
                message = _("Draft saved. You can return later and continue.")
            progress.save(update_fields=["status", "submitted_at", "updated_at"])
        messages.success(request, message)
        return redirect(
            "conferences:participant_session_guiding_questions",
            participant_token=participant_token,
            session_id=session.pk,
        )

    response_values = {key: item.response for key, item in existing.items()}
    topic_rows = []
    for topic in session.guiding_topics.all():
        if not topic.is_active:
            continue
        topic_rows.append({
            "topic": topic,
            "questions": [
                {"question": question, "value": response_values.get(question.pk, "")}
                for question in topic.questions.all()
                if question.is_active
            ],
        })
    return render(request, "conferences/participant_session_guiding_questions.html", {
        "submission": submission,
        "event": submission.event_form.event,
        "session": session,
        "progress": progress,
        "topic_rows": topic_rows,
    })
@require_GET
def programme_download(request, event_slug):
    event = _public_conference(event_slug)
    session_ids = request.GET.getlist("session")

    if not session_ids or any(not value.isdigit() for value in session_ids):
        return HttpResponse("Select at least one valid session.", status=400)

    sessions = list(
        ConferenceSession.objects.filter(
            event=event,
            is_active=True,
            pk__in=session_ids,
        )
        .prefetch_related(
            Prefetch(
                "programme_items",
                queryset=ConferenceProgrammeItem.objects.filter(
                    is_active=True,
                    is_published=True,
                ).prefetch_related(
                    Prefetch(
                        "contributors",
                        queryset=ConferenceProgrammeContributor.objects.filter(
                            is_active=True,
                            speaker__is_active=True,
                        ).select_related("speaker"),
                    ),
                ),
            ),
        )
        .order_by("starts_at", "display_order")
    )

    if len(sessions) != len(set(session_ids)):
        return HttpResponse("One or more sessions are unavailable.", status=400)

    response = HttpResponse(
        generate_programme_pdf(event, sessions),
        content_type="application/pdf",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{event.slug}-timetable.pdf"'
    )
    response["X-Content-Type-Options"] = "nosniff"
    return response


@require_http_methods(["GET", "POST"])
def feedback_submit(request, event_slug):
    event = _public_conference(event_slug)
    if request.method == "POST":
        form = ConferenceFeedbackForm(request.POST, event=event)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.event = event
            feedback.save()
            return redirect(
                "conferences:feedback_thanks",
                public_token=feedback.public_token,
            )
    else:
        form = ConferenceFeedbackForm(event=event)
    return render(request, "conferences/feedback_form.html", {
        "event": event,
        "form": form,
    })


@require_GET
def feedback_thanks(request, public_token):
    feedback = get_object_or_404(
        ConferenceFeedback.objects.select_related("event", "session"),
        public_token=public_token,
        is_active=True,
    )
    return render(request, "conferences/feedback_thanks.html", {
        "feedback": feedback,
    })


@require_GET
def feedback_qr(request, event_slug):
    event = _public_conference(event_slug)
    feedback_url = request.build_absolute_uri(reverse(
        "conferences:feedback_submit",
        kwargs={"event_slug": event.slug},
    ))
    return HttpResponse(generate_qr_png(feedback_url), content_type="image/png")


def _public_call(event_slug):
    call = get_object_or_404(
        ConferenceCallForPapers.objects.select_related("event", "event__category"),
        event__slug=event_slug,
        event__is_active=True,
        event__is_public=True,
        is_active=True,
        is_published=True,
    )
    now = timezone.now()
    if (call.opens_at and now < call.opens_at) or (call.closes_at and now > call.closes_at):
        raise Http404("This call for papers is not open.")
    return call


@require_http_methods(["GET", "POST"])
def paper_submit(request, event_slug):
    call = _public_call(event_slug)
    if request.method == "POST":
        form = ConferencePaperSubmissionForm(request.POST, request.FILES)
        if form.is_valid():
            duplicate = ConferencePaper.objects.filter(
                call=call,
                email__iexact=form.cleaned_data["email"],
                title__iexact=form.cleaned_data["title"],
                is_active=True,
            ).exists()
            if duplicate:
                form.add_error(
                    "title",
                    _("This author has already submitted a paper with this title."),
                )
            else:
                paper = form.save(commit=False)
                paper.call = call
                paper.save()
                return redirect(
                    "conferences:paper_status",
                    public_token=paper.public_token,
                )
    else:
        form = ConferencePaperSubmissionForm()
    return render(request, "conferences/paper_submit.html", {"call": call, "form": form})


@require_GET
def paper_status(request, public_token):
    paper = get_object_or_404(
        ConferencePaper.objects.select_related(
            "call__event", "assigned_session", "presentation__session",
        ),
        public_token=public_token,
        is_active=True,
    )
    return render(request, "conferences/paper_status.html", {"paper": paper})


@require_GET
def paper_document(request, public_token):
    paper = get_object_or_404(ConferencePaper, public_token=public_token, is_active=True)
    if not paper.document:
        raise Http404("No document was uploaded.")
    return FileResponse(
        paper.document.open("rb"),
        as_attachment=True,
        filename=paper.document.name.rsplit("/", 1)[-1],
    )


@login_required
def paper_review_list(request, form_id):
    _require_access(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    papers = ConferencePaper.objects.filter(
        call__event=event_form.event,
        is_active=True,
    ).select_related("assigned_session", "reviewed_by")
    status_filter = request.GET.get("status", "").strip()
    if status_filter in ConferencePaper.Status.values:
        papers = papers.filter(status=status_filter)
    summary = {
        "total": ConferencePaper.objects.filter(call__event=event_form.event, is_active=True).count(),
        "submitted": ConferencePaper.objects.filter(call__event=event_form.event, is_active=True, status=ConferencePaper.Status.SUBMITTED).count(),
        "review": ConferencePaper.objects.filter(call__event=event_form.event, is_active=True, status=ConferencePaper.Status.UNDER_REVIEW).count(),
        "accepted": ConferencePaper.objects.filter(call__event=event_form.event, is_active=True, status=ConferencePaper.Status.ACCEPTED).count(),
    }
    return render(request, "conferences/paper_review_list.html", {
        "event_form": event_form,
        "papers": papers,
        "summary": summary,
        "status_filter": status_filter,
    })


@login_required
@require_http_methods(["GET", "POST"])
def paper_review(request, form_id, paper_id):
    _require_manager(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    paper = get_object_or_404(
        ConferencePaper.objects.select_related("call__event", "assigned_session"),
        pk=paper_id,
        call__event=event_form.event,
        is_active=True,
    )
    if request.method == "POST":
        decision = request.POST.get("decision", "")
        allowed = {
            ConferencePaper.Status.UNDER_REVIEW,
            ConferencePaper.Status.REVISION_REQUIRED,
            ConferencePaper.Status.ACCEPTED,
            ConferencePaper.Status.REJECTED,
        }
        if decision not in allowed:
            raise PermissionDenied
        message = request.POST.get("decision_message", "").strip()
        internal_notes = request.POST.get("internal_notes", "").strip()
        if decision in {ConferencePaper.Status.REVISION_REQUIRED, ConferencePaper.Status.REJECTED} and not message:
            messages.error(request, _("Enter a message to the author for this decision."))
        else:
            session_id = request.POST.get("assigned_session", "").strip()
            assigned_session = None
            if session_id:
                assigned_session = get_object_or_404(
                    ConferenceSession,
                    pk=session_id,
                    event=event_form.event,
                    is_active=True,
                )
            with transaction.atomic():
                paper.status = decision
                paper.decision_message = message
                paper.internal_notes = internal_notes
                paper.assigned_session = assigned_session
                paper.reviewed_by = request.user
                paper.reviewed_at = timezone.now()
                paper.updated_by = request.user
                paper.save()
                ConferencePaperReview.objects.create(
                    paper=paper,
                    decision=decision,
                    message_to_author=message,
                    internal_notes=internal_notes,
                    assigned_session=assigned_session,
                    reviewer=request.user,
                    created_by=request.user,
                    updated_by=request.user,
                )
            messages.success(request, _("The paper review decision was saved."))
            return redirect("conferences:paper_review_list", form_id=event_form.pk)
    sessions = ConferenceSession.objects.filter(event=event_form.event, is_active=True)
    reviewers = ConferenceReviewer.objects.filter(event=event_form.event, is_active=True).select_related("user")
    assignments = paper.peer_review_assignments.filter(is_active=True).select_related("reviewer__user")
    return render(request, "conferences/paper_review.html", {
        "event_form": event_form,
        "paper": paper,
        "sessions": sessions,
        "reviewers": reviewers,
        "assignments": assignments,
    })


@login_required
@require_POST
def assign_paper_reviewer(request, form_id, paper_id):
    _require_manager(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    paper = get_object_or_404(
        ConferencePaper, pk=paper_id, call__event=event_form.event, is_active=True,
    )
    reviewer = get_object_or_404(
        ConferenceReviewer,
        pk=request.POST.get("reviewer"),
        event=event_form.event,
        is_active=True,
    )
    due_at = parse_datetime(request.POST.get("due_at", ""))
    if due_at and timezone.is_naive(due_at):
        due_at = timezone.make_aware(due_at)
    assignment, created = ConferencePaperReviewAssignment.objects.get_or_create(
        paper=paper,
        reviewer=reviewer,
        defaults={
            "assigned_by": request.user,
            "due_at": due_at,
            "created_by": request.user,
            "updated_by": request.user,
        },
    )
    if not created:
        assignment.due_at = due_at
        assignment.updated_by = request.user
        assignment.save()
        messages.info(request, _("The existing reviewer assignment was updated."))
    else:
        if paper.status == ConferencePaper.Status.SUBMITTED:
            paper.status = ConferencePaper.Status.UNDER_REVIEW
            paper.updated_by = request.user
            paper.save()
        messages.success(request, _("Reviewer assigned successfully."))
    return redirect("conferences:paper_review", form_id=event_form.pk, paper_id=paper.pk)


@login_required
def reviewer_workspace(request):
    profiles = ConferenceReviewer.objects.filter(user=request.user, is_active=True)
    assignments = ConferencePaperReviewAssignment.objects.filter(
        reviewer__in=profiles,
        is_active=True,
    ).select_related("paper__call__event", "reviewer")
    return render(request, "conferences/reviewer_workspace.html", {
        "assignments": assignments,
        "summary": {
            "total": assignments.count(),
            "pending": assignments.exclude(status=ConferencePaperReviewAssignment.Status.COMPLETED).exclude(status=ConferencePaperReviewAssignment.Status.CONFLICT).count(),
            "completed": assignments.filter(status=ConferencePaperReviewAssignment.Status.COMPLETED).count(),
        },
    })


@login_required
@require_http_methods(["GET", "POST"])
def peer_review(request, assignment_id):
    assignment = get_object_or_404(
        ConferencePaperReviewAssignment.objects.select_related(
            "paper__call__event", "reviewer__user",
        ),
        pk=assignment_id,
        is_active=True,
    )
    is_manager = request.user.is_superuser or request.user.role in CONFERENCE_MANAGER_ROLES
    if assignment.reviewer.user_id != request.user.id and not is_manager:
        raise PermissionDenied
    if request.method == "POST":
        form = ConferencePeerReviewForm(request.POST, instance=assignment)
        if form.is_valid():
            assignment = form.save(commit=False)
            assignment.updated_by = request.user
            assignment.submitted_at = (
                timezone.now()
                if assignment.status == ConferencePaperReviewAssignment.Status.COMPLETED
                else None
            )
            assignment.save()
            messages.success(request, _("Peer review saved successfully."))
            return redirect("conferences:reviewer_workspace")
    else:
        form = ConferencePeerReviewForm(instance=assignment)
    return render(request, "conferences/peer_review.html", {
        "assignment": assignment,
        "paper": assignment.paper,
        "form": form,
    })


@login_required
@require_http_methods(["GET", "POST"])
def presentation_schedule(request, form_id, paper_id):
    _require_manager(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    paper = get_object_or_404(
        ConferencePaper,
        pk=paper_id,
        call__event=event_form.event,
        status=ConferencePaper.Status.ACCEPTED,
        is_active=True,
    )
    presentation = ConferencePresentation.objects.filter(paper=paper).first()
    if request.method == "POST":
        form = ConferencePresentationScheduleForm(
            request.POST,
            instance=presentation,
            event=event_form.event,
        )
        if form.is_valid():
            presentation = form.save(commit=False)
            presentation.paper = paper
            presentation.created_by = presentation.created_by or request.user
            presentation.updated_by = request.user
            presentation.save()
            messages.success(request, _("Presentation schedule saved successfully."))
            return redirect("conferences:presentation_list", form_id=event_form.pk)
    else:
        form = ConferencePresentationScheduleForm(
            instance=presentation,
            event=event_form.event,
            initial={"presenter_name": paper.corresponding_author},
        )
    return render(request, "conferences/presentation_schedule.html", {
        "event_form": event_form, "paper": paper, "form": form,
    })


@login_required
def presentation_list(request, form_id):
    _require_access(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    presentations = ConferencePresentation.objects.filter(
        paper__call__event=event_form.event,
        is_active=True,
    ).select_related("paper", "session", "programme_item")
    accepted_unscheduled = ConferencePaper.objects.filter(
        call__event=event_form.event,
        status=ConferencePaper.Status.ACCEPTED,
        is_active=True,
        presentation__isnull=True,
    )
    return render(request, "conferences/presentation_list.html", {
        "event_form": event_form,
        "presentations": presentations,
        "accepted_unscheduled": accepted_unscheduled,
    })


@require_http_methods(["GET", "POST"])
def presentation_confirm(request, public_token):
    paper = get_object_or_404(
        ConferencePaper.objects.select_related("presentation__session"),
        public_token=public_token,
        status=ConferencePaper.Status.ACCEPTED,
        is_active=True,
    )
    try:
        presentation = paper.presentation
    except ConferencePresentation.DoesNotExist:
        raise Http404("This paper has not yet been scheduled.")
    if request.method == "POST":
        form = ConferencePresentationConfirmationForm(
            request.POST, request.FILES, instance=presentation,
        )
        if form.is_valid():
            presentation = form.save(commit=False)
            presentation.status = (
                ConferencePresentation.Status.READY
                if presentation.slides
                else ConferencePresentation.Status.CONFIRMED
            )
            presentation.confirmed_at = timezone.now()
            presentation.save()
            return redirect("conferences:paper_status", public_token=paper.public_token)
    else:
        form = ConferencePresentationConfirmationForm(instance=presentation)
    return render(request, "conferences/presentation_confirm.html", {
        "paper": paper, "presentation": presentation, "form": form,
    })


@require_GET
def presentation_slides(request, public_token):
    paper = get_object_or_404(ConferencePaper, public_token=public_token, is_active=True)
    try:
        presentation = paper.presentation
    except ConferencePresentation.DoesNotExist:
        raise Http404
    if not presentation.slides:
        raise Http404
    return FileResponse(
        presentation.slides.open("rb"), as_attachment=True,
        filename=presentation.slides.name.rsplit("/", 1)[-1],
    )


def _communication_defaults(paper, request):
    status_url = request.build_absolute_uri(
        reverse("conferences:paper_status", kwargs={"public_token": paper.public_token})
    )
    if paper.status == ConferencePaper.Status.ACCEPTED:
        communication_type = ConferencePaperCommunication.CommunicationType.ACCEPTANCE
        subject = f"{paper.call.event.code}: Paper acceptance — {paper.reference_number}"
        decision = paper.decision_message or "Your submission has been accepted."
    elif paper.status == ConferencePaper.Status.REVISION_REQUIRED:
        communication_type = ConferencePaperCommunication.CommunicationType.REVISION
        subject = f"{paper.call.event.code}: Revision required — {paper.reference_number}"
        decision = paper.decision_message or "Please revise your submission as advised."
    elif paper.status == ConferencePaper.Status.REJECTED:
        communication_type = ConferencePaperCommunication.CommunicationType.REJECTION
        subject = f"{paper.call.event.code}: Submission decision — {paper.reference_number}"
        decision = paper.decision_message or "We regret that your submission was not accepted."
    else:
        communication_type = ConferencePaperCommunication.CommunicationType.ACKNOWLEDGEMENT
        subject = f"{paper.call.event.code}: Submission update — {paper.reference_number}"
        decision = "Your submission is currently being processed."
    schedule = ""
    try:
        presentation = paper.presentation
        schedule = (
            f"\n\nPresentation schedule:\n{presentation.session.title}\n"
            f"{presentation.starts_at:%d %B %Y, %H:%M}–{presentation.ends_at:%H:%M}\n"
            f"Venue: {presentation.venue_name}"
        )
        if paper.status == ConferencePaper.Status.ACCEPTED:
            communication_type = ConferencePaperCommunication.CommunicationType.PRESENTATION_INVITATION
            subject = f"{paper.call.event.code}: Presentation invitation — {paper.reference_number}"
    except ConferencePresentation.DoesNotExist:
        pass
    message = (
        f"Dear {paper.corresponding_author},\n\n"
        f"Reference: {paper.reference_number}\nPaper: {paper.title}\n\n"
        f"{decision}{schedule}\n\n"
        f"View your private submission record and respond here:\n{status_url}\n\n"
        "Ministry of Education, Science and Technology\n"
        f"{paper.call.event.title_en}"
    )
    return {
        "communication_type": communication_type,
        "recipient_email": paper.email,
        "subject": subject,
        "message": message,
    }


@login_required
@require_http_methods(["GET", "POST"])
def paper_communication(request, form_id, paper_id):
    _require_manager(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    paper = get_object_or_404(
        ConferencePaper.objects.select_related("call__event", "presentation__session"),
        pk=paper_id, call__event=event_form.event, is_active=True,
    )
    if request.method == "POST":
        form = ConferencePaperCommunicationForm(request.POST)
        if form.is_valid():
            communication = form.save(commit=False)
            communication.paper = paper
            communication.sent_by = request.user
            communication.created_by = request.user
            communication.updated_by = request.user
            try:
                send_mail(
                    communication.subject,
                    communication.message,
                    None,
                    [communication.recipient_email],
                    fail_silently=False,
                )
                communication.delivery_status = ConferencePaperCommunication.DeliveryStatus.SENT
                communication.sent_at = timezone.now()
                messages.success(request, _("Communication sent successfully."))
            except Exception as exc:
                communication.delivery_status = ConferencePaperCommunication.DeliveryStatus.FAILED
                communication.failure_message = str(exc)[:1000]
                messages.error(request, _("The email could not be sent. Its failure was logged."))
            communication.save()
            return redirect("conferences:paper_communication", form_id=event_form.pk, paper_id=paper.pk)
    else:
        form = ConferencePaperCommunicationForm(initial=_communication_defaults(paper, request))
    return render(request, "conferences/paper_communication.html", {
        "event_form": event_form,
        "paper": paper,
        "form": form,
        "communications": paper.communications.select_related("sent_by"),
    })


@require_GET
def paper_letter(request, public_token):
    paper = get_object_or_404(
        ConferencePaper.objects.select_related("call__event", "presentation__session"),
        public_token=public_token,
        is_active=True,
    )
    return render(request, "conferences/paper_letter.html", {"paper": paper})


@login_required
@require_http_methods(["GET", "POST"])
def certificate_list(request, form_id):
    _require_access(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    recipient_types = ConferenceCertificate.RecipientType
    eligible = {
        recipient_type: _eligible_certificate_sources(event_form, recipient_type).count()
        for recipient_type in recipient_types.values
    }
    if request.method == "POST":
        _require_manager(request.user)
        selected_type = request.POST.get("recipient_type", "")
        types_to_generate = list(recipient_types.values) if selected_type == "ALL" else [selected_type]
        if any(value not in recipient_types.values for value in types_to_generate):
            raise PermissionDenied
        created_count = 0
        with transaction.atomic():
            for recipient_type in types_to_generate:
                source_field = {
                    recipient_types.PARTICIPANT: "participant_submission",
                    recipient_types.PRESENTER: "paper",
                    recipient_types.REVIEWER: "reviewer",
                }[recipient_type]
                for source in _eligible_certificate_sources(event_form, recipient_type):
                    recipient_name, institution = _certificate_defaults(recipient_type, source)
                    certificate, created = ConferenceCertificate.objects.get_or_create(
                        event=event_form.event,
                        recipient_type=recipient_type,
                        **{source_field: source},
                        defaults={
                            "recipient_name": recipient_name,
                            "institution": institution,
                            "certificate_number": _certificate_number(
                                event_form.event, recipient_type, source.pk,
                            ),
                            "issued_by": request.user,
                            "created_by": request.user,
                            "updated_by": request.user,
                        },
                    )
                    created_count += int(created)
        messages.success(
            request,
            _("%(count)s new certificate(s) generated successfully.") % {"count": created_count},
        )
        return redirect("conferences:certificate_list", form_id=event_form.pk)
    certificates = ConferenceCertificate.objects.filter(
        event=event_form.event,
    ).select_related("issued_by").order_by("recipient_name", "recipient_type")
    return render(request, "conferences/certificate_list.html", {
        "event_form": event_form,
        "certificates": certificates,
        "eligible": eligible,
        "eligible_participants": eligible[recipient_types.PARTICIPANT],
        "eligible_presenters": eligible[recipient_types.PRESENTER],
        "eligible_reviewers": eligible[recipient_types.REVIEWER],
        "recipient_types": recipient_types,
        "can_manage": request.user.is_superuser
        or request.user.role in CONFERENCE_MANAGER_ROLES,
    })


@require_GET
def certificate_print(request, verification_token):
    certificate = get_object_or_404(
        ConferenceCertificate.objects.select_related("event", "event__venue", "paper"),
        verification_token=verification_token,
        is_active=True,
    )
    return render(request, "conferences/conference_certificate.html", {
        "certificate": certificate,
    })


@require_GET
def certificate_verify(request, verification_token):
    certificate = get_object_or_404(
        ConferenceCertificate.objects.select_related("event"),
        verification_token=verification_token,
        is_active=True,
    )
    return render(request, "conferences/conference_certificate_verify.html", {
        "certificate": certificate,
    })


@require_GET
def certificate_qr(request, verification_token):
    certificate = get_object_or_404(
        ConferenceCertificate,
        verification_token=verification_token,
        is_active=True,
    )
    verification_url = request.build_absolute_uri(reverse(
        "conferences:certificate_verify",
        kwargs={"verification_token": certificate.verification_token},
    ))
    return HttpResponse(generate_qr_png(verification_url), content_type="image/png")


@login_required
@require_POST
def certificate_revoke(request, form_id, certificate_id):
    _require_manager(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    certificate = get_object_or_404(
        ConferenceCertificate,
        pk=certificate_id,
        event=event_form.event,
        is_active=True,
    )
    reason = request.POST.get("reason", "").strip()
    if not reason:
        messages.error(request, _("Enter a reason before revoking the certificate."))
    else:
        certificate.is_revoked = True
        certificate.revocation_reason = reason
        certificate.updated_by = request.user
        certificate.save()
        messages.success(request, _("The certificate has been revoked."))
    return redirect("conferences:certificate_list", form_id=event_form.pk)


@login_required
@require_GET
def feedback_dashboard(request, form_id):
    _require_access(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    responses = ConferenceFeedback.objects.filter(
        event=event_form.event,
        is_active=True,
    ).select_related("session")
    session_filter = request.GET.get("session", "").strip()
    if session_filter.isdigit():
        responses = responses.filter(session_id=int(session_filter))
    aggregates = responses.aggregate(
        overall=Avg("overall_rating"),
        content=Avg("content_rating"),
        speakers=Avg("speakers_rating"),
        organization=Avg("organization_rating"),
        venue=Avg("venue_rating"),
    )
    response_count = responses.count()
    recommend_count = responses.filter(would_recommend=True).count()
    summary = {
        "total": response_count,
        "anonymous": responses.filter(is_anonymous=True).count(),
        "recommend_percent": round(recommend_count * 100 / response_count) if response_count else 0,
        **aggregates,
    }
    sessions = ConferenceSession.objects.filter(
        event=event_form.event,
        is_active=True,
    ).annotate(
        response_count=Count(
            "feedback_responses",
            filter=Q(feedback_responses__is_active=True),
        ),
        average_rating=Avg(
            "feedback_responses__overall_rating",
            filter=Q(feedback_responses__is_active=True),
        ),
    ).order_by("starts_at", "display_order")
    return render(request, "conferences/feedback_dashboard.html", {
        "event_form": event_form,
        "responses": responses,
        "sessions": sessions,
        "summary": summary,
        "session_filter": session_filter,
    })


@login_required
@require_GET
def feedback_csv(request, form_id):
    _require_access(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    responses = ConferenceFeedback.objects.filter(
        event=event_form.event,
        is_active=True,
    ).select_related("session").order_by("created_at")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = (
        f'attachment; filename="{event_form.event.code}-conference-feedback.csv"'
    )
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow((
        "Reference", "Submitted at", "Anonymous", "Name", "Institution", "Email",
        "Session", "Overall rating", "Content rating", "Speakers rating",
        "Organization rating", "Venue rating", "Would recommend",
        "Most valuable aspect", "Suggested improvements", "Additional comments",
    ))
    for item in responses:
        writer.writerow(tuple(safe_spreadsheet_value(value) for value in (
            item.reference_number,
            timezone.localtime(item.created_at).strftime("%Y-%m-%d %H:%M"),
            "Yes" if item.is_anonymous else "No",
            "" if item.is_anonymous else item.respondent_name,
            "" if item.is_anonymous else item.institution,
            "" if item.is_anonymous else item.email,
            item.session.title if item.session else "Overall conference",
            item.overall_rating,
            item.content_rating,
            item.speakers_rating,
            item.organization_rating,
            item.venue_rating,
            "Yes" if item.would_recommend else "No",
            item.most_valuable,
            item.improvements,
            item.additional_comments,
        )))
    return response


@login_required
def conference_list(request):
    _require_access(request.user)
    registrations = []
    for event_form in _conference_registration_forms(request.user):
        registrations.append({
            "form": event_form,
            "public_path": public_form_path(event_form, language="en"),
            "submission_count": event_form.submissions.filter(
                is_complete=True,
            ).count(),
        })
    return render(
        request,
        "conferences/conference_list.html",
        {"registrations": registrations},
    )


@login_required
def conference_detail(request, form_id):
    _require_access(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    submissions = list(
        event_form.submissions.filter(is_active=True, is_complete=True)
        .prefetch_related("answers__selected_options")
        .order_by("-created_at")
    )
    sessions = list(
        ConferenceSession.objects.filter(event=event_form.event, is_active=True)
        .prefetch_related("attendance_records")
        .order_by("starts_at", "display_order")
    )
    sessions_by_value = {
        session.registration_option_value: session for session in sessions
    }
    for submission in submissions:
        submission.selected_conference_sessions = [
            sessions_by_value[value]
            for value in _submission_session_values(submission)
            if value in sessions_by_value
        ]
    for session in sessions:
        session.registered_count = _selected_submissions(
            event_form,
            session,
        ).count()
        session.checked_in_count = session.attendance_records.filter(
            is_active=True,
        ).count()
    summary = {
        "total": len(submissions),
        "pending": sum(
            item.review_status == FormSubmission.ReviewStatus.PENDING
            for item in submissions
        ),
        "approved": sum(
            item.review_status == FormSubmission.ReviewStatus.APPROVED
            for item in submissions
        ),
        "rejected": sum(
            item.review_status == FormSubmission.ReviewStatus.REJECTED
            for item in submissions
        ),
    }
    return render(request, "conferences/conference_detail.html", {
        "event_form": event_form,
        "submissions": submissions,
        "sessions": sessions,
        "summary": summary,
        "can_manage": request.user.is_superuser
        or request.user.role in CONFERENCE_MANAGER_ROLES,
    })


def _participant_export_rows(event_form, session_id="", status="", query=""):
    sessions = list(
        ConferenceSession.objects.filter(event=event_form.event, is_active=True)
        .order_by("starts_at", "display_order", "id")
    )
    sessions_by_value = {
        session.registration_option_value: session for session in sessions
    }
    submissions = event_form.submissions.filter(is_active=True, is_complete=True)
    selected_session = None
    if str(session_id).isdigit():
        selected_session = next(
            (session for session in sessions if session.pk == int(session_id)),
            None,
        )
        if selected_session:
            submissions = submissions.filter(
                answers__selected_options__value=selected_session.registration_option_value
            )
    valid_statuses = {value for value, _label in FormSubmission.ReviewStatus.choices}
    if status in valid_statuses:
        submissions = submissions.filter(review_status=status)
    query = query.strip()
    if query:
        submissions = submissions.filter(
            Q(reference_number__icontains=query)
            | Q(badge_name__icontains=query)
            | Q(badge_organization__icontains=query)
            | Q(submitter_email__icontains=query)
            | Q(submitter_phone__icontains=query)
        )
    submissions = list(
        submissions.distinct().prefetch_related("answers__selected_options")
        .order_by("badge_name", "created_at", "id")
    )
    rows = []
    for number, submission in enumerate(submissions, start=1):
        selected_sessions = [
            sessions_by_value[value]
            for value in _submission_session_values(submission)
            if value in sessions_by_value
        ]
        selected_sessions.sort(key=lambda item: (item.starts_at, item.display_order))
        rows.append({
            "number": number,
            "submission": submission,
            "sessions": selected_sessions,
            "session_names": ", ".join(item.title for item in selected_sessions),
        })
    return {
        "rows": rows,
        "sessions": sessions,
        "selected_session": selected_session,
        "selected_status": status if status in valid_statuses else "",
        "query": query,
    }


@login_required
@require_GET
def participant_list_print(request, form_id):
    _require_access(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    export_data = _participant_export_rows(
        event_form,
        session_id=request.GET.get("session", ""),
        status=request.GET.get("status", ""),
        query=request.GET.get("q", ""),
    )
    return render(request, "conferences/participant_list_print.html", {
        "event_form": event_form,
        "event": event_form.event,
        **export_data,
        "status_choices": FormSubmission.ReviewStatus.choices,
        "generated_at": timezone.localtime(),
    })


@login_required
@require_GET
def participant_list_excel(request, form_id):
    _require_access(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    event = event_form.event
    export_data = _participant_export_rows(
        event_form,
        session_id=request.GET.get("session", ""),
        status=request.GET.get("status", ""),
        query=request.GET.get("q", ""),
    )
    rows = export_data["rows"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Registered participants"
    navy = "173B67"
    teal = "087F73"
    pale_blue = "EAF2F8"
    white = "FFFFFF"
    thin = Side(style="thin", color="C8D6E5")

    sheet.merge_cells("A1:H1")
    sheet["A1"] = event.title_en
    sheet["A1"].font = Font(size=16, bold=True, color=white)
    sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.merge_cells("A2:H2")
    sheet["A2"] = f"{event.code} — REGISTERED PARTICIPANTS"
    sheet["A2"].font = Font(size=12, bold=True, color=teal)
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A3:H3")
    date_range = f"{event.starts_at:%d %B %Y} – {event.ends_at:%d %B %Y}"
    venue = event.venue.name if event.venue else ""
    sheet["A3"] = f"{date_range}{' · ' + venue if venue else ''} · Total: {len(rows)}"
    sheet["A3"].alignment = Alignment(horizontal="center")
    filter_labels = []
    if export_data["selected_session"]:
        filter_labels.append(f"Session: {export_data['selected_session'].title}")
    if export_data["selected_status"]:
        filter_labels.append(f"Status: {export_data['selected_status'].title()}")
    if export_data["query"]:
        filter_labels.append(f"Search: {export_data['query']}")
    sheet.merge_cells("A4:H4")
    sheet["A4"] = " · ".join(filter_labels) if filter_labels else "All participants"
    sheet["A4"].font = Font(italic=True, color="526579")
    sheet["A4"].alignment = Alignment(horizontal="center")

    headers = (
        "No.", "Reference", "Full name", "Institution", "Email address",
        "Phone number", "Selected sessions", "Status",
    )
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=column, value=header)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=teal)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    for row_number, item in enumerate(rows, start=6):
        submission = item["submission"]
        values = (
            item["number"],
            safe_spreadsheet_value(submission.reference_number),
            safe_spreadsheet_value(submission.badge_display_name),
            safe_spreadsheet_value(submission.badge_organization),
            safe_spreadsheet_value(submission.submitter_email),
            safe_spreadsheet_value(submission.submitter_phone),
            safe_spreadsheet_value(item["session_names"]),
            submission.get_review_status_display(),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
            if row_number % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=pale_blue)

    widths = (6, 23, 24, 25, 28, 18, 45, 15)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + column)].width = width
    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A5:H{max(5, len(rows) + 5)}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:5"
    sheet.sheet_view.showGridLines = False

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{event.code}-registered-participants.xlsx"'
    )
    return response


@login_required
@require_POST
def registration_decision(request, form_id, submission_id, decision):
    _require_manager(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    submission = get_object_or_404(
        event_form.submissions,
        pk=submission_id,
        is_active=True,
        is_complete=True,
    )
    statuses = {
        "approve": FormSubmission.ReviewStatus.APPROVED,
        "reject": FormSubmission.ReviewStatus.REJECTED,
    }
    status = statuses.get(decision)
    if status is None:
        raise PermissionDenied
    reason = request.POST.get("reason", "").strip()
    if status == FormSubmission.ReviewStatus.REJECTED and not reason:
        messages.error(request, _("Enter a reason before rejecting the registration."))
        return redirect("conferences:conference_detail", form_id=event_form.pk)
    submission.review_status = status
    submission.reviewed_by = request.user
    submission.reviewed_at = timezone.now()
    submission.review_notes = reason if status == FormSubmission.ReviewStatus.REJECTED else ""
    submission.updated_by = request.user
    submission.save(update_fields=(
        "review_status",
        "reviewed_by",
        "reviewed_at",
        "review_notes",
        "updated_by",
        "updated_at",
    ))
    notification_type = (
        NotificationLog.NotificationType.REGISTRATION_APPROVED
        if status == FormSubmission.ReviewStatus.APPROVED
        else NotificationLog.NotificationType.REGISTRATION_REJECTED
    )
    send_submission_notification(submission, notification_type, request=request)
    return redirect("conferences:conference_detail", form_id=event_form.pk)


def _identifier_token(identifier):
    match = UUID_PATTERN.search(identifier)
    return match.group(0) if match else ""


@login_required
@require_http_methods(["GET", "POST"])
def session_register(request, form_id, session_id):
    _require_checkin(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    session = get_object_or_404(
        ConferenceSession,
        pk=session_id,
        event=event_form.event,
        is_active=True,
    )
    lookup_error = ""
    checked_submission = None
    if request.method == "POST":
        identifier = request.POST.get("identifier", "").strip()
        token = _identifier_token(identifier)
        candidates = FormSubmission.objects.filter(
            event_form=event_form,
            review_status=FormSubmission.ReviewStatus.APPROVED,
            is_active=True,
            is_complete=True,
        )
        checked_submission = candidates.filter(
            Q(reference_number__iexact=identifier)
            | Q(participant_token=token if token else None)
        ).first()
        if checked_submission is None:
            lookup_error = _("Participant not found or registration is not approved.")
        elif not _selected_submissions(
            event_form,
            session,
        ).filter(pk=checked_submission.pk).exists():
            lookup_error = _("This participant did not select this session.")
            checked_submission = None
        else:
            with transaction.atomic():
                attendance, created = ConferenceSessionAttendance.objects.get_or_create(
                    session=session,
                    submission=checked_submission,
                    defaults={
                        "checked_in_by": request.user,
                        "method": (
                            ConferenceSessionAttendance.Method.QR
                            if token
                            else ConferenceSessionAttendance.Method.MANUAL
                        ),
                        "created_by": request.user,
                        "updated_by": request.user,
                    },
                )
            if created:
                messages.success(
                    request,
                    _("%(name)s checked in successfully.")
                    % {"name": checked_submission.badge_display_name},
                )
            else:
                messages.info(
                    request,
                    _("This participant was already checked in to this session."),
                )
            return redirect(
                "conferences:session_register",
                form_id=event_form.pk,
                session_id=session.pk,
            )
    selected = list(
        _selected_submissions(event_form, session)
        .select_related("reviewed_by")
        .order_by("badge_name", "reference_number")
    )
    attendance_by_submission = {
        record.submission_id: record
        for record in session.attendance_records.filter(is_active=True)
        .select_related("checked_in_by")
    }
    for submission in selected:
        submission.session_attendance = attendance_by_submission.get(submission.pk)
    return render(request, "conferences/session_register.html", {
        "event_form": event_form,
        "session": session,
        "submissions": selected,
        "lookup_error": lookup_error,
        "checked_in_count": len(attendance_by_submission),
    })


@login_required
@require_GET
def session_register_csv(request, form_id, session_id):
    _require_access(request.user)
    event_form = get_object_or_404(_conference_registration_forms(request.user), pk=form_id)
    session = get_object_or_404(
        ConferenceSession,
        pk=session_id,
        event=event_form.event,
        is_active=True,
    )
    selected = _selected_submissions(event_form, session).order_by(
        "badge_name",
        "reference_number",
    )
    attendance_by_submission = {
        record.submission_id: record
        for record in session.attendance_records.filter(is_active=True)
        .select_related("checked_in_by")
    }
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response.write("\ufeff")
    response["Content-Disposition"] = (
        f'attachment; filename="{event_form.event.code}-{session.code}-attendance.csv"'
    )
    writer = csv.writer(response)
    writer.writerow([
        _("Reference number"),
        _("Participant"),
        _("Institution"),
        _("Checked in"),
        _("Checked in at"),
        _("Checked in by"),
    ])
    for submission in selected:
        attendance = attendance_by_submission.get(submission.pk)
        writer.writerow([
            safe_spreadsheet_value(submission.reference_number),
            safe_spreadsheet_value(submission.badge_display_name),
            safe_spreadsheet_value(submission.badge_organization),
            _("Yes") if attendance else _("No"),
            attendance.checked_in_at.isoformat() if attendance else "",
            safe_spreadsheet_value(str(attendance.checked_in_by) if attendance else ""),
        ])
    return response


@login_required
def registration_qr(request, form_id):
    _require_access(request.user)
    event_form = get_object_or_404(
        _conference_registration_forms(request.user),
        pk=form_id,
        is_published=True,
    )
    registration_url = public_form_url(
        event_form,
        request=request,
        language="en",
    )
    response = HttpResponse(
        generate_qr_png(registration_url),
        content_type="image/png",
    )
    if request.GET.get("download") == "1":
        response["Content-Disposition"] = (
            f'attachment; filename="{event_form.event.code}-registration-QR.png"'
        )
    response["X-Content-Type-Options"] = "nosniff"
    return response
